from datetime import datetime

from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import TemplateView
from django.db.models import Count, Avg
from django.views.decorators.http import require_GET

from .forms import StationForm
from .pump_forms import PumpForm
from .nozzle_forms import NozzleForm
from .models import Station, Pump, Nozzle
from inventory.models import FuelTank

from accounts.mixins import AdminMixin, OperationsRoleMixin
from accounts.station_access import (
    filter_nozzles_queryset_for_user,
    filter_pumps_queryset_for_user,
    filter_tanks_queryset_for_user,
    require_station_access,
    user_can_edit_station,
    visible_stations,
)


class StationListView(OperationsRoleMixin, TemplateView):
    template_name = "stations/list.html"
    extra_context = {"page_title": "Stations", "active_menu": "stations"}

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        qs = visible_stations(self.request.user).select_related("manager")
        search = self.request.GET.get("search", "").strip()
        if search:
            qs = qs.filter(name__icontains=search) | qs.filter(code__icontains=search)
        ctx["stations"] = qs
        ctx["search"] = search
        ctx["kpi_total"] = visible_stations(self.request.user).count()
        ctx["kpi_active"] = visible_stations(self.request.user).filter(is_active=True).count()
        return ctx

    def render_to_response(self, context, **response_kwargs):
        if self.request.GET.get("partial") == "list":
            return render(self.request, "stations/_list_content.html", context)
        return super().render_to_response(context, **response_kwargs)


class StationCreateView(AdminMixin, View):
    template_name = "stations/_modal_form.html"

    def get(self, request):
        form = StationForm()
        return render(request, self.template_name, {"form": form, "title": "Add Station"})

    def post(self, request):
        form = StationForm(request.POST)
        if form.is_valid():
            form.save()
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": True})
            messages.success(request, "Station created successfully.")
            return redirect("stations:list")
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors, "non_field_errors": form.non_field_errors()})
        return render(request, self.template_name, {"form": form, "title": "Add Station"})


class StationUpdateView(OperationsRoleMixin, View):
    template_name = "stations/_modal_form.html"

    def get_object(self, pk):
        station = get_object_or_404(visible_stations(self.request.user), pk=pk)
        if not user_can_edit_station(self.request.user, station):
            raise PermissionDenied("You cannot edit this station.")
        return station

    def get(self, request, pk):
        station = self.get_object(pk)
        form = StationForm(instance=station)
        return render(request, self.template_name, {"form": form, "title": "Edit Station"})

    def post(self, request, pk):
        station = self.get_object(pk)
        form = StationForm(request.POST, instance=station)
        if form.is_valid():
            form.save()
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": True})
            messages.success(request, "Station updated successfully.")
            return redirect("stations:list")
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors, "non_field_errors": form.non_field_errors()})
        return render(request, self.template_name, {"form": form, "title": "Edit Station"})


class StationDeleteView(AdminMixin, View):
    template_name = "stations/_confirm_delete.html"

    def get_object(self, pk):
        return get_object_or_404(Station, pk=pk)

    def get(self, request, pk):
        station = self.get_object(pk)
        return render(request, self.template_name, {"station": station})

    def post(self, request, pk):
        station = self.get_object(pk)
        station.delete()
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": True})
        messages.success(request, "Station deleted.")
        return redirect("stations:list")


class StationExportPDFView(AdminMixin, View):
    def get(self, request):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet

        buffer = []
        filename = f"stations_report_{datetime.now():%Y%m%d}.pdf"
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        doc = SimpleDocTemplate(response, pagesize=A4)
        styles = getSampleStyleSheet()

        data = [["Name", "Code", "Location", "Manager", "Status"]]
        for s in Station.objects.select_related("manager"):
            data.append([
                s.name,
                s.code,
                s.location,
                s.manager.full_name if s.manager else "",
                "Active" if s.is_active else "Inactive",
            ])
        table = Table(data, hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f7ea6")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
        ]))
        buffer.append(Paragraph("Stations Report", styles["Title"]))
        buffer.append(table)
        doc.build(buffer)
        return response


class StationExportExcelView(AdminMixin, View):
    def get(self, request):
        import openpyxl
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stations"
        headers = ["Name", "Code", "Location", "Manager", "Status"]
        ws.append(headers)
        for s in Station.objects.select_related("manager"):
            ws.append([
                s.name,
                s.code,
                s.location,
                s.manager.full_name if s.manager else "",
                "Active" if s.is_active else "Inactive",
            ])
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        filename = f"stations_report_{datetime.now():%Y%m%d}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
        wb.save(response)
        return response


# Pump CRUD


class PumpListView(OperationsRoleMixin, TemplateView):
    template_name = "stations/pumps/list.html"
    extra_context = {"page_title": "Pumps", "active_menu": "pumps"}

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from django.db.models import Prefetch

        qs = filter_pumps_queryset_for_user(
            Pump.objects.select_related("station", "tank").prefetch_related(
                Prefetch("nozzles", queryset=Nozzle.objects.select_related("tank").order_by("fuel_type")),
                "nozzles__tank",
            ),
            self.request.user,
        )
        search = self.request.GET.get("search", "").strip()
        if search:
            qs = qs.filter(label__icontains=search) | qs.filter(station__name__icontains=search)
        station_filter = self.request.GET.get("station")
        status_filter = self.request.GET.get("status")
        if station_filter:
            try:
                require_station_access(self.request.user, int(station_filter))
            except (PermissionDenied, TypeError, ValueError):
                qs = qs.none()
            else:
                qs = qs.filter(station_id=station_filter)
        if status_filter == "active":
            qs = qs.filter(is_active=True)
        elif status_filter == "inactive":
            qs = qs.filter(is_active=False)
        ctx["pumps"] = qs
        ctx["search"] = search
        ctx["stations"] = visible_stations(self.request.user)
        ctx["station_filter"] = station_filter or ""
        ctx["status_filter"] = status_filter or ""
        pump_scope = filter_pumps_queryset_for_user(Pump.objects.all(), self.request.user)
        ctx["kpi_total"] = pump_scope.count()
        ctx["kpi_active"] = pump_scope.filter(is_active=True).count()
        ctx["kpi_nozzles"] = filter_nozzles_queryset_for_user(Nozzle.objects.all(), self.request.user).count()
        ctx["kpi_avg_per_station"] = pump_scope.values("station").annotate(c=Count("id")).aggregate(avg=Avg("c"))[
            "avg"
        ] or 0
        return ctx

    def render_to_response(self, context, **response_kwargs):
        if self.request.GET.get("partial") == "list":
            return render(self.request, "stations/pumps/_list_content.html", context)
        return super().render_to_response(context, **response_kwargs)


class PumpCreateView(OperationsRoleMixin, View):
    template_name = "stations/pumps/_modal_form.html"

    def get(self, request):
        form = PumpForm()
        return render(request, self.template_name, {"form": form, "title": "Add Pump", "action": request.path})

    def post(self, request):
        form = PumpForm(request.POST)
        if form.is_valid():
            pump = form.save(commit=False)
            try:
                require_station_access(request.user, pump.station_id)
            except PermissionDenied:
                return JsonResponse(
                    {"success": False, "non_field_errors": ["You cannot add a pump at this station."]},
                    status=403,
                )
            pump.save()
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": True})
            messages.success(request, "Pump created successfully.")
            return redirect("stations:pumps")
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors, "non_field_errors": form.non_field_errors()})
        return render(request, self.template_name, {"form": form, "title": "Add Pump", "action": request.path})


class PumpUpdateView(OperationsRoleMixin, View):
    template_name = "stations/pumps/_modal_form.html"

    def get_object(self, pk):
        qs = filter_pumps_queryset_for_user(Pump.objects.all(), self.request.user)
        return get_object_or_404(qs, pk=pk)

    def get(self, request, pk):
        pump = self.get_object(pk)
        form = PumpForm(instance=pump)
        return render(request, self.template_name, {"form": form, "title": "Edit Pump", "action": request.path})

    def post(self, request, pk):
        pump = self.get_object(pk)
        form = PumpForm(request.POST, instance=pump)
        if form.is_valid():
            obj = form.save(commit=False)
            try:
                require_station_access(request.user, obj.station_id)
            except PermissionDenied:
                return JsonResponse(
                    {"success": False, "non_field_errors": ["You cannot move this pump to that station."]},
                    status=403,
                )
            obj.save()
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": True})
            messages.success(request, "Pump updated successfully.")
            return redirect("stations:pumps")
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors, "non_field_errors": form.non_field_errors()})
        return render(request, self.template_name, {"form": form, "title": "Edit Pump", "action": request.path})


class PumpDeleteView(OperationsRoleMixin, View):
    template_name = "stations/pumps/_confirm_delete.html"

    def get_object(self, pk):
        qs = filter_pumps_queryset_for_user(Pump.objects.all(), self.request.user)
        return get_object_or_404(qs, pk=pk)

    def get(self, request, pk):
        pump = self.get_object(pk)
        return render(request, self.template_name, {"pump": pump})

    def post(self, request, pk):
        pump = self.get_object(pk)
        pump.delete()
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": True})
        messages.success(request, "Pump deleted.")
        return redirect("stations:pumps")


# Nozzle CRUD (from pumps page)


class NozzleCreateView(OperationsRoleMixin, View):
    template_name = "stations/pumps/_nozzle_modal_form.html"

    def get(self, request):
        pump_id = request.GET.get("pump")
        form = NozzleForm(pump_id=pump_id)
        return render(request, self.template_name, {"form": form, "title": "Add Nozzle", "action": request.path})

    def post(self, request):
        pump_id = request.POST.get("pump")
        form = NozzleForm(request.POST, pump_id=pump_id)
        if form.is_valid():
            pump = form.cleaned_data.get("pump") or get_object_or_404(Pump, pk=pump_id)
            try:
                require_station_access(request.user, pump.station_id)
            except PermissionDenied:
                return JsonResponse(
                    {"success": False, "non_field_errors": ["You cannot add a nozzle to this pump."]},
                    status=403,
                )
            form.save()
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": True})
            messages.success(request, "Nozzle created successfully.")
            return redirect("stations:pumps")
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors, "non_field_errors": form.non_field_errors()})
        return render(request, self.template_name, {"form": form, "title": "Add Nozzle", "action": request.path})


class NozzleUpdateView(OperationsRoleMixin, View):
    template_name = "stations/pumps/_nozzle_modal_form.html"

    def get_object(self, pk):
        qs = filter_nozzles_queryset_for_user(Nozzle.objects.select_related("pump"), self.request.user)
        return get_object_or_404(qs, pk=pk)

    def get(self, request, pk):
        nozzle = self.get_object(pk)
        form = NozzleForm(instance=nozzle, pump_id=nozzle.pump_id)
        return render(request, self.template_name, {"form": form, "title": "Edit Nozzle", "action": request.path})

    def post(self, request, pk):
        nozzle = self.get_object(pk)
        form = NozzleForm(request.POST, instance=nozzle, pump_id=nozzle.pump_id)
        if form.is_valid():
            obj = form.save(commit=False)
            try:
                require_station_access(request.user, obj.pump.station_id)
            except PermissionDenied:
                return JsonResponse(
                    {"success": False, "non_field_errors": ["You cannot reassign this nozzle to that pump."]},
                    status=403,
                )
            obj.save()
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": True})
            messages.success(request, "Nozzle updated successfully.")
            return redirect("stations:pumps")
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors, "non_field_errors": form.non_field_errors()})
        return render(request, self.template_name, {"form": form, "title": "Edit Nozzle", "action": request.path})


class NozzleDeleteView(OperationsRoleMixin, View):
    template_name = "stations/pumps/_nozzle_confirm_delete.html"

    def get_object(self, pk):
        qs = filter_nozzles_queryset_for_user(Nozzle.objects.all(), self.request.user)
        return get_object_or_404(qs, pk=pk)

    def get(self, request, pk):
        nozzle = self.get_object(pk)
        return render(request, self.template_name, {"nozzle": nozzle})

    def post(self, request, pk):
        nozzle = self.get_object(pk)
        nozzle.delete()
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": True})
        messages.success(request, "Nozzle deleted.")
        return redirect("stations:pumps")


class TanksForPumpView(OperationsRoleMixin, View):
    def get(self, request):
        pump_id = request.GET.get("pump")
        fuel_type = request.GET.get("fuel_type")
        if not pump_id:
            return JsonResponse({"results": []})
        try:
            pump = filter_pumps_queryset_for_user(
                Pump.objects.select_related("station"),
                request.user,
            ).get(pk=pump_id)
        except Pump.DoesNotExist:
            return JsonResponse({"results": []})
        tanks = filter_tanks_queryset_for_user(
            FuelTank.objects.filter(station=pump.station),
            request.user,
        )
        if fuel_type:
            tanks = tanks.filter(fuel_type=fuel_type)
        data = [{"id": t.id, "name": f"{t.name} ({t.current_volume_liters} / {t.capacity_liters} L)"} for t in tanks]
        return JsonResponse({"results": data})
