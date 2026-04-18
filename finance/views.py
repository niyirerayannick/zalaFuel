from datetime import datetime
import csv
from decimal import Decimal

from django.conf import settings as django_settings
from django.views.generic import TemplateView, View
from django.db.models import DecimalField, ExpressionWrapper, F, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.core.exceptions import ValidationError

from accounts.mixins import FinanceRoleMixin
from accounts.station_access import (
    filter_delivery_receipts_queryset_for_user,
    filter_fuel_sales_queryset_for_user,
    filter_shifts_queryset_for_user,
    filter_tanks_queryset_for_user,
    require_station_access,
    visible_stations,
)
from django.core.exceptions import PermissionDenied
from sales.models import CreditPayment, CreditTransaction, Customer, FuelSale, ShiftSession
from sales.services import receive_credit_payment
from inventory.models import FuelTank
from stations.models import Pump
from suppliers.models import DeliveryReceipt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from accounts.models import SystemSettings
from .forms import CreditPaymentForm, CustomerForm


class FinanceDashboardView(FinanceRoleMixin, TemplateView):
    template_name = "finance/index.html"
    extra_context = {"page_title": "Finance", "active_menu": "finance"}

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        user = self.request.user
        sales_scope = filter_fuel_sales_queryset_for_user(FuelSale.objects.all(), user)
        total_revenue = sales_scope.aggregate(s=Sum("total_amount"))["s"] or 0
        exp_scope = filter_delivery_receipts_queryset_for_user(
            DeliveryReceipt.objects.filter(status=DeliveryReceipt.Status.RECEIVED).exclude(unit_cost__isnull=True),
            user,
        )
        total_expenses = exp_scope.annotate(
            line_total=ExpressionWrapper(
                F("delivered_volume") * F("unit_cost"),
                output_field=DecimalField(max_digits=14, decimal_places=2),
            )
        ).aggregate(s=Sum("line_total"))["s"] or 0
        recent_revenue = sales_scope.select_related("shift", "shift__station", "nozzle", "nozzle__pump").order_by(
            "-created_at"
        )[:8]
        recent_expenses = filter_delivery_receipts_queryset_for_user(
            DeliveryReceipt.objects.select_related("purchase_order", "purchase_order__supplier", "tank"),
            user,
        ).order_by("-delivery_date", "-created_at")[:8]
        ctx.update(
            {
                "total_revenue": total_revenue,
                "total_expenses": total_expenses,
                "gross_margin": total_revenue - total_expenses,
                "credit_balance": Customer.objects.aggregate(s=Sum("current_balance"))["s"] or 0,
                "recent_revenue": recent_revenue,
                "recent_expenses": recent_expenses,
            }
        )
        return ctx


class RevenueListView(FinanceRoleMixin, TemplateView):
    template_name = "finance/revenue.html"
    extra_context = {"page_title": "Revenue", "active_menu": "finance"}

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        user = self.request.user
        station_filter = self.request.GET.get("station", "").strip()
        payment_filter = self.request.GET.get("payment_method", "").strip()
        revenue = filter_fuel_sales_queryset_for_user(
            FuelSale.objects.select_related("shift", "shift__station", "nozzle", "nozzle__pump"),
            user,
        ).order_by("-created_at")
        if station_filter:
            try:
                require_station_access(user, int(station_filter))
            except (PermissionDenied, TypeError, ValueError):
                revenue = revenue.none()
            else:
                revenue = revenue.filter(shift__station_id=station_filter)
        if payment_filter:
            revenue = revenue.filter(payment_method=payment_filter)
        by_station = revenue.values("shift__station__name").annotate(total=Sum("total_amount")).order_by("-total")
        by_payment = revenue.values("payment_method").annotate(total=Sum("total_amount")).order_by("-total")
        ctx.update(
            {
                "revenue_rows": revenue,
                "stations": visible_stations(user),
                "payment_methods": FuelSale.PaymentMethod.choices,
                "filters": {"station": station_filter, "payment_method": payment_filter},
                "kpi_total_revenue": revenue.aggregate(s=Sum("total_amount"))["s"] or 0,
                "kpi_total_liters": revenue.aggregate(s=Sum("volume_liters"))["s"] or 0,
                "kpi_sales_count": revenue.count(),
                "station_breakdown": by_station[:5],
                "payment_breakdown": by_payment[:5],
            }
        )
        return ctx


class ExpenseListView(FinanceRoleMixin, TemplateView):
    template_name = "finance/expenses.html"
    extra_context = {"page_title": "Expenses", "active_menu": "finance"}

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        user = self.request.user
        station_filter = self.request.GET.get("station", "").strip()
        supplier_filter = self.request.GET.get("supplier", "").strip()
        status_filter = self.request.GET.get("status", "").strip()
        expenses = filter_delivery_receipts_queryset_for_user(
            DeliveryReceipt.objects.select_related(
                "purchase_order",
                "purchase_order__supplier",
                "purchase_order__station",
                "tank",
                "received_by",
            ),
            user,
        ).annotate(
            line_total=ExpressionWrapper(
                F("delivered_volume") * F("unit_cost"),
                output_field=DecimalField(max_digits=14, decimal_places=2),
            )
        ).order_by("-delivery_date", "-created_at")
        if station_filter:
            try:
                require_station_access(user, int(station_filter))
            except (PermissionDenied, TypeError, ValueError):
                expenses = expenses.none()
            else:
                expenses = expenses.filter(purchase_order__station_id=station_filter)
        if supplier_filter:
            expenses = expenses.filter(purchase_order__supplier_id=supplier_filter)
        if status_filter:
            expenses = expenses.filter(status=status_filter)
        received_expenses = expenses.filter(status=DeliveryReceipt.Status.RECEIVED).exclude(unit_cost__isnull=True)
        total_cost = received_expenses.aggregate(s=Sum(F("delivered_volume") * F("unit_cost")))["s"] or 0
        ctx.update(
            {
                "expense_rows": expenses,
                "stations": visible_stations(user),
                "suppliers": DeliveryReceipt.objects.values("purchase_order__supplier_id", "purchase_order__supplier__name").distinct().order_by("purchase_order__supplier__name"),
                "delivery_statuses": DeliveryReceipt.Status.choices,
                "filters": {"station": station_filter, "supplier": supplier_filter, "status": status_filter},
                "kpi_total_expenses": total_cost,
                "kpi_received_count": expenses.filter(status=DeliveryReceipt.Status.RECEIVED).count(),
                "kpi_pending_count": expenses.filter(status__in=[DeliveryReceipt.Status.DRAFT, DeliveryReceipt.Status.PENDING]).count(),
                "kpi_volume": expenses.filter(status=DeliveryReceipt.Status.RECEIVED).aggregate(s=Sum("delivered_volume"))["s"] or 0,
            }
        )
        return ctx


class CustomerListView(FinanceRoleMixin, TemplateView):
    template_name = "finance/customers.html"

    def get_queryset(self):
        search = self.request.GET.get("search", "").strip()
        credit_filter = self.request.GET.get("credit")

        qs = Customer.objects.all().order_by("name")
        if search:
            qs = qs.filter(Q(name__icontains=search) | Q(phone__icontains=search) | Q(email__icontains=search))
        if credit_filter == "allowed":
            qs = qs.filter(is_credit_allowed=True)
        elif credit_filter == "blocked":
            qs = qs.filter(is_credit_allowed=False)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        user = self.request.user
        qs = self.get_queryset()
        search = self.request.GET.get("search", "").strip()
        credit_filter = self.request.GET.get("credit")
        st = visible_stations(user)
        pump_qs = Pump.objects.filter(station_id__in=st.values_list("pk", flat=True)) if st.exists() else Pump.objects.none()
        customer_rows = list(qs)
        for customer in customer_rows:
            customer.available_credit = max(Decimal("0"), (customer.credit_limit or 0) - (customer.current_balance or 0))
            customer.recent_credit_transactions = list(
                customer.credit_transactions.select_related("sale").order_by("-created_at")[:5]
            )
            customer.recent_credit_payments = list(
                customer.credit_payments.select_related("received_by").order_by("-created_at")[:5]
            )
        ctx.update(
            {
                "page_title": "Customers",
                "active_menu": "finance",
                "customers": customer_rows,
                "filters": {"search": search, "credit": credit_filter or ""},
                "kpi_total": Customer.objects.count(),
                "kpi_credit": Customer.objects.filter(is_credit_allowed=True).count(),
                "kpi_balance": Customer.objects.aggregate(s=Sum("current_balance"))["s"] or 0,
                "kpi_avg_limit": Customer.objects.filter(is_credit_allowed=True).aggregate(s=Sum("credit_limit"))["s"] or 0,
                "kpi_available": Customer.objects.filter(is_credit_allowed=True).annotate(
                    available=F("credit_limit") - F("current_balance")
                ).aggregate(s=Sum("available"))["s"]
                or 0,
                "kpi_tanks": filter_tanks_queryset_for_user(FuelTank.objects.all(), user).count(),
                "kpi_shifts": filter_shifts_queryset_for_user(ShiftSession.objects.all(), user).count(),
                "kpi_pumps": pump_qs.count(),
            }
        )
        return ctx

    def render_to_response(self, context, **response_kwargs):
        if self.request.GET.get("partial") == "list":
            return render(self.request, "finance/_customers_list_content.html", context)
        return super().render_to_response(context, **response_kwargs)


class CustomerCreateView(FinanceRoleMixin, View):
    template_name = "finance/_customers_modal_form.html"

    def get(self, request):
        form = CustomerForm()
        return render(request, self.template_name, {"form": form, "title": "New Customer", "action": request.path})

    def post(self, request):
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save()
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": True})
            messages.success(request, "Customer created successfully.")
            return redirect("finance:customers")
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors, "non_field_errors": form.non_field_errors()}, status=400)
        return render(request, self.template_name, {"form": form, "title": "New Customer", "action": request.path}, status=400)


class CustomerUpdateView(FinanceRoleMixin, View):
    template_name = "finance/_customers_modal_form.html"

    def get_object(self, pk):
        return get_object_or_404(Customer, pk=pk)

    def get(self, request, pk):
        customer = self.get_object(pk)
        form = CustomerForm(instance=customer)
        return render(request, self.template_name, {"form": form, "title": "Edit Customer", "action": request.path})

    def post(self, request, pk):
        customer = self.get_object(pk)
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": True})
            messages.success(request, "Customer updated successfully.")
            return redirect("finance:customers")
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors, "non_field_errors": form.non_field_errors()}, status=400)
        return render(request, self.template_name, {"form": form, "title": "Edit Customer", "action": request.path}, status=400)


class CustomerDeleteView(FinanceRoleMixin, View):
    template_name = "finance/_customers_confirm_delete.html"

    def get_object(self, pk):
        return get_object_or_404(Customer, pk=pk)

    def get(self, request, pk):
        customer = self.get_object(pk)
        return render(request, self.template_name, {"customer": customer, "action": request.path})

    def post(self, request, pk):
        customer = self.get_object(pk)
        customer.delete()
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": True})
        messages.success(request, "Customer deleted successfully.")
        return redirect("finance:customers")


class CustomerCreditPaymentCreateView(FinanceRoleMixin, View):
    template_name = "finance/_credit_payment_modal_form.html"

    def get_object(self, pk):
        return get_object_or_404(Customer, pk=pk)

    def get(self, request, pk):
        customer = self.get_object(pk)
        form = CreditPaymentForm(customer=customer)
        return render(
            request,
            self.template_name,
            {
                "form": form,
                "customer": customer,
                "title": "Receive Credit Payment",
                "action": request.path,
            },
        )

    def post(self, request, pk):
        customer = self.get_object(pk)
        form = CreditPaymentForm(request.POST, customer=customer)
        if form.is_valid():
            try:
                receive_credit_payment(
                    customer=customer,
                    amount=form.cleaned_data["amount"],
                    method=form.cleaned_data["method"],
                    received_by=request.user,
                    reference=form.cleaned_data.get("reference", ""),
                    notes=form.cleaned_data.get("notes", ""),
                )
            except ValidationError as exc:
                errors = getattr(exc, "message_dict", None)
                if errors:
                    for field, messages_list in errors.items():
                        for message in messages_list:
                            form.add_error(field if field in form.fields else None, message)
                else:
                    form.add_error(None, "; ".join(exc.messages))
            else:
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return JsonResponse({"success": True})
                messages.success(request, "Credit payment recorded successfully.")
                return redirect("finance:customers")

        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {"success": False, "errors": form.errors, "non_field_errors": form.non_field_errors()},
                status=400,
            )
        return render(
            request,
            self.template_name,
            {"form": form, "customer": customer, "title": "Receive Credit Payment", "action": request.path},
            status=400,
        )


class CustomerExportCSVView(CustomerListView):
    def get(self, request, *args, **kwargs):
        qs = self.get_queryset()
        response = HttpResponse(content_type="text/csv")
        filename = f"customers_{datetime.now().strftime('%Y%m%d')}.csv"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)
        writer.writerow(["Name", "Phone", "Email", "Customer Type", "Credit Allowed", "Credit Limit", "Current Balance", "Address", "Notes"])
        for c in qs:
            writer.writerow(
                [
                    c.name,
                    c.phone,
                    c.email,
                    c.get_customer_type_display(),
                    "Yes" if c.is_credit_allowed else "No",
                    f"{c.credit_limit}",
                    f"{c.current_balance}",
                    c.address,
                    c.notes,
                ]
            )
        return response


class CustomerExportMixin(CustomerListView):
    def get_export_system_info(self):
        system_settings = SystemSettings.get_settings()
        company_name = (
            getattr(system_settings, "company_name", "")
            or getattr(django_settings, "BRAND_NAME", "ZALA/ECO ENERGY")
        )
        currency_code = getattr(system_settings, "currency", getattr(django_settings, "DEFAULT_CURRENCY", "USD"))
        currency_symbol = getattr(system_settings, "currency_symbol", currency_code) or currency_code
        exported_at = datetime.now().strftime("%Y-%m-%d %H:%M")
        return {
            "company_name": company_name,
            "currency_code": currency_code,
            "currency_symbol": currency_symbol,
            "exported_at": exported_at,
        }

    def get_export_rows(self):
        rows = []
        for customer in self.get_queryset():
            rows.append(
                [
                    customer.name,
                    customer.phone or "-",
                    customer.email or "-",
                    customer.get_customer_type_display(),
                    "Yes" if customer.is_credit_allowed else "No",
                    float(customer.credit_limit or 0),
                    float(customer.current_balance or 0),
                    float(max(Decimal("0"), (customer.credit_limit or 0) - (customer.current_balance or 0))),
                    customer.address or "-",
                    customer.notes or "-",
                ]
            )
        return rows


class CustomerExportExcelView(CustomerExportMixin):
    def get(self, request, *args, **kwargs):
        system_info = self.get_export_system_info()
        rows = self.get_export_rows()
        wb = Workbook()
        ws = wb.active
        ws.title = "Customers"
        ws.append([system_info["company_name"]])
        ws.append(["Customers Export"])
        ws.append([f"Currency: {system_info['currency_code']} ({system_info['currency_symbol']})"])
        ws.append([f"Exported at: {system_info['exported_at']}"])
        ws.append([])
        headers = [
            "Name",
            "Phone",
            "Email",
            "Customer Type",
            "Credit Allowed",
            "Credit Limit",
            "Current Balance",
            "Available Credit",
            "Address",
            "Notes",
        ]
        ws.append(headers)
        for row in rows:
            ws.append(row)

        ws["A1"].font = Font(size=16, bold=True)
        ws["A2"].font = Font(size=12, bold=True)
        for cell in ws[6]:
            cell.font = Font(bold=True)
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            ws.column_dimensions[col].width = 18
        ws["A1"].alignment = Alignment(horizontal="left")
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"customers_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class CustomerExportPDFView(CustomerExportMixin):
    def get(self, request, *args, **kwargs):
        system_info = self.get_export_system_info()
        rows = self.get_export_rows()
        response = HttpResponse(content_type="application/pdf")
        filename = f"customers_{datetime.now().strftime('%Y%m%d')}.pdf"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        doc = SimpleDocTemplate(response, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        story = [
            Paragraph(system_info["company_name"], styles["Title"]),
            Spacer(1, 6),
            Paragraph("Customers Export", styles["Heading2"]),
            Paragraph(
                f"Currency: {system_info['currency_code']} ({system_info['currency_symbol']}) | Exported at: {system_info['exported_at']}",
                styles["Normal"],
            ),
            Spacer(1, 12),
        ]

        table_data = [[
            "Name",
            "Phone",
            "Email",
            "Type",
            "Credit",
            "Limit",
            "Balance",
            "Available",
        ]]
        for row in rows:
            table_data.append(row[:8])

        table = Table(table_data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f7ea6")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (5, 1), (7, -1), "RIGHT"),
                ]
            )
        )
        story.append(table)
        doc.build(story)
        return response
