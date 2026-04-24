from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.db.models import Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse_lazy
from django.views.generic import CreateView, ListView, TemplateView, UpdateView
from django.shortcuts import get_object_or_404
from django.utils import timezone
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from accounts.currency import format_currency
from accounts.mixins import FinanceRoleMixin
from accounts.models import SystemSettings
from omcs.models import OMC
from products.models import Product
from terminals.models import Terminal

from .forms import OMCSalesEntryForm
from .models import OMCSalesEntry


def _logo_path():
    candidate = Path(__file__).resolve().parents[1] / "static" / "img" / "ZALA Terminal.png"
    return candidate if candidate.exists() else None


def _logo_stream(max_width=900):
    logo = _logo_path()
    if not logo:
        return None
    try:
        with PILImage.open(logo) as img:
            img = img.convert("RGBA")
            if img.width > max_width:
                ratio = max_width / float(img.width)
                img = img.resize((int(img.width * ratio), int(img.height * ratio)))
            buffer = BytesIO()
            img.save(buffer, format="PNG")
            buffer.seek(0)
            return buffer
    except Exception:
        return None


class OMCSalesQuerysetMixin:
    def get_base_queryset(self):
        return OMCSalesEntry.objects.select_related("terminal", "omc", "product", "submitted_by").order_by(
            "-sale_date", "-created_at"
        )

    def get_filtered_queryset(self):
        search = (self.request.GET.get("search") or "").strip()
        omc_filter = (self.request.GET.get("omc") or "").strip()
        product_filter = (self.request.GET.get("product") or "").strip()
        terminal_filter = (self.request.GET.get("terminal") or "").strip()
        date_from = (self.request.GET.get("date_from") or "").strip()
        date_to = (self.request.GET.get("date_to") or "").strip()

        entries = self.get_base_queryset()

        if search:
            entries = entries.filter(submission_reference__icontains=search) | entries.filter(
                omc__name__icontains=search
            )
        if omc_filter:
            entries = entries.filter(omc_id=omc_filter)
        if product_filter:
            entries = entries.filter(product_id=product_filter)
        if terminal_filter:
            entries = entries.filter(terminal_id=terminal_filter)
        if date_from:
            entries = entries.filter(sale_date__gte=date_from)
        if date_to:
            entries = entries.filter(sale_date__lte=date_to)

        return entries.order_by("-sale_date", "-created_at")

    def get_filter_context(self):
        return {
            "omcs": OMC.objects.order_by("name"),
            "products": Product.objects.order_by("product_name"),
            "terminals": Terminal.objects.order_by("name"),
            "filters": {
                "search": self.request.GET.get("search", ""),
                "omc": self.request.GET.get("omc", ""),
                "product": self.request.GET.get("product", ""),
                "terminal": self.request.GET.get("terminal", ""),
                "date_from": self.request.GET.get("date_from", ""),
                "date_to": self.request.GET.get("date_to", ""),
            },
        }

    def get_system_branding(self):
        settings_obj = SystemSettings.get_settings()
        currency_code = getattr(settings_obj, "currency", getattr(settings, "DEFAULT_CURRENCY", "USD"))
        currency_symbol = getattr(settings_obj, "currency_symbol", "") or currency_code
        company_name = getattr(settings_obj, "company_name", None) or getattr(settings, "BRAND_NAME", "ZALA Terminal")
        return settings_obj, company_name, currency_code, currency_symbol


class OMCSalesDashboardView(OMCSalesQuerysetMixin, FinanceRoleMixin, ListView):
    model = OMCSalesEntry
    template_name = "sales/dashboard.html"
    context_object_name = "entries"
    paginate_by = 25

    def get_queryset(self):
        return self.get_filtered_queryset()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_entries = self.get_filtered_queryset()
        _settings_obj, company_name, currency_code, currency_symbol = self.get_system_branding()

        total_volume = all_entries.aggregate(total=Sum("volume_liters"))["total"] or 0
        total_revenue = all_entries.aggregate(total=Sum("total_amount"))["total"] or 0
        active_omcs = all_entries.values("omc").distinct().count()

        top_by_volume = all_entries.order_by("-volume_liters").first()
        top_by_revenue = all_entries.order_by("-total_amount").first()

        context.update(
            {
                "page_title": "OMC Sales",
                "active_menu": "omc_sales",
                "kpi_total_volume": total_volume,
                "kpi_total_revenue": total_revenue,
                "kpi_active_omcs": active_omcs,
                "kpi_total_records": all_entries.count(),
                "kpi_top_volume_omc": top_by_volume.omc.name if top_by_volume else "-",
                "kpi_top_revenue_omc": top_by_revenue.omc.name if top_by_revenue else "-",
                "company_name": company_name,
                "currency_code": currency_code,
                "currency_symbol": currency_symbol,
                "average_price": (total_revenue / total_volume) if total_volume else 0,
                "current_query": self.request.GET.urlencode(),
                **self.get_filter_context(),
            }
        )
        return context

    def render_to_response(self, context, **response_kwargs):
        if self.request.GET.get("partial") == "list":
            return render(self.request, "sales/_list_content.html", context)
        return super().render_to_response(context, **response_kwargs)


class OMCSalesExportMixin(OMCSalesQuerysetMixin, FinanceRoleMixin, TemplateView):
    def get_filtered_queryset(self):
        return super().get_filtered_queryset()


class OMCSalesEntryExcelExportView(OMCSalesExportMixin):
    def get(self, request, *args, **kwargs):
        queryset = self.get_filtered_queryset()
        _settings_obj, company_name, currency_code, currency_symbol = self.get_system_branding()
        filename = "omc_sales_report.xlsx"
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError:
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="omc_sales_report.csv"'
            response.write("Sale Date,Reference,Terminal,OMC,Product,Volume (L),Unit Price,Total Amount,Submitted By,Remarks\n")
            for row in queryset:
                response.write(
                    f'"{row.sale_date:%Y-%m-%d}","{row.submission_reference}","{row.terminal.name}","{row.omc.name}","{row.product.product_name}",'
                    f'"{row.volume_liters}","{row.unit_price}","{row.total_amount}","{getattr(row.submitted_by, "full_name", "")}","{(row.remarks or "").replace(chr(10), " ")}"\n'
                )
            return response

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "OMC Sales"
        header_fill = PatternFill(fill_type="solid", fgColor="0F5B2A")
        header_font = Font(color="FFFFFF", bold=True)
        title_font = Font(color="0F5B2A", bold=True, size=16)
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )
        sheet.merge_cells("A1:J1")
        sheet["A1"] = f"{company_name} OMC Sales Report"
        sheet["A1"].font = title_font
        sheet.merge_cells("A2:J2")
        sheet["A2"] = (
            f"Generated on {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')} | "
            f"Currency: {currency_code} ({currency_symbol}) | Records: {queryset.count()}"
        )
        sheet["A2"].font = Font(color="475569", italic=True, size=10)

        headers = [
            "Sale Date",
            "Reference",
            "Terminal",
            "OMC",
            "Product",
            "Volume (L)",
            f"Unit Price ({currency_code})",
            f"Total Amount ({currency_code})",
            "Submitted By",
            "Remarks",
        ]
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=4, column=idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        row_no = 5
        for entry in queryset:
            values = [
                entry.sale_date.strftime("%d/%m/%Y"),
                entry.submission_reference,
                entry.terminal.name,
                entry.omc.name,
                entry.product.product_name,
                float(entry.volume_liters or 0),
                float(entry.unit_price or 0),
                float(entry.total_amount or 0),
                getattr(entry.submitted_by, "full_name", "-") if entry.submitted_by else "-",
                entry.remarks or "",
            ]
            for col_no, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_no, column=col_no, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top")
                if col_no in (6, 7, 8):
                    cell.number_format = '#,##0.00'
            row_no += 1

        total_row = row_no + 1
        sheet.cell(row=total_row, column=5, value="Totals").font = Font(bold=True, color="0F5B2A")
        sheet.cell(row=total_row, column=6, value=float(sum((e.volume_liters or 0) for e in queryset))).number_format = '#,##0.00'
        sheet.cell(row=total_row, column=8, value=float(sum((e.total_amount or 0) for e in queryset))).number_format = '#,##0.00'
        for col in (5, 6, 8):
            sheet.cell(row=total_row, column=col).border = thin_border

        sheet.auto_filter.ref = f"A4:J{max(row_no - 1, 4)}"
        sheet.freeze_panes = "A5"
        for column_cells in sheet.columns:
            column_index = column_cells[0].column
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 3, 30)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type=mime_type)
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class OMCSalesEntryPdfExportView(OMCSalesExportMixin):
    def get(self, request, *args, **kwargs):
        queryset = self.get_filtered_queryset()
        _settings_obj, company_name, currency_code, currency_symbol = self.get_system_branding()
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=12 * mm,
            leftMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
        )
        styles = getSampleStyleSheet()
        logo_stream = _logo_stream()

        header_left = []
        if logo_stream:
            header_left.append(Image(logo_stream, width=34 * mm, height=16 * mm))
            header_left.append(Spacer(1, 2 * mm))
        header_left.extend(
            [
                Paragraph(f"<font color='#0F5B2A'><b>{company_name} OMC Sales Report</b></font>", styles["Title"]),
                Paragraph("Commercial sales register export generated from ZALA Terminal.", styles["Normal"]),
            ]
        )
        total_volume = sum((entry.volume_liters or 0) for entry in queryset)
        total_amount = sum((entry.total_amount or 0) for entry in queryset)
        header_right = [
            Paragraph("<b>Report</b><br/>OMC Sales Register", styles["Normal"]),
            Spacer(1, 2),
            Paragraph(f"<b>Generated</b><br/>{timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}", styles["Normal"]),
            Spacer(1, 2),
            Paragraph(f"<b>Currency</b><br/>{currency_code} ({currency_symbol})", styles["Normal"]),
            Spacer(1, 2),
            Paragraph(f"<b>Total Revenue</b><br/>{format_currency(total_amount, currency_code)}", styles["Normal"]),
        ]
        header_table = Table([[header_left, header_right]], colWidths=[175 * mm, 80 * mm])
        header_table.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1E7D7")),
                    ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#EAF7EF")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        rows = [[
            "Date",
            "Reference",
            "Terminal",
            "OMC",
            "Product",
            "Volume (L)",
            f"Unit Price ({currency_code})",
            f"Revenue ({currency_code})",
        ]]
        for entry in queryset:
            rows.append(
                [
                    entry.sale_date.strftime("%d/%m/%Y"),
                    entry.submission_reference,
                    entry.terminal.name,
                    entry.omc.name,
                    entry.product.product_name,
                    f"{entry.volume_liters:,.2f}",
                    format_currency(entry.unit_price, currency_code),
                    format_currency(entry.total_amount, currency_code),
                ]
            )
        if len(rows) == 1:
            rows.append(["-", "-", "-", "-", "-", "-", "-", "-"])
        rows.append(["", "", "", "", "Totals", f"{total_volume:,.2f}", "", format_currency(total_amount, currency_code)])
        table = Table(
            rows,
            colWidths=[22 * mm, 34 * mm, 34 * mm, 30 * mm, 34 * mm, 24 * mm, 34 * mm, 34 * mm],
            repeatRows=1,
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F5B2A")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1D5DB")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F8FAFC")]),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#ECFDF5")),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        doc.build([header_table, Spacer(1, 5 * mm), table])
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="omc_sales_report.pdf"'
        return response


class OMCSalesEntryCreateView(FinanceRoleMixin, CreateView):
    model = OMCSalesEntry
    form_class = OMCSalesEntryForm
    template_name = "sales/form.html"
    success_url = reverse_lazy("sales:dashboard")

    def get_template_names(self):
        if self.request.GET.get("partial") == "form":
            return ["sales/_modal_form.html"]
        return [self.template_name]

    def form_valid(self, form):
        form.instance.submitted_by = self.request.user
        self.object = form.save()
        if self.request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": True})
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        if self.request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors}, status=400)
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Add Sales Entry"
        context["action"] = self.request.path
        return context


class OMCSalesEntryUpdateView(FinanceRoleMixin, UpdateView):
    model = OMCSalesEntry
    form_class = OMCSalesEntryForm
    template_name = "sales/form.html"
    success_url = reverse_lazy("sales:dashboard")

    def get_template_names(self):
        if self.request.GET.get("partial") == "form":
            return ["sales/_modal_form.html"]
        return [self.template_name]

    def form_valid(self, form):
        self.object = form.save()
        if self.request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": True})
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        if self.request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors}, status=400)
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Edit Sales Entry"
        context["action"] = self.request.path
        return context
