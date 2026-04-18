import logging
import warnings
from io import BytesIO
from pathlib import Path

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
from django.core import signing
from django.core.exceptions import PermissionDenied, ValidationError
from django.db import transaction
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.views import View
from django.views.generic import CreateView, DetailView, ListView, TemplateView, UpdateView
from django.db.models import Q, Count
from django.utils.crypto import get_random_string
from django.utils import timezone
from datetime import datetime, timedelta
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from django.conf import settings as app_settings
try:
    from transport.messaging.twilio_client import send_sms_message
except Exception:  # transport app removed in fuel edition
    def send_sms_message(*args, **kwargs):
        return None
from .emailing import build_public_url, send_atms_email

from .forms import (
    ForgotPasswordRequestForm,
    LoginForm,
    LoginVerificationCodeForm,
    PasswordResetCodeForm,
    PasswordResetNewPasswordForm,
    PasswordChangeForm,
    UserCreationForm,
    UserProfileForm,
    UserUpdateForm,
    SystemSettingsForm,
)
from .mixins import AdminMixin, ClientMixin, DriverMixin, ManagerMixin, SuperAdminMixin
from .models import User, UserProfile, ActivityLog, LoginVerification, PasswordResetVerification, RolePermission, SystemSettings
from .rbac import SYSTEM_ROLE_CHOICES, SystemGroup, current_system_role, sync_user_to_system_role, user_has_role
from stations.models import Station
from django.contrib.auth.decorators import login_required as login_required_fn

logger = logging.getLogger(__name__)


PASSWORD_RESET_VERIFICATION_ID_SESSION_KEY = "password_reset_verification_id"
PASSWORD_RESET_VERIFIED_USER_SESSION_KEY = "password_reset_verified_user_id"
PASSWORD_RESET_FALLBACK_CODE_SESSION_KEY = "password_reset_fallback_code"
PASSWORD_RESET_EMAIL_WARNING_SESSION_KEY = "password_reset_email_warning"
LOGIN_VERIFICATION_ID_SESSION_KEY = "login_verification_id"
LOGIN_PENDING_USER_ID_SESSION_KEY = "login_pending_user_id"
LOGIN_PENDING_BACKEND_SESSION_KEY = "login_pending_backend"
LOGIN_FALLBACK_CODE_SESSION_KEY = "login_fallback_code"
LOGIN_EMAIL_WARNING_SESSION_KEY = "login_email_warning"
LOGIN_AUTHENTICATED_AT_SESSION_KEY = "session_authenticated_at"
INTERNAL_STAFF_ROLE_VALUES = (
    User.Role.SUPERADMIN,
    User.Role.ADMIN,
    User.Role.STATION_MANAGER,
    User.Role.SUPERVISOR,
    User.Role.PUMP_ATTENDANT,
    User.Role.ACCOUNTANT,
)
INTERNAL_STAFF_SYSTEM_ROLES = (
    SystemGroup.ADMIN,
    SystemGroup.STATION_MANAGER,
    SystemGroup.SUPERVISOR,
    SystemGroup.PUMP_ATTENDANT,
    SystemGroup.ACCOUNTANT,
)
STATION_MANAGER_MANAGED_ROLES = (
    SystemGroup.SUPERVISOR,
    SystemGroup.PUMP_ATTENDANT,
    SystemGroup.ACCOUNTANT,
)


def internal_staff_role_choices_for(user):
    if user_has_role(user, SystemGroup.ADMIN):
        return [(value, label) for value, label in SYSTEM_ROLE_CHOICES if value in INTERNAL_STAFF_SYSTEM_ROLES]
    return [(value, label) for value, label in SYSTEM_ROLE_CHOICES if value in STATION_MANAGER_MANAGED_ROLES]


def manageable_stations_for(user):
    stations = Station.objects.filter(is_active=True)
    if user_has_role(user, SystemGroup.ADMIN):
        return stations.order_by("name")
    station_ids = []
    if getattr(user, "assigned_station_id", None):
        station_ids.append(user.assigned_station_id)
    managed_ids = list(user.managed_stations.filter(is_active=True).values_list("pk", flat=True))
    station_ids.extend(managed_ids)
    return stations.filter(pk__in=station_ids).distinct().order_by("name")


class UserManagementAccessMixin(LoginRequiredMixin):
    """Staff user management with station scoping for station managers."""

    def dispatch(self, request, *args, **kwargs):
        if not (
            user_has_role(request.user, SystemGroup.ADMIN)
            or user_has_role(request.user, SystemGroup.STATION_MANAGER)
        ):
            raise PermissionDenied("You do not have permission to manage users.")
        return super().dispatch(request, *args, **kwargs)

    def get_station_queryset(self):
        return manageable_stations_for(self.request.user)

    def get_role_choices(self):
        return internal_staff_role_choices_for(self.request.user)

    def get_user_queryset(self):
        qs = User.objects.filter(role__in=INTERNAL_STAFF_ROLE_VALUES).select_related("assigned_station", "profile").prefetch_related("groups")
        if user_has_role(self.request.user, SystemGroup.ADMIN):
            return qs
        station_ids = list(self.get_station_queryset().values_list("pk", flat=True))
        return qs.filter(assigned_station_id__in=station_ids, role__in=(User.Role.SUPERVISOR, User.Role.PUMP_ATTENDANT, User.Role.ACCOUNTANT))

    def can_manage_user(self, user):
        if user_has_role(self.request.user, SystemGroup.ADMIN):
            return True
        if user.pk == self.request.user.pk:
            return False
        allowed_stations = set(self.get_station_queryset().values_list("pk", flat=True))
        return (
            user.assigned_station_id in allowed_stations
            and current_system_role(user) in STATION_MANAGER_MANAGED_ROLES
        )

    def get_object(self, queryset=None):
        user = get_object_or_404(self.get_user_queryset(), pk=self.kwargs["pk"])
        if not self.can_manage_user(user) and not user_has_role(self.request.user, SystemGroup.ADMIN):
            raise PermissionDenied("You cannot manage this user.")
        return user


def _logo_path():
    candidate = Path(__file__).resolve().parents[1] / "static" / "img" / "ZALA/ECO ENERGY.png"
    return candidate if candidate.exists() else None


def _logo_stream(max_width=900):
    logo = _logo_path()
    if not logo:
        return None

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", PILImage.DecompressionBombWarning)
        img = PILImage.open(logo)
        img.load()
    with img:
        img.thumbnail((max_width, max_width))
        stream = BytesIO()
        img.save(stream, format="PNG", optimize=True)
        stream.seek(0)
        return stream


def client_ip_from_request(request):
    x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    if x_forwarded_for:
        return x_forwarded_for.split(",")[0]
    return request.META.get("REMOTE_ADDR")


def user_agent_from_request(request):
    return (request.META.get("HTTP_USER_AGENT") or "").strip()


def approximate_location_from_request(request, ip_address):
    forwarded_country = (
        request.META.get("HTTP_CF_IPCOUNTRY")
        or request.META.get("HTTP_X_COUNTRY_CODE")
        or request.META.get("HTTP_X_APPENGINE_COUNTRY")
        or ""
    ).strip()
    if forwarded_country:
        return forwarded_country
    if not ip_address:
        return "Unknown location"
    if ip_address.startswith("127.") or ip_address == "::1":
        return "Localhost / same device"
    if ip_address.startswith("10.") or ip_address.startswith("192.168.") or ip_address.startswith("172."):
        return "Private network"
    return "Approximate location unavailable"


def mask_email(email):
    if not email or "@" not in email:
        return email
    local_part, domain = email.split("@", 1)
    if len(local_part) <= 2:
        masked_local = local_part[0] + "*" * max(len(local_part) - 1, 0)
    else:
        masked_local = f"{local_part[0]}{'*' * max(len(local_part) - 2, 1)}{local_part[-1]}"
    return f"{masked_local}@{domain}"


def mask_phone(phone):
    clean = "".join(ch for ch in (phone or "") if ch.isdigit() or ch == "+")
    if not clean:
        return ""
    if len(clean) <= 4:
        return clean[:1] + "*" * max(len(clean) - 1, 0)
    return f"{clean[:3]}{'*' * max(len(clean) - 5, 1)}{clean[-2:]}"


def clear_login_verification_session(request):
    request.session.pop(LOGIN_VERIFICATION_ID_SESSION_KEY, None)
    request.session.pop(LOGIN_PENDING_USER_ID_SESSION_KEY, None)
    request.session.pop(LOGIN_PENDING_BACKEND_SESSION_KEY, None)
    request.session.pop(LOGIN_FALLBACK_CODE_SESSION_KEY, None)
    request.session.pop(LOGIN_EMAIL_WARNING_SESSION_KEY, None)


def build_account_security_token(user):
    return signing.dumps({"user_id": str(user.pk)}, salt="accounts.secure-account")


def log_user_activity(request, user, action, description):
    ActivityLog.objects.create(
        user=user,
        action=action,
        description=description,
        ip_address=client_ip_from_request(request),
        user_agent=user_agent_from_request(request),
    )


def send_login_security_alert(user, *, ip_address, user_agent, happened_at, location, secure_account_url):
    readable_time = timezone.localtime(happened_at).strftime("%d/%m/%Y %H:%M")
    device_text = user_agent or "Unknown device"
    sms_message = (
        f"ZALA/ECO ENERGY security alert: new login detected at {readable_time} from IP "
        f"{ip_address or 'unknown'}. If this was not you, change your password now."
    )

    try:
        send_atms_email(
            subject="ZALA/ECO ENERGY security alert: new login detected",
            to=[user.email],
            greeting=f"Hello {user.full_name}",
            headline="Security Alert",
            intro="ZALA/ECO ENERGY noticed a sign-in to your account from a new device or location.",
            details=[
                {"label": "Time", "value": readable_time},
                {"label": "Location", "value": location},
                {"label": "IP Address", "value": ip_address or "Unknown"},
                {"label": "Device", "value": device_text},
            ],
            note="If this was not you, change your password immediately and contact ZALA/ECO ENERGY support.",
            cta_label="Secure Account Now",
            cta_url=secure_account_url,
        )
    except Exception:
        pass

    if user.phone:
        send_sms_message(user.phone, sms_message)


class PasswordResetRequestView(View):
    template_name = "accounts/password_reset.html"
    form_class = ForgotPasswordRequestForm

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, {"form": self.form_class()})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if not form.is_valid():
            return render(request, self.template_name, {"form": form}, status=400)

        user = User.objects.filter(email=form.cleaned_data["email"], is_active=True).first()
        if not user:
            messages.success(request, "If that email exists in ZALA/ECO ENERGY, a recovery message has been sent.")
            return redirect("accounts:password_reset")

        verification = PasswordResetVerification.issue_for_user(
            user,
            ip_address=client_ip_from_request(request),
        )
        request.session[PASSWORD_RESET_VERIFICATION_ID_SESSION_KEY] = str(verification.pk)
        request.session.pop(PASSWORD_RESET_VERIFIED_USER_SESSION_KEY, None)
        request.session.pop(PASSWORD_RESET_FALLBACK_CODE_SESSION_KEY, None)
        request.session.pop(PASSWORD_RESET_EMAIL_WARNING_SESSION_KEY, None)
        try:
            send_atms_email(
                subject="ZALA/ECO ENERGY password reset verification code",
                to=[user.email],
                greeting=f"Hello {user.full_name}",
                headline="Password Reset Verification",
                intro="Use the one-time verification code below to continue resetting your ZALA/ECO ENERGY password.",
                details=[
                    {"label": "Verification Code", "value": verification.raw_code},
                    {"label": "Validity", "value": "10 minutes"},
                ],
                note="If you did not request this change, ignore this message and keep your account credentials private.",
            )
        except Exception:
            request.session[PASSWORD_RESET_FALLBACK_CODE_SESSION_KEY] = verification.raw_code
            request.session[PASSWORD_RESET_EMAIL_WARNING_SESSION_KEY] = (
                "We could not send the verification email from this environment. "
                "Use the one-time code shown on the next step."
            )
        return redirect("accounts:password_reset_done")


class PasswordResetVerifyView(View):
    template_name = "accounts/password_reset_done.html"
    form_class = PasswordResetCodeForm

    def _get_verification(self, request):
        verification_id = request.session.get(PASSWORD_RESET_VERIFICATION_ID_SESSION_KEY)
        if not verification_id:
            return None
        return PasswordResetVerification.objects.select_related("user").filter(pk=verification_id).first()

    def get(self, request, *args, **kwargs):
        verification = self._get_verification(request)
        if not verification:
            messages.info(request, "Start a password reset request first.")
            return redirect("accounts:password_reset")
        if verification.is_expired or verification.is_used:
            messages.error(request, "That verification code has expired. Request a new one.")
            return redirect("accounts:password_reset")
        context = {
            "form": self.form_class(),
            "masked_email": mask_email(verification.user.email),
            "fallback_code": request.session.get(PASSWORD_RESET_FALLBACK_CODE_SESSION_KEY),
            "email_warning": request.session.get(PASSWORD_RESET_EMAIL_WARNING_SESSION_KEY),
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        verification = self._get_verification(request)
        if not verification:
            messages.info(request, "Start a password reset request first.")
            return redirect("accounts:password_reset")

        form = self.form_class(request.POST)
        context = {
            "form": form,
            "masked_email": mask_email(verification.user.email),
            "fallback_code": request.session.get(PASSWORD_RESET_FALLBACK_CODE_SESSION_KEY),
            "email_warning": request.session.get(PASSWORD_RESET_EMAIL_WARNING_SESSION_KEY),
        }
        if not form.is_valid():
            return render(request, self.template_name, context, status=400)

        if verification.is_expired or verification.is_used:
            form.add_error("code", "This verification code has expired. Request a new one.")
            return render(request, self.template_name, context, status=400)

        if verification.attempt_count >= 5:
            verification.used_at = timezone.now()
            verification.save(update_fields=["used_at"])
            form.add_error("code", "Too many attempts. Request a new verification code.")
            return render(request, self.template_name, context, status=400)

        if not verification.check_code(form.cleaned_data["code"]):
            form.add_error("code", "The verification code is invalid.")
            return render(request, self.template_name, context, status=400)

        request.session[PASSWORD_RESET_VERIFIED_USER_SESSION_KEY] = str(verification.user_id)
        request.session.pop(PASSWORD_RESET_FALLBACK_CODE_SESSION_KEY, None)
        request.session.pop(PASSWORD_RESET_EMAIL_WARNING_SESSION_KEY, None)
        return redirect("accounts:password_reset_confirm")


class PasswordResetNewPasswordView(View):
    template_name = "accounts/password_reset_confirm.html"
    form_class = PasswordResetNewPasswordForm

    def _get_user(self, request):
        user_id = request.session.get(PASSWORD_RESET_VERIFIED_USER_SESSION_KEY)
        if not user_id:
            return None
        return User.objects.filter(pk=user_id, is_active=True).first()

    def _get_verification(self, request):
        verification_id = request.session.get(PASSWORD_RESET_VERIFICATION_ID_SESSION_KEY)
        if not verification_id:
            return None
        return PasswordResetVerification.objects.filter(pk=verification_id).first()

    def get(self, request, *args, **kwargs):
        user = self._get_user(request)
        if not user:
            messages.info(request, "Verify your recovery code first.")
            return redirect("accounts:password_reset")
        return render(request, self.template_name, {"form": self.form_class(user)})

    def post(self, request, *args, **kwargs):
        user = self._get_user(request)
        verification = self._get_verification(request)
        if not user or not verification:
            messages.info(request, "Verify your recovery code first.")
            return redirect("accounts:password_reset")

        form = self.form_class(user, request.POST)
        if not form.is_valid():
            return render(request, self.template_name, {"form": form}, status=400)

        form.save()
        user.must_change_password = False
        user.save(update_fields=["must_change_password"])
        verification.used_at = timezone.now()
        verification.save(update_fields=["used_at"])
        request.session.pop(PASSWORD_RESET_VERIFIED_USER_SESSION_KEY, None)
        request.session.pop(PASSWORD_RESET_VERIFICATION_ID_SESSION_KEY, None)
        messages.success(request, "Your password has been updated successfully.")
        return redirect("accounts:password_reset_complete")


class PasswordResetCompleteStepView(TemplateView):
    template_name = "accounts/password_reset_complete.html"


class SecureAccountView(View):
    def get(self, request, *args, **kwargs):
        token = request.GET.get("token", "")
        try:
            payload = signing.loads(token, salt="accounts.secure-account", max_age=60 * 60 * 24)
            user = User.objects.get(pk=payload.get("user_id"), is_active=True)
        except Exception:
            messages.error(request, "This security action link is invalid or expired.")
            return redirect("accounts:login")

        user.session_invalid_before = timezone.now()
        user.must_change_password = True
        user.save(update_fields=["session_invalid_before", "must_change_password"])
        if request.user.is_authenticated and request.user.pk == user.pk:
            logout(request)
        messages.success(
            request,
            "Your account sessions were secured. Please sign in again and change your password immediately.",
        )
        return redirect("accounts:login")


class ForcePasswordChangeView(LoginRequiredMixin, View):
    template_name = "accounts/password_change_required.html"
    form_class = PasswordChangeForm

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, {"form": self.form_class(request.user)})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.user, request.POST)
        if not form.is_valid():
            return render(request, self.template_name, {"form": form}, status=400)

        form.save()
        request.user.must_change_password = False
        request.user.save(update_fields=["must_change_password"])
        update_session_auth_hash(request, request.user)
        log_user_activity(
            request,
            request.user,
            ActivityLog.ActionType.UPDATE,
            "User changed password after first login requirement.",
        )
        messages.success(request, "Your password has been updated successfully.")
        return redirect(dashboard_name_for_user(request.user))


def dashboard_name_for_user(user):
    if user_has_role(user, SystemGroup.CUSTOMER):
        return "finance:customers"
    if user_has_role(user, SystemGroup.ACCOUNTANT):
        return "finance:dashboard"
    if user_has_role(user, SystemGroup.STATION_MANAGER):
        return "core:dashboard"
    if user_has_role(user, SystemGroup.SUPERVISOR):
        return "core:dashboard"
    if user_has_role(user, SystemGroup.PUMP_ATTENDANT):
        return "sales:shifts"
    if user_has_role(user, SystemGroup.ADMIN):
        return "core:dashboard"
    return "core:dashboard"


class LoginVerificationView(View):
    template_name = "accounts/login_verify.html"
    form_class = LoginVerificationCodeForm

    def _get_verification(self, request):
        verification_id = request.session.get(LOGIN_VERIFICATION_ID_SESSION_KEY)
        if not verification_id:
            return None
        return LoginVerification.objects.select_related("user").filter(pk=verification_id).first()

    def _get_user(self, request):
        user_id = request.session.get(LOGIN_PENDING_USER_ID_SESSION_KEY)
        if not user_id:
            return None
        return User.objects.filter(pk=user_id, is_active=True).first()

    def _context(self, request, form, verification):
        user = verification.user if verification and verification.user_id else None
        masked_phone = mask_phone(getattr(user, "phone", ""))
        masked_email = mask_email(getattr(user, "email", ""))
        return {
            "form": form,
            "masked_phone": masked_phone,
            "masked_email": masked_email,
            "fallback_code": request.session.get(LOGIN_FALLBACK_CODE_SESSION_KEY),
            "delivery_warning": request.session.get(LOGIN_EMAIL_WARNING_SESSION_KEY),
        }

    def get(self, request, *args, **kwargs):
        verification = self._get_verification(request)
        user = self._get_user(request)
        if not verification or not user:
            messages.info(request, "Sign in first to continue with verification.")
            return redirect("accounts:login")
        if verification.is_expired or verification.is_used:
            clear_login_verification_session(request)
            messages.error(request, "Your login verification code expired. Sign in again.")
            return redirect("accounts:login")
        return render(request, self.template_name, self._context(request, self.form_class(), verification))

    def post(self, request, *args, **kwargs):
        verification = self._get_verification(request)
        user = self._get_user(request)
        if not verification or not user:
            messages.info(request, "Sign in first to continue with verification.")
            return redirect("accounts:login")

        form = self.form_class(request.POST)
        context = self._context(request, form, verification)
        if not form.is_valid():
            return render(request, self.template_name, context, status=400)
        if verification.is_expired or verification.is_used:
            clear_login_verification_session(request)
            form.add_error("code", "The verification code has expired. Sign in again.")
            return render(request, self.template_name, context, status=400)
        if verification.attempt_count >= 5:
            verification.used_at = timezone.now()
            verification.save(update_fields=["used_at"])
            clear_login_verification_session(request)
            form.add_error("code", "Too many attempts. Sign in again to receive a new code.")
            return render(request, self.template_name, context, status=400)
        if not verification.check_code(form.cleaned_data["code"]):
            form.add_error("code", "The verification code is invalid.")
            return render(request, self.template_name, context, status=400)

        verification.used_at = timezone.now()
        verification.save(update_fields=["used_at"])
        backend = request.session.get(LOGIN_PENDING_BACKEND_SESSION_KEY) or app_settings.AUTHENTICATION_BACKENDS[0]
        ip_address = client_ip_from_request(request)
        user_agent = user_agent_from_request(request)
        location = approximate_location_from_request(request, ip_address)
        previous_login = (
            ActivityLog.objects.filter(user=user, action=ActivityLog.ActionType.LOGIN)
            .order_by("-created_at")
            .first()
        )
        clear_login_verification_session(request)
        login(request, user, backend=backend)
        request.session[LOGIN_AUTHENTICATED_AT_SESSION_KEY] = timezone.now().isoformat()
        log_user_activity(
            request,
            user,
            ActivityLog.ActionType.LOGIN,
            f"User signed in to ZALA/ECO ENERGY dashboard from {ip_address or 'unknown IP'}",
        )
        if previous_login and (
            (previous_login.ip_address or "") != (ip_address or "")
            or (previous_login.user_agent or "") != (user_agent or "")
        ):
            send_login_security_alert(
                user,
                ip_address=ip_address,
                user_agent=user_agent,
                happened_at=timezone.now(),
                location=location,
                secure_account_url=build_public_url(
                    f"/secure-account/?token={build_account_security_token(user)}"
                ),
            )
        return redirect(dashboard_name_for_user(user))


class LoginView(View):
    form_class = LoginForm
    template_name = "accounts/login.html"

    def get(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            return redirect("accounts:dashboard")
        clear_login_verification_session(request)
        return render(request, self.template_name, {"form": self.form_class()})

    def user_allowed_for_login(self, user):
        return True, None

    def begin_login_verification(self, request, user):
        verification = LoginVerification.issue_for_user(
            user,
            ip_address=client_ip_from_request(request),
        )
        request.session[LOGIN_VERIFICATION_ID_SESSION_KEY] = str(verification.pk)
        request.session[LOGIN_PENDING_USER_ID_SESSION_KEY] = str(user.pk)
        request.session[LOGIN_PENDING_BACKEND_SESSION_KEY] = getattr(user, "backend", "") or app_settings.AUTHENTICATION_BACKENDS[0]
        request.session.pop(LOGIN_FALLBACK_CODE_SESSION_KEY, None)
        request.session.pop(LOGIN_EMAIL_WARNING_SESSION_KEY, None)
        sms_message = (
            f"ZALA/ECO ENERGY verification code: {verification.raw_code}. "
            "It expires in 10 minutes. Do not share this code."
        )
        email_sent = False
        sms_sent = False
        delivery_warning = ""

        if user.phone:
            sms_sent = bool(send_sms_message(user.phone, sms_message))

        try:
            send_atms_email(
                subject="ZALA/ECO ENERGY login verification code",
                to=[user.email],
                greeting=f"Hello {user.full_name}",
                headline="Login Verification",
                intro="Use the verification code below to complete your ZALA/ECO ENERGY sign-in securely.",
                details=[
                    {"label": "Verification Code", "value": verification.raw_code},
                    {"label": "Validity", "value": "10 minutes"},
                ],
                note="If this was not you, change your password immediately.",
            )
            email_sent = True
        except Exception:
            if sms_sent:
                delivery_warning = (
                    "We could not send the login code by email, but it was sent to your phone."
                )

        if sms_sent and email_sent:
            return

        if not sms_sent and not email_sent:
            request.session[LOGIN_FALLBACK_CODE_SESSION_KEY] = verification.raw_code
            delivery_warning = "Use the one-time code shown on the next step."

        if delivery_warning:
            request.session[LOGIN_EMAIL_WARNING_SESSION_KEY] = delivery_warning

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if not form.is_valid():
            return render(request, self.template_name, {"form": form}, status=400)

        email = form.cleaned_data["email"].lower()
        password = form.cleaned_data["password"]
        user = authenticate(request, email=email, password=password)

        if user is None:
            form.add_error(None, "Invalid email or password.")
            return render(request, self.template_name, {"form": form}, status=401)

        if not user.is_active:
            form.add_error(None, "Your account is inactive.")
            return render(request, self.template_name, {"form": form}, status=403)

        is_allowed, error_message = self.user_allowed_for_login(user)
        if not is_allowed:
            form.add_error(None, error_message or "You do not have access to this login.")
            return render(request, self.template_name, {"form": form}, status=403)

        clear_login_verification_session(request)
        self.begin_login_verification(request, user)
        return redirect("accounts:login-verify")


class CustomerLoginView(LoginView):
    template_name = "accounts/login.html"

    def user_allowed_for_login(self, user):
        if not user_has_role(user, SystemGroup.CUSTOMER):
            return False, "Use the staff login for internal accounts."
        return True, None


class LogoutView(LoginRequiredMixin, View):
    template_name = "accounts/logout_confirm.html"

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        log_user_activity(request, request.user, ActivityLog.ActionType.LOGOUT, "User logged out of ZALA/ECO ENERGY")
        logout(request)
        messages.success(request, "You have been logged out.")
        return redirect("accounts:login")


class DashboardView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        return redirect(dashboard_name_for_user(request.user))


class SuperAdminDashboardView(SuperAdminMixin, View):
    def get(self, request, *args, **kwargs):
        return redirect("core:dashboard")


class AdminDashboardView(AdminMixin, View):
    def get(self, request, *args, **kwargs):
        return redirect("core:dashboard")


class ManagerDashboardView(ManagerMixin, View):
    def get(self, request, *args, **kwargs):
        return redirect("core:dashboard")


class DriverDashboardView(DriverMixin, View):
    def get(self, request, *args, **kwargs):
        return redirect("sales:shifts")


class ClientDashboardView(ClientMixin, View):
    def get(self, request, *args, **kwargs):
        return redirect("finance:customers")


class ProfileView(LoginRequiredMixin, TemplateView):
    template_name = "accounts/profile.html"

    def _get_or_create_profile(self, user):
        profile = User.objects.filter(pk=user.pk).values_list("profile__pk", flat=True).first()
        if profile:
            return user.profile
        return UserProfile.objects.get_or_create(user=user)[0]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        profile = self._get_or_create_profile(self.request.user)
        context["profile_user"] = self.request.user
        context["profile_form"] = UserProfileForm(instance=profile)
        context["user_form"] = UserUpdateForm(instance=self.request.user)
        return context

    def post(self, request, *args, **kwargs):
        profile = self._get_or_create_profile(request.user)
        profile_form = UserProfileForm(request.POST, instance=profile)

        user = request.user
        user.full_name = request.POST.get("full_name", user.full_name)
        user.phone = request.POST.get("phone", user.phone)
        new_profile_photo = request.FILES.get("profile_photo")

        if profile_form.is_valid():
            try:
                user.full_clean(exclude=["password", "profile_photo"])
            except ValidationError as exc:
                for field, errors in exc.message_dict.items():
                    target_field = field if field in profile_form.fields else None
                    for error in errors:
                        profile_form.add_error(target_field, error)
            else:
                try:
                    with transaction.atomic():
                        user.save(update_fields=["full_name", "phone"])
                        updated_profile = profile_form.save(commit=False)
                        updated_profile.user = request.user
                        updated_profile.save()
                except Exception:
                    logger.exception("Profile details update failed for user %s", request.user.pk)
                    profile_form.add_error(None, "We could not update your profile right now. Please try again.")
                else:
                    if new_profile_photo:
                        try:
                            user.profile_photo = new_profile_photo
                            user.save(update_fields=["profile_photo"])
                        except Exception as exc:
                            logger.exception("Profile photo upload failed for user %s", request.user.pk)
                            messages.warning(
                                request,
                                f"Your profile details were saved, but the photo upload failed: {exc}",
                            )
                            return redirect("accounts:profile")

                    messages.success(request, "Profile updated successfully.")
                    return redirect("accounts:profile")

        messages.error(request, "Please correct the errors below.")
        return render(
            request,
            self.template_name,
            {
                "profile_user": request.user,
                "profile_form": profile_form,
                "user_form": UserUpdateForm(instance=request.user),
            },
            status=400,
        )

    from django.views.generic import TemplateView

    # Custom password reset confirmation view
    class CustomPasswordResetDoneView(TemplateView):
        template_name = "accounts/password_reset_done.html"



class UserListView(UserManagementAccessMixin, ListView):
    model = User
    template_name = "accounts/user_list.html"
    context_object_name = "users"
    paginate_by = 25

    def get_queryset(self):
        search = (self.request.GET.get("search") or "").strip().lower()
        status = (self.request.GET.get("status") or "").strip().lower()
        role_filter = (self.request.GET.get("role") or "").strip()
        station_filter = (self.request.GET.get("station") or "").strip()

        users = list(self.get_user_queryset().order_by("-created_at"))
        for user in users:
            user.system_role = current_system_role(user)

        if search:
            users = [
                user for user in users
                if search in (user.full_name or "").lower()
                or search in (user.email or "").lower()
                or search in (user.phone or "").lower()
                or search in (user.staff_id or "").lower()
                or search in (getattr(user.assigned_station, "name", "") or "").lower()
            ]
        if status == "active":
            users = [user for user in users if user.is_active]
        elif status == "inactive":
            users = [user for user in users if not user.is_active]
        if role_filter:
            users = [user for user in users if user.system_role == role_filter]
        if station_filter:
            users = [user for user in users if str(user.assigned_station_id or "") == station_filter]

        return users

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_users = list(self.get_user_queryset().order_by("-created_at"))
        for user in all_users:
            user.system_role = current_system_role(user)

        context.update(
            {
                "total_users": len(all_users),
                "active_users": sum(1 for user in all_users if user.is_active),
                "inactive_users": sum(1 for user in all_users if not user.is_active),
                "admin_users": sum(1 for user in all_users if user.system_role == SystemGroup.ADMIN),
                "station_users": sum(1 for user in all_users if user.assigned_station_id),
                "role_choices": self.get_role_choices(),
                "stations": self.get_station_queryset(),
                "filters": {
                    "search": self.request.GET.get("search", ""),
                    "status": self.request.GET.get("status", ""),
                    "role": self.request.GET.get("role", ""),
                    "station": self.request.GET.get("station", ""),
                },
            }
        )
        return context

    def render_to_response(self, context, **response_kwargs):
        if self.request.GET.get("partial") == "list":
            return render(self.request, "accounts/users/_list_content.html", context)
        return super().render_to_response(context, **response_kwargs)


class UserExportMixin(UserManagementAccessMixin):
    def get_filtered_queryset(self):
        view = UserListView()
        view.request = self.request
        return view.get_queryset()


class UserExcelExportView(UserExportMixin, ListView):
    def get(self, request, *args, **kwargs):
        users = self.get_filtered_queryset()
        filename = "users_report.xlsx"
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError:
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="users_report.csv"'
            response.write("Name,Email,Phone,Staff ID,Role,Station,Status,Last Login,Created\n")
            for user in users:
                last_login = timezone.localtime(user.last_login).strftime("%Y-%m-%d %H:%M") if user.last_login else "Never"
                response.write(
                    f'"{user.full_name or ""}","{user.email or ""}","{user.phone or ""}","{user.staff_id or ""}","{current_system_role(user)}","{getattr(user.assigned_station, "name", "") or ""}","{"Active" if user.is_active else "Inactive"}","{last_login}","{user.created_at:%Y-%m-%d}"\n'
                )
            return response

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Users"
        header_fill = PatternFill(fill_type="solid", fgColor="0C5069")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"), top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"))
        sheet.merge_cells("A1:I1")
        sheet["A1"] = "ZALA/ECO ENERGY Users Report"
        sheet["A1"].font = Font(color="0C5069", bold=True, size=16)
        sheet.merge_cells("A2:I2")
        sheet["A2"] = f"Generated on {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
        sheet["A2"].font = Font(color="475569", italic=True, size=10)
        headers = ["Name", "Email", "Phone", "Staff ID", "Role", "Station", "Status", "Last Login", "Created"]
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=4, column=idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        row_no = 5
        for user in users:
            values = [
                user.full_name or "-",
                user.email or "-",
                user.phone or "-",
                user.staff_id or "-",
                getattr(user, "system_role", current_system_role(user)),
                getattr(user.assigned_station, "name", None) or "-",
                "Active" if user.is_active else "Inactive",
                timezone.localtime(user.last_login).strftime("%d/%m/%Y %H:%M") if user.last_login else "Never",
                user.created_at.strftime("%d/%m/%Y"),
            ]
            for col_no, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_no, column=col_no, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top")
            row_no += 1
        sheet.auto_filter.ref = f"A4:I{max(row_no - 1, 4)}"
        sheet.freeze_panes = "A5"
        for column_cells in sheet.columns:
            column_index = column_cells[0].column
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 3, 28)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type=mime_type)
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class UserPdfExportView(UserExportMixin, ListView):
    def get(self, request, *args, **kwargs):
        users = self.get_filtered_queryset()
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
        styles = getSampleStyleSheet()
        logo_stream = _logo_stream()
        header_left = []
        if logo_stream:
            header_left.append(Image(logo_stream, width=34 * mm, height=16 * mm))
            header_left.append(Spacer(1, 2 * mm))
        header_left.extend([Paragraph("<font color='#0C5069'><b>ZALA/ECO ENERGY Users Report</b></font>", styles["Title"]), Paragraph("Internal fuel station staff directory export.", styles["Normal"])])
        header_right = [Paragraph("<b>Report</b><br/>Users Register", styles["Normal"]), Spacer(1, 2), Paragraph(f"<b>Generated</b><br/>{timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}", styles["Normal"]), Spacer(1, 2), Paragraph(f"<b>Total Users</b><br/>{len(users)}", styles["Normal"])]
        header_table = Table([[header_left, header_right]], colWidths=[170 * mm, 85 * mm])
        header_table.setStyle(TableStyle([("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1E7D7")), ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#EAF7EF")), ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10), ("TOPPADDING", (0, 0), (-1, -1), 10), ("BOTTOMPADDING", (0, 0), (-1, -1), 10), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
        rows = [["Name", "Email", "Phone", "Staff ID", "Role", "Station", "Status", "Last Login", "Created"]]
        for user in users:
            rows.append([
                user.full_name or "-",
                user.email or "-",
                user.phone or "-",
                user.staff_id or "-",
                getattr(user, "system_role", current_system_role(user)),
                getattr(user.assigned_station, "name", None) or "-",
                "Active" if user.is_active else "Inactive",
                timezone.localtime(user.last_login).strftime("%d/%m/%Y %H:%M") if user.last_login else "Never",
                user.created_at.strftime("%d/%m/%Y"),
            ])
        if len(rows) == 1:
            rows.append(["-", "-", "-", "-", "-", "-", "-", "-", "-"])
        table = Table(rows, colWidths=[34 * mm, 45 * mm, 25 * mm, 22 * mm, 29 * mm, 32 * mm, 20 * mm, 28 * mm, 22 * mm], repeatRows=1)
        table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0C5069")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 7), ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1D5DB")), ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]), ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4), ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
        doc.build([header_table, Spacer(1, 5 * mm), table])
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="users_report.pdf"'
        return response


class UserCreateView(UserManagementAccessMixin, CreateView):
    model = User
    form_class = UserCreationForm
    template_name = "accounts/user_create.html"
    success_url = reverse_lazy("accounts:user-list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["request_user"] = self.request.user
        kwargs["station_queryset"] = self.get_station_queryset()
        kwargs["role_choices"] = self.get_role_choices()
        return kwargs

    def get_template_names(self):
        if self.request.GET.get("partial") == "form":
            return ["accounts/users/_modal_form.html"]
        return [self.template_name]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Add User"
        context["action"] = self.request.path
        return context

    def form_valid(self, form):
        self.object = form.save()
        raw_password = form.cleaned_data.get("password1", "")
        user = self.object
        selected_role = getattr(user, "_selected_system_role", current_system_role(user))

        try:
            send_atms_email(
                subject="Your ZALA/ECO ENERGY account has been created",
                to=[user.email],
                greeting=f"Hello {user.full_name}",
                headline="Account Created Successfully",
                intro="Your user account has been created in the ZALA/ECO ENERGY. Use the credentials below to sign in.",
                details=[
                    {"label": "Email", "value": user.email},
                    {"label": "Temporary Password", "value": raw_password},
                    {"label": "Role", "value": selected_role},
                    {"label": "Phone", "value": user.phone or "Not provided"},
                    {"label": "Login URL", "value": build_public_url("/")},
                ],
                note="For security, sign in as soon as possible and change your password after your first login.",
                cta_label="Open Login",
                cta_url=build_public_url("/"),
            )
            messages.success(self.request, f'User "{user.full_name}" created and credentials sent to {user.email}.')
        except Exception as exc:
            messages.warning(
                self.request,
                f'User "{user.full_name}" was created, but email could not be sent: {exc}. Password: {raw_password}'
            )

        if self.request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": True})
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        if self.request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors, "non_field_errors": form.non_field_errors()}, status=400)
        return super().form_invalid(form)


class UserDetailView(UserManagementAccessMixin, DetailView):
    model = User
    template_name = "accounts/user_detail.html"
    context_object_name = "managed_user"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.object
        user.system_role = current_system_role(user)
        context["recent_activity"] = user.activity_logs.order_by("-created_at")[:10]
        context["can_change_lifecycle"] = user.pk != self.request.user.pk
        return context


class UserUpdateView(UserManagementAccessMixin, UpdateView):
    model = User
    form_class = UserUpdateForm
    template_name = "accounts/user_edit.html"
    success_url = reverse_lazy("accounts:user-list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["request_user"] = self.request.user
        kwargs["station_queryset"] = self.get_station_queryset()
        kwargs["role_choices"] = self.get_role_choices()
        return kwargs

    def get_template_names(self):
        if self.request.GET.get("partial") == "form":
            return ["accounts/users/_modal_form.html"]
        return [self.template_name]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Edit User"
        context["action"] = self.request.path
        return context

    def form_valid(self, form):
        self.object = form.save()
        if self.request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": True})
        messages.success(self.request, "User updated successfully.")
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        if self.request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors, "non_field_errors": form.non_field_errors()}, status=400)
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse("accounts:user-detail", kwargs={"pk": self.object.pk})


class UserToggleActiveView(UserManagementAccessMixin, View):
    def post(self, request, *args, **kwargs):
        user = self.get_object()
        if user == request.user:
            messages.error(request, "You cannot deactivate your own account.")
            return redirect("accounts:user-detail", pk=user.pk)
        user.is_active = not user.is_active
        if not user.is_active:
            user.session_invalid_before = timezone.now()
        user.save(update_fields=["is_active", "session_invalid_before"])
        log_user_activity(
            request,
            request.user,
            ActivityLog.ActionType.UPDATE,
            f"{'Activated' if user.is_active else 'Deactivated'} user: {user.full_name} ({user.email})",
        )
        messages.success(request, f"{user.full_name or user.email} is now {'active' if user.is_active else 'inactive'}.")
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": True})
        return redirect(request.POST.get("next") or reverse("accounts:user-detail", kwargs={"pk": user.pk}))


class UserPasswordResetTriggerView(UserManagementAccessMixin, View):
    def post(self, request, *args, **kwargs):
        user = self.get_object()
        if not user.is_active:
            messages.error(request, "Activate this user before sending a password reset.")
            return redirect("accounts:user-detail", pk=user.pk)
        temporary_password = get_random_string(14)
        user.set_password(temporary_password)
        user.must_change_password = True
        user.session_invalid_before = timezone.now()
        user.save(update_fields=["password", "must_change_password", "session_invalid_before"])
        try:
            send_atms_email(
                subject="ZALA/ECO ENERGY temporary password",
                to=[user.email],
                greeting=f"Hello {user.full_name}",
                headline="Temporary Password Issued",
                intro="An administrator reset your ZALA/ECO ENERGY password. Sign in with this temporary password and create a new one immediately.",
                details=[
                    {"label": "Login Email", "value": user.email},
                    {"label": "Temporary Password", "value": temporary_password},
                ],
                note="If this was unexpected, contact your station manager or system administrator.",
                cta_label="Open Login",
                cta_url=build_public_url("/"),
            )
            messages.success(request, f"Temporary password sent to {user.email}.")
        except Exception as exc:
            messages.warning(request, f"Temporary password created but email failed: {exc}. Password: {temporary_password}")
        log_user_activity(
            request,
            request.user,
            ActivityLog.ActionType.UPDATE,
            f"Triggered password reset for user: {user.full_name} ({user.email})",
        )
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": True})
        return redirect(request.POST.get("next") or reverse("accounts:user-detail", kwargs={"pk": user.pk}))


class UserDeleteView(UserManagementAccessMixin, View):
    template_name = "accounts/users/_confirm_delete.html"

    def get(self, request, *args, **kwargs):
        user = self.get_object()
        return render(request, self.template_name, {"managed_user": user, "action": request.path})

    def post(self, request, *args, **kwargs):
        user = self.get_object()
        if user == request.user:
            messages.error(request, "You cannot delete your own account.")
            return redirect("accounts:user-list")

        linked_counts = [
            user.attendant_shifts.exists(),
            user.opened_shifts.exists(),
            user.closed_shifts.exists(),
            user.sales_made.exists(),
            user.deliveries_received.exists(),
            user.payments_received.exists(),
            user.tank_readings.exists(),
            user.activity_logs.exists(),
        ]
        if any(linked_counts):
            user.is_active = False
            user.session_invalid_before = timezone.now()
            user.save(update_fields=["is_active", "session_invalid_before"])
            log_user_activity(
                request,
                request.user,
                ActivityLog.ActionType.UPDATE,
                f"Deactivated user instead of deleting due to linked records: {user.full_name} ({user.email})",
            )
            messages.warning(request, "This user has operational records, so they were deactivated instead of deleted.")
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": True})
            return redirect("accounts:user-list")

        log_user_activity(
            request,
            request.user,
            ActivityLog.ActionType.DELETE,
            f"Deleted user: {user.full_name} ({user.email})",
        )
        user.delete()
        messages.success(request, "User deleted successfully.")
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": True})
        return redirect("accounts:user-list")


class RoleManagementView(SuperAdminMixin, TemplateView):
    """Role and permission management interface"""
    template_name = 'accounts/role_management.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # User role statistics
        role_stats = []
        for role_choice in User.Role.choices:
            role_code, role_name = role_choice
            count = User.objects.filter(role=role_code).count()
            role_stats.append({
                'code': role_code,
                'name': role_name,
                'count': count
            })
        
        context.update({
            'role_stats': role_stats,
            'total_users': User.objects.count(),
            'permissions': Permission.objects.all()[:20],  # Show first 20 permissions
            'recent_role_changes': self.get_recent_role_changes(),
        })
        
        return context
    
    def get_recent_role_changes(self):
        """Get recent role/permission changes"""
        return ActivityLog.objects.filter(
            action__in=['create', 'update'], 
            description__icontains='role'
        )[:10]


class UserRoleUpdateView(SuperAdminMixin, View):
    """Update user role"""
    
    def post(self, request, *args, **kwargs):
        user_id = kwargs.get('pk')
        new_role = request.POST.get('role')
        
        user = get_object_or_404(User, pk=user_id)
        old_role = user.role
        system_role_map = {
            User.Role.SUPERADMIN: SystemGroup.ADMIN,
            User.Role.ADMIN: SystemGroup.ADMIN,
            User.Role.STATION_MANAGER: SystemGroup.STATION_MANAGER,
            User.Role.SUPERVISOR: SystemGroup.SUPERVISOR,
            User.Role.PUMP_ATTENDANT: SystemGroup.PUMP_ATTENDANT,
            User.Role.ACCOUNTANT: SystemGroup.ACCOUNTANT,
            User.Role.CLIENT: SystemGroup.CUSTOMER,
        }

        if new_role in system_role_map:
            user.role = new_role
            user.save()
            sync_user_to_system_role(user, system_role_map[new_role])
            
            # Log the role change
            ActivityLog.objects.create(
                user=request.user,
                action=ActivityLog.ActionType.UPDATE,
                description=f"Changed role of {user.full_name} from {old_role} to {new_role}",
                ip_address=self.get_client_ip(request)
            )
            
            messages.success(request, f"Successfully updated {user.full_name}'s role to {new_role}")
        else:
            messages.error(request, "Invalid role selected")
        
        return redirect('accounts:role-management')
    
    def get_client_ip(self, request):
        """Get client IP address"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip


class ActivityLogView(AdminMixin, ListView):
    """Display user activity logs"""
    model = ActivityLog
    template_name = 'accounts/activity_logs.html'
    context_object_name = 'logs'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = ActivityLog.objects.select_related('user').all()
        
        # Filtering
        user_filter = self.request.GET.get('user')
        action_filter = self.request.GET.get('action')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        search = self.request.GET.get('search')
        
        if user_filter:
            queryset = queryset.filter(user__id=user_filter)
        if action_filter:
            queryset = queryset.filter(action=action_filter)
        if date_from:
            queryset = queryset.filter(created_at__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__lte=date_to)
        if search:
            queryset = queryset.filter(
                Q(description__icontains=search) |
                Q(user__full_name__icontains=search) |
                Q(user__email__icontains=search)
            )
            
        return queryset.order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Statistics for the page
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        
        current_user = self.request.GET.get('user')
        current_action = self.request.GET.get('action')
        current_date_from = self.request.GET.get('date_from')
        current_date_to = self.request.GET.get('date_to')
        current_search = self.request.GET.get('search')

        query_params = self.request.GET.copy()
        query_params.pop('page', None)
        preserved_query = query_params.urlencode()

        filtered_logs = self.get_queryset()
        active_filter_count = sum(
            1 for value in [current_user, current_action, current_date_from, current_date_to, current_search] if value
        )

        context.update({
            'users': User.objects.all(),
            'action_choices': ActivityLog.ActionType.choices,
            'total_activities': ActivityLog.objects.count(),
            'activities_today': ActivityLog.objects.filter(created_at__date=today).count(),
            'activities_this_week': ActivityLog.objects.filter(created_at__date__gte=week_ago).count(),
            'unique_users': ActivityLog.objects.exclude(user__isnull=True).values('user').distinct().count(),
            'filtered_total': filtered_logs.count(),
            'active_filter_count': active_filter_count,
            'preserved_query': f"&{preserved_query}" if preserved_query else "",

            # Current filter values
            'current_user': current_user,
            'current_action': current_action,
            'current_date_from': current_date_from,
            'current_date_to': current_date_to,
            'current_search': current_search,
        })
        
        return context


class PermissionManagementView(SuperAdminMixin, TemplateView):
    """Manage permissions for different roles"""
    template_name = 'accounts/permission_management.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Group permissions by content type for better organization
        permissions_by_app = {}
        for permission in Permission.objects.select_related('content_type').all():
            app_label = permission.content_type.app_label
            if app_label not in permissions_by_app:
                permissions_by_app[app_label] = []
            permissions_by_app[app_label].append(permission)
        
        context.update({
            'permissions_by_app': permissions_by_app,
            'role_choices': User.Role.choices,
            'role_permissions': RolePermission.objects.select_related('permission', 'granted_by').all()
        })
        
        return context
    
    def post(self, request, *args, **kwargs):
        """Handle permission assignment"""
        role = request.POST.get('role')
        permission_id = request.POST.get('permission_id')
        action = request.POST.get('action')  # 'grant' or 'revoke'
        
        if not all([role, permission_id, action]):
            messages.error(request, "Missing required parameters")
            return self.get(request, *args, **kwargs)
        
        try:
            permission = Permission.objects.get(id=permission_id)
            
            if action == 'grant':
                role_perm, created = RolePermission.objects.get_or_create(
                    role=role,
                    permission=permission,
                    defaults={'granted_by': request.user}
                )
                if created:
                    ActivityLog.objects.create(
                        user=request.user,
                        action=ActivityLog.ActionType.CREATE,
                        description=f"Granted permission '{permission.name}' to role '{role}'",
                        ip_address=self.get_client_ip(request)
                    )
                    messages.success(request, f"Permission granted successfully")
                else:
                    messages.info(request, f"Permission already granted to this role")
                    
            elif action == 'revoke':
                deleted_count = RolePermission.objects.filter(
                    role=role, 
                    permission=permission
                ).delete()[0]
                
                if deleted_count > 0:
                    ActivityLog.objects.create(
                        user=request.user,
                        action=ActivityLog.ActionType.DELETE,
                        description=f"Revoked permission '{permission.name}' from role '{role}'",
                        ip_address=self.get_client_ip(request)
                    )
                    messages.success(request, f"Permission revoked successfully")
                else:
                    messages.warning(request, f"Permission was not assigned to this role")
                    
        except Permission.DoesNotExist:
            messages.error(request, "Permission not found")
        
        return self.get(request, *args, **kwargs)
    
    def get_client_ip(self, request):
        """Get client IP address"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip


class SystemSettingsView(AdminMixin, View):
    """Manage system settings like colors, currency, etc."""
    template_name = 'accounts/system_settings.html'
    
    def get(self, request):
        """Display settings form"""
        settings = SystemSettings.get_settings()
        form = SystemSettingsForm(instance=settings)
        
        context = {
            'form': form,
            'settings': settings,
            'page_title': 'System Settings'
        }
        return render(request, self.template_name, context)
    
    def post(self, request):
        """Handle settings update"""
        settings = SystemSettings.get_settings()
        original_currency = (settings.currency or "").upper() if settings else ""
        original_petrol_price = settings.petrol_unit_price if settings else None
        original_diesel_price = settings.diesel_unit_price if settings else None
        form = SystemSettingsForm(request.POST, request.FILES, instance=settings)
        
        if form.is_valid():
            updated_settings = form.save(commit=False)
            previous_currency = original_currency
            new_currency = (updated_settings.currency or "").upper()
            updated_settings.updated_by = request.user
            conversion_summary = None
            if previous_currency and new_currency and previous_currency != new_currency:
                from .currency import convert_currency
                from .currency_conversion import convert_system_money_values

                if original_petrol_price is not None and updated_settings.petrol_unit_price == original_petrol_price:
                    updated_settings.petrol_unit_price = convert_currency(original_petrol_price, previous_currency, new_currency)
                if original_diesel_price is not None and updated_settings.diesel_unit_price == original_diesel_price:
                    updated_settings.diesel_unit_price = convert_currency(original_diesel_price, previous_currency, new_currency)
                conversion_summary = convert_system_money_values(
                    previous_currency,
                    new_currency,
                    exclude_settings_pk=updated_settings.pk,
                )
            updated_settings.save()
            
            # Log the activity
            ActivityLog.objects.create(
                user=request.user,
                action=ActivityLog.ActionType.UPDATE,
                description=f"Updated system settings",
                ip_address=self.get_client_ip(request)
            )
            
            if conversion_summary:
                messages.success(
                    request,
                    (
                        f"System settings updated successfully! Existing monetary values were converted "
                        f"from {previous_currency} to {new_currency} across {conversion_summary['converted_rows']} records."
                    ),
                )
            else:
                messages.success(request, "System settings updated successfully!")
            return redirect('accounts:system-settings')
        
        context = {
            'form': form,
            'settings': settings,
            'page_title': 'System Settings'
        }
        return render(request, self.template_name, context)
    
    def get_client_ip(self, request):
        """Get client IP address"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip


# Ã¢â€â‚¬Ã¢â€â‚¬ Currency exchange-rate API Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
@login_required_fn
def exchange_rates_api(request):
    """
    AJAX endpoint returning live exchange rates.
    GET /accounts/exchange-rates/
    GET /accounts/exchange-rates/?base=RWF
    """
    from .currency import get_exchange_rates, CURRENCY_SYMBOLS, CURRENCY_DECIMALS

    base = request.GET.get('base', '').upper()
    if not base:
        settings_obj = SystemSettings.get_settings()
        base = settings_obj.currency if settings_obj else 'USD'

    rates = get_exchange_rates(base)

    supported = ['USD', 'RWF', 'EUR', 'GBP', 'KES', 'UGX', 'TZS']
    result = {}
    for code in supported:
        if code in rates:
            result[code] = {
                'rate': rates[code],
                'symbol': CURRENCY_SYMBOLS.get(code, code),
                'decimals': CURRENCY_DECIMALS.get(code, 2),
            }

    return JsonResponse({
        'success': True,
        'base': base,
        'rates': result,
    })
