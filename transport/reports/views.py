from collections import defaultdict
from decimal import Decimal
from io import BytesIO
import csv
import os
import tempfile
import warnings
from pathlib import Path

from django.core.management import call_command
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.generic import TemplateView, View

from accounts.rbac import RBACRequiredMixin, SystemGroup
from transport.customers.models import Customer
from transport.drivers.models import Driver
from transport.finance.models import Payment
from transport.orders.models import Order
from transport.trips.models import Shipment, Trip
from transport.vehicles.models import Vehicle


REPORT_LABELS = {
    "order_number": "Order Number",
    "customer": "Customer",
    "status": "Status",
    "quoted_price": "Quoted Price",
    "outstanding_balance": "Outstanding Balance",
    "vehicle": "Vehicle",
    "ownership": "Ownership",
    "trip_count": "Trip Count",
    "active_trip_count": "Active Trip Count",
    "route": "Route",
    "trip": "Trip",
    "distance": "Distance",
    "cost_per_km": "Cost per KM",
    "gross_profit": "Gross Profit",
    "expenses": "Expenses",
    "net_profit": "Net Profit",
    "revenue": "Revenue",
}


def _report_title(report_type):
    report_map = {
        "job-status": "Job Status Report",
        "fleet-utilization": "Fleet Utilization Report",
        "delivery-performance": "Delivery Performance Report",
        "profit-trip": "Profit Per Trip Report",
        "profit-customer": "Profit Per Customer Report",
        "profit-route": "Profit Per Route Report",
    }
    return report_map.get(report_type, "ZALA Terminal Report")


def _styled_report_workbook(report_type, headers, rows, filters):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from PIL import Image as PILImage
    from PIL import Image as PILModule

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ZALA Terminal Report"

    max_column = max(len(headers), 1)
    sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=max_column)
    title_cell = sheet.cell(row=1, column=2, value=_report_title(report_type))
    title_cell.font = Font(name="Arial", size=15, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="166534")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    logo_path = Path(__file__).resolve().parents[2] / "static" / "img" / "ZALA Terminal.png"
    if logo_path.exists():
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", PILModule.DecompressionBombWarning)
                source_logo = PILImage.open(logo_path)
            with source_logo:
                source_logo.thumbnail((180, 90))
                logo_buffer = BytesIO()
                source_logo.save(logo_buffer, format="PNG")
                logo_buffer.seek(0)
            logo = XLImage(logo_buffer)
            sheet.add_image(logo, "A1")
        except Exception:
            pass

    generated_at = timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")
    sheet.cell(row=2, column=1, value="Generated")
    sheet.cell(row=2, column=2, value=generated_at)
    sheet.cell(row=3, column=1, value="Rows")
    sheet.cell(row=3, column=2, value=len(rows))

    meta_label_font = Font(name="Arial", bold=True, color="166534")
    meta_value_font = Font(name="Arial", color="1F2937")
    for row_idx in (2, 3):
        sheet.cell(row=row_idx, column=1).font = meta_label_font
        sheet.cell(row=row_idx, column=2).font = meta_value_font

    header_row = 5
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="DCFCE7")
    alt_fill = PatternFill("solid", fgColor="F9FAFB")
    money_format = '#,##0.00;(#,##0.00);"-"'

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(name="Arial", bold=True, color="166534")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_offset, row in enumerate(rows, start=1):
        row_idx = header_row + row_offset
        use_alt = row_offset % 2 == 0
        for col_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", color="111827")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            if use_alt:
                cell.fill = alt_fill
            if isinstance(value, (int, float, Decimal)):
                cell.number_format = money_format if any(
                    key in headers[col_idx - 1].lower() for key in ("price", "balance", "revenue", "profit", "expense", "cost")
                ) else "0"

    for idx, header in enumerate(headers, start=1):
        values = [str(header)]
        for row in rows:
            if idx - 1 < len(row):
                values.append("" if row[idx - 1] is None else str(row[idx - 1]))
        max_len = min(max(len(value) for value in values) + 2, 28)
        sheet.column_dimensions[sheet.cell(row=header_row, column=idx).column_letter].width = max_len

    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.sheet_view.showGridLines = False
    return workbook


def _decimal(value):
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _parse_filters(filters):
    date_from = parse_date(filters.get("date_from", ""))
    date_to = parse_date(filters.get("date_to", ""))
    status = filters.get("status", "").strip()
    return {"date_from": date_from, "date_to": date_to, "status": status}


def _trip_queryset(filters=None):
    queryset = Trip.objects.select_related("customer", "route", "vehicle", "driver").prefetch_related(
        "shipments__order__customer",
        "payments",
        "expenses",
    )
    if not filters:
        return queryset
    if filters.get("status"):
        queryset = queryset.filter(status=filters["status"])
    if filters.get("date_from"):
        queryset = queryset.filter(created_at__date__gte=filters["date_from"])
    if filters.get("date_to"):
        queryset = queryset.filter(created_at__date__lte=filters["date_to"])
    return queryset


def _order_queryset(filters=None):
    queryset = Order.objects.select_related("customer", "route", "unit")
    if not filters:
        return queryset
    if filters.get("status"):
        queryset = queryset.filter(status=filters["status"])
    if filters.get("date_from"):
        queryset = queryset.filter(created_at__date__gte=filters["date_from"])
    if filters.get("date_to"):
        queryset = queryset.filter(created_at__date__lte=filters["date_to"])
    return queryset


def _payment_queryset(filters=None):
    queryset = Payment.objects.exclude(status=Payment.Status.FAILED).select_related("customer", "trip", "order")
    if not filters:
        return queryset
    if filters.get("status"):
        queryset = queryset.filter(status=filters["status"])
    if filters.get("date_from"):
        queryset = queryset.filter(payment_date__gte=filters["date_from"])
    if filters.get("date_to"):
        queryset = queryset.filter(payment_date__lte=filters["date_to"])
    return queryset


def _trip_revenue_allocations(trip):
    allocations = []
    shipments = list(trip.shipments.select_related("order__customer"))

    if shipments:
        total_load = trip.total_load or Decimal("0")
        for shipment in shipments:
            order = shipment.order
            order_total_quantity = order.total_quantity_value
            order_revenue_basis = order.total_invoiced or order.quoted_price or Decimal("0")
            revenue_share = Decimal("0")
            if order_total_quantity > 0:
                revenue_share = (order_revenue_basis * shipment.quantity) / order_total_quantity

            expense_share = Decimal("0")
            if total_load > 0:
                expense_share = (trip.total_expenses * shipment.quantity) / total_load

            allocations.append(
                {
                    "order": order,
                    "customer": order.customer,
                    "route": trip.route,
                    "quantity": shipment.quantity,
                    "revenue": revenue_share,
                    "expenses": expense_share,
                    "net_profit": revenue_share - expense_share,
                }
            )
        return allocations

    if trip.job_id and trip.job:
        allocations.append(
            {
                "order": trip.job,
                "customer": trip.job.customer,
                "route": trip.route,
                "quantity": trip.quantity or Decimal("0"),
                "revenue": trip.total_revenue,
                "expenses": trip.total_expenses,
                "net_profit": trip.net_profit,
            }
        )
        return allocations

    allocations.append(
        {
            "order": None,
            "customer": trip.customer,
            "route": trip.route,
            "quantity": trip.quantity or Decimal("0"),
            "revenue": trip.total_revenue,
            "expenses": trip.total_expenses,
            "net_profit": trip.net_profit,
        }
    )
    return allocations


def _report_rows_job_status(filters):
    rows = []
    for order in _order_queryset(filters):
        rows.append(
            {
                "order_number": order.order_number,
                "customer": order.customer.company_name,
                "status": order.get_status_display(),
                "quoted_price": order.quoted_price or Decimal("0"),
                "outstanding_balance": order.outstanding_balance or Decimal("0"),
            }
        )
    return rows, ["order_number", "customer", "status", "quoted_price", "outstanding_balance"]


def _report_rows_fleet_utilization(filters):
    trip_filters = Q()
    if filters.get("date_from"):
        trip_filters &= Q(trips__created_at__date__gte=filters["date_from"])
    if filters.get("date_to"):
        trip_filters &= Q(trips__created_at__date__lte=filters["date_to"])
    if filters.get("status"):
        trip_filters &= Q(trips__status=filters["status"])

    vehicles = (
        Vehicle.objects.select_related("owner")
        .annotate(
            trip_count=Count("trips", filter=trip_filters, distinct=True),
            active_trip_count=Count(
                "trips",
                filter=trip_filters & Q(trips__status__in=[Trip.TripStatus.ASSIGNED, Trip.TripStatus.IN_TRANSIT]),
                distinct=True,
            ),
        )
        .order_by("plate_number")
    )

    rows = [
        {
            "vehicle": vehicle.plate_number,
            "ownership": vehicle.get_ownership_type_display(),
            "status": vehicle.get_status_display(),
            "trip_count": vehicle.trip_count,
            "active_trip_count": vehicle.active_trip_count,
        }
        for vehicle in vehicles
    ]
    return rows, ["vehicle", "ownership", "status", "trip_count", "active_trip_count"]


def _report_rows_delivery_performance(filters):
    rows = []
    for trip in _trip_queryset(filters):
        rows.append(
            {
                "trip": trip.order_number,
                "route": f"{trip.route.origin} -> {trip.route.destination}",
                "status": trip.get_status_display(),
                "distance": trip.distance or Decimal("0"),
                "cost_per_km": trip.cost_per_km or Decimal("0"),
            }
        )
    return rows, ["trip", "route", "status", "distance", "cost_per_km"]


def _report_rows_profit_trip(filters):
    rows = []
    for trip in _trip_queryset(filters):
        rows.append(
            {
                "trip": trip.order_number,
                "customer": trip.customer.company_name,
                "gross_profit": trip.gross_profit or Decimal("0"),
                "expenses": trip.total_expenses,
                "net_profit": trip.net_profit,
            }
        )
    return rows, ["trip", "customer", "gross_profit", "expenses", "net_profit"]


def _report_rows_profit_customer(filters):
    aggregates = defaultdict(lambda: {"revenue": Decimal("0"), "expenses": Decimal("0"), "net_profit": Decimal("0")})

    for trip in _trip_queryset(filters):
        for allocation in _trip_revenue_allocations(trip):
            customer = allocation["customer"]
            key = customer.company_name if customer else "Unassigned"
            aggregates[key]["revenue"] += allocation["revenue"]
            aggregates[key]["expenses"] += allocation["expenses"]
            aggregates[key]["net_profit"] += allocation["net_profit"]

    rows = []
    for customer_name in sorted(aggregates.keys()):
        rows.append(
            {
                "customer": customer_name,
                "revenue": aggregates[customer_name]["revenue"],
                "expenses": aggregates[customer_name]["expenses"],
                "net_profit": aggregates[customer_name]["net_profit"],
            }
        )
    return rows, ["customer", "revenue", "expenses", "net_profit"]


def _report_rows_profit_route(filters):
    aggregates = defaultdict(lambda: {"revenue": Decimal("0"), "expenses": Decimal("0"), "net_profit": Decimal("0")})

    for trip in _trip_queryset(filters):
        route_name = f"{trip.route.origin} -> {trip.route.destination}"
        aggregates[route_name]["revenue"] += trip.total_revenue
        aggregates[route_name]["expenses"] += trip.total_expenses
        aggregates[route_name]["net_profit"] += trip.net_profit

    rows = []
    for route_name in sorted(aggregates.keys()):
        rows.append(
            {
                "route": route_name,
                "revenue": aggregates[route_name]["revenue"],
                "expenses": aggregates[route_name]["expenses"],
                "net_profit": aggregates[route_name]["net_profit"],
            }
        )
    return rows, ["route", "revenue", "expenses", "net_profit"]


REPORT_BUILDERS = {
    "job-status": _report_rows_job_status,
    "fleet-utilization": _report_rows_fleet_utilization,
    "delivery-performance": _report_rows_delivery_performance,
    "profit-trip": _report_rows_profit_trip,
    "profit-customer": _report_rows_profit_customer,
    "profit-route": _report_rows_profit_route,
}


class ReportsDashboardView(RBACRequiredMixin, TemplateView):
    template_name = "transport/reports/dashboard.html"
    allowed_roles = (SystemGroup.ADMIN, SystemGroup.OPERATIONS_MANAGER, SystemGroup.FINANCE)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        payment_queryset = _payment_queryset()
        trip_queryset = _trip_queryset()

        total_revenue = payment_queryset.aggregate(total=Sum("amount")).get("total") or Decimal("0")
        net_profit = sum((trip.net_profit for trip in trip_queryset), Decimal("0"))
        active_trips = trip_queryset.filter(
            status__in=[Trip.TripStatus.APPROVED, Trip.TripStatus.ASSIGNED, Trip.TripStatus.IN_TRANSIT]
        ).count()
        completed_trips = trip_queryset.filter(
            status__in=[Trip.TripStatus.DELIVERED, Trip.TripStatus.COMPLETED, Trip.TripStatus.CLOSED]
        ).count()

        context.update(
            {
                "total_trips": trip_queryset.count(),
                "active_trips": active_trips,
                "completed_trips": completed_trips,
                "total_orders": Order.objects.count(),
                "total_vehicles": Vehicle.objects.count(),
                "available_vehicles": Vehicle.objects.filter(status=Vehicle.VehicleStatus.AVAILABLE).count(),
                "total_revenue": total_revenue,
                "net_profit": net_profit,
                "recent_trips": trip_queryset.order_by("-created_at")[:10],
                "soa_years": list(range(timezone.now().year, timezone.now().year - 5, -1)),
                "report_types": [
                    {"id": "job-status", "label": "Job Status"},
                    {"id": "fleet-utilization", "label": "Fleet Utilization"},
                    {"id": "delivery-performance", "label": "Delivery Performance"},
                    {"id": "profit-trip", "label": "Profit per Trip"},
                    {"id": "profit-customer", "label": "Profit per Customer"},
                    {"id": "profit-route", "label": "Profit per Route"},
                ],
            }
        )
        return context


class GenerateSOAReportView(RBACRequiredMixin, View):
    allowed_roles = (SystemGroup.ADMIN, SystemGroup.OPERATIONS_MANAGER, SystemGroup.FINANCE)

    def post(self, request, *args, **kwargs):
        try:
            year = int(request.POST.get("year") or timezone.now().year)
        except (TypeError, ValueError):
            year = timezone.now().year

        fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            call_command("generate_soa_report", year=year, output=temp_path)
            with open(temp_path, "rb") as report_file:
                response = HttpResponse(
                    report_file.read(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            response["Content-Disposition"] = f'attachment; filename="SOA_{year}.xlsx"'
            return response
        finally:
            if os.path.exists(temp_path):
                os.unlink(temp_path)


class ExportExcelView(RBACRequiredMixin, View):
    allowed_roles = (SystemGroup.ADMIN, SystemGroup.OPERATIONS_MANAGER, SystemGroup.FINANCE)

    def post(self, request, *args, **kwargs):
        report_type = request.POST.get("report_type", "profit-trip")
        raw_columns = request.POST.get("columns", "")
        columns = [col.strip() for col in raw_columns.split(",") if col.strip()]
        parsed_filters = _parse_filters(request.POST)
        headers, rows = build_report_rows(report_type, columns, request.POST)
        try:
            workbook = _styled_report_workbook(report_type, headers, rows, parsed_filters)

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{report_type}.xlsx"'
            return response
        except ModuleNotFoundError:
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = f'attachment; filename="{report_type}.csv"'
            writer = csv.writer(response)
            writer.writerow(headers)
            writer.writerows(rows)
            return response


def build_report_rows(report_type, columns, filters):
    parsed_filters = _parse_filters(filters)
    builder = REPORT_BUILDERS.get(report_type, _report_rows_profit_trip)
    row_dicts, default_columns = builder(parsed_filters)
    selected_columns = columns or default_columns
    headers = [REPORT_LABELS.get(column, column.replace("_", " ").title()) for column in selected_columns]

    rows = []
    for row_dict in row_dicts:
        rows.append([row_dict.get(column, "") for column in selected_columns])
    return headers, rows
