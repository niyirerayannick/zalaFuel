from collections import defaultdict
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db.models import Prefetch, Sum
from django.utils import timezone

from transport.finance.models import Expense, Payment
from transport.trips.models import Shipment


TITLE_FILL = "00205B"
HEADER_FILL = "BDD7EE"
WHITE = "FFFFFF"
NAVY = "00205B"
LIGHT_ROW = "F2F2F2"
POSITIVE_GREEN = "006100"
NEGATIVE_RED = "9C0006"
MONEY_FORMAT = '#,##0;(#,##0);"-"'
DATE_FORMAT = "DD/MM/YYYY"


def _decimal(value):
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal("0")


def _safe_date(value):
    if not value:
        return None
    if hasattr(value, "date"):
        return value.date()
    return value


def _excel_date(value):
    value = _safe_date(value)
    return value or ""


def _expense_name(expense):
    if expense.type_id and expense.type:
        return (expense.type.name or "").strip().lower()
    return (expense.category or "").strip().lower()


def _contains_any(name, keywords):
    return any(keyword in name for keyword in keywords)


class Command(BaseCommand):
    help = "Generate an ZALA Terminal SOA (Statement of Account) Excel report."

    def add_arguments(self, parser):
        parser.add_argument("--year", type=int, default=timezone.now().year)
        parser.add_argument("--output", type=str, default=None)

    def handle(self, *args, **options):
        try:
            from openpyxl import Workbook
            from openpyxl.formatting.rule import CellIsRule
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except ModuleNotFoundError as exc:
            raise CommandError(
                "openpyxl is required for generate_soa_report. Install it in the active environment first."
            ) from exc

        year = options["year"] or timezone.now().year
        output_path = Path(options["output"] or f"soa_report_{year}.xlsx")
        output_path.parent.mkdir(parents=True, exist_ok=True)

        self.Alignment = Alignment

        shipments = self._get_shipments(year)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f"SOA {year}"

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_fill = PatternFill("solid", fgColor=TITLE_FILL)
        header_fill = PatternFill("solid", fgColor=HEADER_FILL)
        light_fill = PatternFill("solid", fgColor=LIGHT_ROW)
        white_fill = PatternFill("solid", fgColor=WHITE)
        title_font = Font(name="Arial", size=14, bold=True, color=WHITE)
        header_font = Font(name="Arial", bold=True, color=NAVY)
        body_font = Font(name="Arial", size=10)
        balance_font = Font(name="Arial", size=10, bold=True, color=NAVY)

        self._build_headers(sheet, title_fill, header_fill, title_font, header_font, border)

        start_row = 5
        total_profit = Decimal("0")

        for index, shipment in enumerate(shipments, start=start_row):
            row_fill = light_fill if (index - start_row) % 2 else white_fill
            row_values, row_profit = self._build_row(shipment, index)
            total_profit += row_profit

            for column_index, value in enumerate(row_values, start=1):
                cell = sheet.cell(row=index, column=column_index, value=value)
                cell.font = body_font
                cell.alignment = self.Alignment(vertical="center", horizontal="center")
                cell.border = border
                cell.fill = row_fill

            for column in ("A", "R", "U", "X", "AA"):
                cell = sheet[f"{column}{index}"]
                if cell.value:
                    cell.number_format = DATE_FORMAT

            for column in ("G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "S", "V", "Y", "AB", "AD", "AE", "AF"):
                sheet[f"{column}{index}"].number_format = MONEY_FORMAT

            sheet[f"AF{index}"].font = balance_font

        if shipments:
            sheet.conditional_formatting.add(
                f"AE{start_row}:AE{sheet.max_row}",
                CellIsRule(operator="greaterThan", formula=["0"], font=Font(name="Arial", color=POSITIVE_GREEN)),
            )
            sheet.conditional_formatting.add(
                f"AE{start_row}:AE{sheet.max_row}",
                CellIsRule(operator="lessThan", formula=["0"], font=Font(name="Arial", color=NEGATIVE_RED)),
            )

        sheet.freeze_panes = "D5"
        self._set_column_widths(sheet)

        workbook.save(output_path)

        self.stdout.write(
            self.style.SUCCESS(
                f"SOA report generated successfully. Rows written: {len(shipments)} | "
                f"Output: {output_path} | Total profit: {total_profit.quantize(Decimal('0.01'))}"
            )
        )

    def _get_shipments(self, year):
        payment_prefetch = Prefetch(
            "order__payments",
            queryset=Payment.objects.exclude(status=Payment.Status.FAILED).order_by("payment_date", "created_at"),
        )
        expense_prefetch = Prefetch(
            "trip__expenses",
            queryset=Expense.objects.select_related("type").order_by("expense_date", "created_at"),
        )

        queryset = (
            Shipment.objects.filter(status__in=[Shipment.Status.IN_TRANSIT, Shipment.Status.DELIVERED], trip__isnull=False)
            .select_related("trip__vehicle", "trip__driver", "trip__route", "order__customer")
            .prefetch_related(payment_prefetch, expense_prefetch, "trip__allowances")
            .order_by("created_at")
        )

        filtered = []
        for shipment in queryset:
            report_date = self._shipment_report_date(shipment)
            if report_date and report_date.year == year:
                filtered.append(shipment)

        filtered.sort(key=lambda shipment: self._shipment_report_date(shipment) or timezone.now().date())
        return filtered

    def _shipment_report_date(self, shipment):
        trip = shipment.trip
        if trip and trip.created_at:
            return _safe_date(trip.created_at)
        return _safe_date(shipment.created_at)

    def _build_headers(self, sheet, title_fill, header_fill, title_font, header_font, border):
        sheet.merge_cells("A1:AF1")
        title_cell = sheet["A1"]
        title_cell.value = "AFRILLOT   ACCOUNT"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = self.Alignment(horizontal="center", vertical="center")
        title_cell.border = border

        section_ranges = {
            "A3:G3": "PARTICULARS",
            "H3:P3": "ADDITIONAL COST",
            "Q3:AC3": "",
            "AD3:AF3": "BALANCE",
        }
        for cell_range, value in section_ranges.items():
            sheet.merge_cells(cell_range)
            cell = sheet[cell_range.split(":")[0]]
            cell.value = value
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = self.Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        headers = [
            "DATE",
            "FILE NO.",
            "CONSIGNEE",
            "BL NO.",
            "20'",
            "40'",
            "TOTAL AMOUNT",
            "VERF.",
            "STORAGE",
            "REMOVAL",
            "CWR",
            "HANDOVER",
            "CLEARANCE FEES",
            "TRANSPORT",
            "DEMURRAGE",
            "ADDITIONAL TOTAL",
            "GRAND TOTAL",
            "DATE",
            "AMOUNT",
            "MODE",
            "DATE",
            "AMOUNT",
            "MODE",
            "DATE",
            "AMOUNT",
            "MODE",
            "DATE",
            "AMOUNT",
            "MODE",
            "TOTAL RECEIVED",
            "PROFIT",
            "TOTAL PROFIT BALANCE",
        ]
        for index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=4, column=index, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = self.Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    def _build_row(self, shipment, row_number):
        trip = shipment.trip
        order = shipment.order
        vehicle = getattr(trip, "vehicle", None)
        customer = getattr(order, "customer", None)
        report_date = self._shipment_report_date(shipment)
        container_20, container_40 = self._container_columns(vehicle)

        total_amount = self._allocated_order_revenue(order, shipment)
        additional_costs = self._allocated_additional_costs(trip, shipment)
        payments = self._payment_slots(order)

        row = [
            _excel_date(report_date),
            getattr(trip, "order_number", "") or "",
            getattr(customer, "company_name", "") or "",
            self._bl_number(shipment),
            container_20,
            container_40,
            float(total_amount),
            float(additional_costs["verification"]),
            float(additional_costs["storage"]),
            float(additional_costs["removal"]),
            float(additional_costs["cwr"]),
            float(additional_costs["handover"]),
            float(additional_costs["clearance"]),
            float(additional_costs["transport"]),
            float(additional_costs["demurrage"]),
            f'=IFERROR(SUM(H{row_number}:O{row_number}),0)',
            f'=IFERROR(G{row_number}+P{row_number},0)',
        ]

        for payment in payments:
            row.extend(
                [
                    _excel_date(payment["date"]),
                    float(payment["amount"]) if payment["amount"] else "",
                    payment["mode"],
                ]
            )

        row.extend(
            [
                f'=IFERROR(S{row_number},0)+IFERROR(V{row_number},0)+IFERROR(Y{row_number},0)+IFERROR(AB{row_number},0)',
                f'=IFERROR(AD{row_number}-Q{row_number},0)',
                f'=AE{row_number}' if row_number == 5 else f'=AF{row_number - 1}+AE{row_number}',
            ]
        )

        python_profit = sum((payment["amount"] for payment in payments), Decimal("0")) - (total_amount + sum(additional_costs.values(), Decimal("0")))
        return row, python_profit

    def _container_columns(self, vehicle):
        plate_number = getattr(vehicle, "plate_number", "") or ""
        capacity_kg = _decimal(getattr(vehicle, "capacity", 0)) * Decimal("1000")
        if capacity_kg and capacity_kg > Decimal("22000"):
            return "", plate_number
        return plate_number, ""

    def _bl_number(self, shipment):
        return getattr(shipment, "shipment_number", "") or f"SHP-{shipment.pk}"

    def _allocated_order_revenue(self, order, shipment):
        total_quantity = _decimal(getattr(order, "total_quantity", 0))
        shipment_quantity = _decimal(getattr(shipment, "quantity", 0))
        revenue_basis = _decimal(getattr(order, "total_invoiced", 0)) or _decimal(getattr(order, "quoted_price", 0))
        if total_quantity > 0 and shipment_quantity > 0:
            return (revenue_basis * shipment_quantity) / total_quantity
        return revenue_basis

    def _shipment_ratio(self, trip, shipment):
        if not trip:
            return Decimal("1")
        total_load = _decimal(getattr(trip, "total_load", 0))
        shipment_quantity = _decimal(getattr(shipment, "quantity", 0))
        if total_load > 0 and shipment_quantity > 0:
            return shipment_quantity / total_load
        shipment_count = max(trip.shipments.count(), 1)
        return Decimal("1") / Decimal(str(shipment_count))

    def _allocated_additional_costs(self, trip, shipment):
        if not trip:
            return {
                "verification": Decimal("0"),
                "storage": Decimal("0"),
                "removal": Decimal("0"),
                "cwr": Decimal("0"),
                "handover": Decimal("0"),
                "clearance": Decimal("0"),
                "transport": Decimal("0"),
                "demurrage": Decimal("0"),
            }
        ratio = self._shipment_ratio(trip, shipment)
        bucket_totals = defaultdict(lambda: Decimal("0"))

        expense_buckets = {
            "verification": ["verf", "verification"],
            "storage": ["storage"],
            "removal": ["removal", "handling", "loading"],
            "cwr": ["cwr", "warehouse receipt"],
            "handover": ["handover", "offloading"],
            "clearance": ["clearance", "customs"],
            "demurrage": ["demurrage"],
            "transport_misc": ["fuel", "toll", "parking", "vehicle rent", "rent", "maintenance", "miscellaneous"],
        }

        for expense in trip.expenses.all():
            name = _expense_name(expense)
            allocated_amount = _decimal(expense.amount) * ratio
            matched = False
            for bucket, keywords in expense_buckets.items():
                if _contains_any(name, keywords):
                    bucket_totals[bucket] += allocated_amount
                    matched = True
                    break
            if not matched:
                # Keep every unmatched trip expense in the shipment totals instead of excluding it.
                bucket_totals["transport_misc"] += allocated_amount

        transport = (_decimal(trip.fuel_cost) + _decimal(trip.rental_fee)) * ratio
        allowances_total = (
            trip.allowances.filter(status="APPROVED").aggregate(total_amount=Sum("amount")).get("total_amount")
            or Decimal("0")
        )
        transport += _decimal(allowances_total) * ratio
        transport += bucket_totals["transport_misc"]

        return {
            "verification": bucket_totals["verification"],
            "storage": bucket_totals["storage"],
            "removal": bucket_totals["removal"],
            "cwr": bucket_totals["cwr"],
            "handover": bucket_totals["handover"],
            "clearance": bucket_totals["clearance"],
            "transport": transport,
            "demurrage": bucket_totals["demurrage"],
        }

    def _payment_slots(self, order):
        if not order:
            return [{"date": None, "amount": Decimal("0"), "mode": ""} for _ in range(4)]

        received_payments = []
        for payment in order.payments.exclude(status=Payment.Status.FAILED).order_by("payment_date", "created_at"):
            amount = payment.collected_amount
            if amount <= 0:
                continue
            received_payments.append(
                {
                    "date": payment.payment_date,
                    "amount": amount,
                    "mode": payment.get_payment_method_display().upper() if payment.payment_method else "",
                }
            )
            if len(received_payments) == 4:
                break

        while len(received_payments) < 4:
            received_payments.append({"date": None, "amount": Decimal("0"), "mode": ""})
        return received_payments

    def _set_column_widths(self, sheet):
        widths = {
            "A": 14,
            "B": 10,
            "C": 22,
            "D": 18,
            "E": 14,
            "F": 14,
            "G": 14,
            "H": 12,
            "I": 12,
            "J": 12,
            "K": 12,
            "L": 12,
            "M": 12,
            "N": 12,
            "O": 12,
            "P": 13,
            "Q": 13,
            "R": 11,
            "S": 11,
            "T": 11,
            "U": 11,
            "V": 11,
            "W": 11,
            "X": 11,
            "Y": 11,
            "Z": 11,
            "AA": 11,
            "AB": 11,
            "AC": 11,
            "AD": 14,
            "AE": 12,
            "AF": 18,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
