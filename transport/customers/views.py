# Customer Module Views
import string
import secrets
import warnings
from io import BytesIO
from pathlib import Path
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib import messages
from django.urls import reverse_lazy
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count, Avg
from django.utils.decorators import method_decorator
from django.core.mail import send_mail
from django.conf import settings as app_settings
from django.utils import timezone
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from accounts.models import User as AuthUser
from accounts.rbac import SystemGroup, user_has_role
from .models import Customer
from .forms import CustomerForm
from transport.trips.models import Trip


def _logo_path():
    candidate = Path(__file__).resolve().parents[2] / "static" / "img" / "ZALA Terminal.png"
    return candidate if candidate.exists() else None


def _logo_stream(max_width=900):
    logo = _logo_path()
    if not logo:
        return None

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", PILImage.DecompressionBombWarning)
        img = PILImage.open(logo)
        img.load()
    with img:
        img.thumbnail((max_width, max_width))
        stream = BytesIO()
        img.save(stream, format="PNG", optimize=True)
        stream.seek(0)
        return stream


def generate_password(length=10):
    """Generate a secure random password"""
    alphabet = string.ascii_letters + string.digits + '!@#$%'
    password = ''.join(secrets.choice(alphabet) for _ in range(length))
    password = (
        secrets.choice(string.ascii_uppercase) +
        secrets.choice(string.ascii_lowercase) +
        secrets.choice(string.digits) +
        secrets.choice('!@#$%') +
        password[4:]
    )
    return password

class StaffRequiredMixin(UserPassesTestMixin):
    """Mixin to require staff access level"""
    def test_func(self):
        return self.request.user.is_authenticated and user_has_role(
            self.request.user,
            SystemGroup.ADMIN,
            SystemGroup.OPERATIONS_MANAGER,
            SystemGroup.LOGISTICS_COORDINATOR,
        )

class CustomerListView(StaffRequiredMixin, ListView):
    model = Customer
    template_name = 'transport/customers/list.html'
    context_object_name = 'customers'
    paginate_by = 20

    def get_template_names(self):
        if self.request.GET.get("partial") == "list":
            return ["transport/customers/_list_content.html"]
        return [self.template_name]

    def get_queryset(self):
        queryset = Customer.objects.all()
        search = self.request.GET.get('search')
        status = self.request.GET.get('status')
        
        if search:
            queryset = queryset.filter(
                Q(company_name__icontains=search) |
                Q(contact_person__icontains=search) |
                Q(email__icontains=search)
            )
        
        if status:
            queryset = queryset.filter(status=status)
        
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Customer statistics
        all_customers = Customer.objects.all()
        context['total_customers'] = all_customers.count()
        context['active_customers'] = all_customers.filter(status='ACTIVE').count()
        context['inactive_customers'] = all_customers.filter(status='INACTIVE').count() + all_customers.filter(status='SUSPENDED').count()
        
        # Financial metrics - TODO: Implement when Trip model is ready
        context['total_revenue'] = 0  # TODO: Calculate from trips
        context['customers_with_outstanding'] = 0  # TODO: Implement
        
        context['status_choices'] = Customer.STATUS_CHOICES
        return context


class CustomerExportMixin(StaffRequiredMixin):
    def get_filtered_queryset(self):
        view = CustomerListView()
        view.request = self.request
        return view.get_queryset()


class CustomerExcelExportView(CustomerExportMixin, ListView):
    def get(self, request, *args, **kwargs):
        queryset = self.get_filtered_queryset()
        filename = "customers_report.xlsx"
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError:
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="customers_report.csv"'
            response.write("Company,Contact Person,Phone,Email,Status,Created\n")
            for customer in queryset:
                response.write(
                    f'"{customer.company_name}","{customer.contact_person or ""}","{customer.phone or ""}","{customer.email or ""}","{customer.get_status_display()}","{customer.created_at:%Y-%m-%d}"\n'
                )
            return response

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Customers"

        header_fill = PatternFill(fill_type="solid", fgColor="0F5B2A")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        sheet.merge_cells("A1:F1")
        sheet["A1"] = "ZALA Terminal Customer Report"
        sheet["A1"].font = Font(color="0F5B2A", bold=True, size=16)
        sheet.merge_cells("A2:F2")
        sheet["A2"] = f"Generated on {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
        sheet["A2"].font = Font(color="475569", italic=True, size=10)

        headers = ["Company", "Contact Person", "Phone", "Email", "Status", "Created"]
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=4, column=idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        row_no = 5
        for customer in queryset:
            values = [
                customer.company_name,
                customer.contact_person or "-",
                customer.phone or "-",
                customer.email or "-",
                customer.get_status_display(),
                customer.created_at.strftime("%d/%m/%Y"),
            ]
            for col_no, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_no, column=col_no, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top")
            row_no += 1

        sheet.auto_filter.ref = f"A4:F{max(row_no - 1, 4)}"
        sheet.freeze_panes = "A5"
        for column_cells in sheet.columns:
            column_index = column_cells[0].column
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 3, 28)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type=mime_type)
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class CustomerPdfExportView(CustomerExportMixin, ListView):
    def get(self, request, *args, **kwargs):
        queryset = self.get_filtered_queryset()
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=12 * mm,
            leftMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
        )
        styles = getSampleStyleSheet()

        logo_stream = _logo_stream()
        header_left = []
        if logo_stream:
            header_left.append(Image(logo_stream, width=34 * mm, height=16 * mm))
            header_left.append(Spacer(1, 2 * mm))
        header_left.extend(
            [
                Paragraph("<font color='#0F5B2A'><b>ZALA Terminal Customer Report</b></font>", styles["Title"]),
                Paragraph("Customer directory export generated from ZALA Terminal.", styles["Normal"]),
            ]
        )
        header_right = [
            Paragraph("<b>Report</b><br/>Customer Register", styles["Normal"]),
            Spacer(1, 2),
            Paragraph(f"<b>Generated</b><br/>{timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}", styles["Normal"]),
            Spacer(1, 2),
            Paragraph(f"<b>Total Customers</b><br/>{queryset.count()}", styles["Normal"]),
        ]
        header_table = Table([[header_left, header_right]], colWidths=[170 * mm, 85 * mm])
        header_table.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1E7D7")),
                    ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#EAF7EF")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )

        rows = [["Company", "Contact Person", "Phone", "Email", "Status", "Created"]]
        for customer in queryset:
            rows.append(
                [
                    customer.company_name,
                    customer.contact_person or "-",
                    customer.phone or "-",
                    customer.email or "-",
                    customer.get_status_display(),
                    customer.created_at.strftime("%d/%m/%Y"),
                ]
            )
        if len(rows) == 1:
            rows.append(["-", "-", "-", "-", "-", "-"])

        table = Table(rows, colWidths=[45 * mm, 42 * mm, 28 * mm, 55 * mm, 24 * mm, 24 * mm], repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F5B2A")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1D5DB")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )

        doc.build([header_table, Spacer(1, 5 * mm), table])
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="customers_report.pdf"'
        return response

class CustomerDetailView(StaffRequiredMixin, DetailView):
    model = Customer
    template_name = 'transport/customers/detail.html'
    context_object_name = 'customer'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        customer = self.get_object()
        
        # Add tabs data
        tab = self.request.GET.get('tab', 'overview')
        context['current_tab'] = tab
        
        # View mode (grid or list)
        context['view_mode'] = self.request.GET.get('view', 'list')
        
        # Service history filters
        status_filter = self.request.GET.get('status', '')
        search_query = self.request.GET.get('q', '')
        
        # All trips for this customer
        customer_trips = Trip.objects.filter(customer=customer).select_related(
            'route', 'vehicle', 'driver', 'commodity_type'
        )
        
        # Apply filters
        if status_filter:
            customer_trips = customer_trips.filter(status=status_filter)
        if search_query:
            customer_trips = customer_trips.filter(
                Q(order_number__icontains=search_query) |
                Q(route__origin__icontains=search_query) |
                Q(route__destination__icontains=search_query) |
                Q(vehicle__plate_number__icontains=search_query) |
                Q(driver__name__icontains=search_query)
            )
        
        customer_trips = customer_trips.order_by('-created_at')
        
        # Pagination for service history
        page_number = self.request.GET.get('page', 1)
        paginator = Paginator(customer_trips, 12)
        page_obj = paginator.get_page(page_number)
        
        context['service_history'] = page_obj
        context['page_obj'] = page_obj
        context['paginator'] = paginator
        context['is_paginated'] = page_obj.has_other_pages()
        context['status_filter'] = status_filter
        context['search_query'] = search_query
        context['trip_status_choices'] = Trip.TripStatus.choices
        
        # Customer metrics
        all_trips = Trip.objects.filter(customer=customer)
        completed_trips = all_trips.filter(status__in=['DELIVERED', 'CLOSED'])
        active_trips = all_trips.filter(status__in=['ASSIGNED', 'IN_TRANSIT'])
        
        context['customer_metrics'] = {
            'total_trips': all_trips.count(),
            'completed_trips': completed_trips.count(),
            'active_trips': active_trips.count(),
            'total_revenue': completed_trips.aggregate(total=Sum('revenue'))['total'] or 0,
            'avg_trip_value': completed_trips.aggregate(avg=Avg('revenue'))['avg'] or 0,
            'total_distance': completed_trips.aggregate(total=Sum('distance'))['total'] or 0,
            'outstanding_balance': 0,  # TODO: Calculate outstanding payments
        }
        
        # Recent trips (for overview tab)
        context['recent_trips'] = all_trips[:5]
        
        # Last order date
        last_trip = all_trips.first()
        context['last_order_date'] = last_trip.created_at if last_trip else None
        
        # Payment history placeholder
        context['recent_payments'] = []  # TODO: Add payment model and logic
        
        return context

class CustomerCreateView(StaffRequiredMixin, CreateView):
    model = Customer
    form_class = CustomerForm
    template_name = 'transport/customers/create.html'
    success_url = reverse_lazy('transport:customers:list')

    def _is_ajax(self):
        return self.request.headers.get("x-requested-with") == "XMLHttpRequest"

    def get_template_names(self):
        if self.request.GET.get("partial") == "form":
            return ["transport/customers/_modal_form.html"]
        return [self.template_name]

    def form_valid(self, form):
        customer = form.save(commit=False)
        create_account = form.cleaned_data.get('create_account', True)
        email = form.cleaned_data.get('email', '')
        
        customer.email = email
        customer.save()
        
        if create_account and email:
            # Check if a user with this email already exists
            existing_user = AuthUser.objects.filter(email=email.lower()).first()
            
            if existing_user:
                customer.user = existing_user
                customer.save()
                messages.info(
                    self.request,
                    f'Customer "{customer.company_name}" created and linked to existing account ({email}).'
                )
            else:
                password = generate_password()
                
                try:
                    user = AuthUser.objects.create_user(
                        email=email.lower(),
                        full_name=customer.contact_person or customer.company_name,
                        password=password,
                        role=AuthUser.Role.CLIENT,
                        phone=customer.phone or '',
                        must_change_password=True,
                    )
                    
                    customer.user = user
                    customer.save()
                    
                    try:
                        subject = 'Your Client Account Has Been Created'
                        message = (
                            f"Hello {customer.contact_person or customer.company_name},\n\n"
                            f"Your client account has been created in the Transport Management System.\n\n"
                            f"Here are your login credentials:\n"
                            f"  Email: {email}\n"
                            f"  Password: {password}\n\n"
                            f"Company: {customer.company_name}\n"
                            f"Phone: {customer.phone}\n\n"
                            f"Please log in and change your password as soon as possible.\n\n"
                            f"Best regards,\n"
                            f"Transport Management Team"
                        )
                        
                        from accounts.emailing import build_public_url, send_atms_email

                        send_atms_email(
                            subject=subject,
                            to=[email],
                            greeting=f"Hello {customer.company_name}",
                            headline="Customer Portal Account Created",
                            intro="Your customer account has been created in ZALA Terminal. Use the credentials below to access your orders and trips.",
                            details=[
                                {"label": "Email", "value": email},
                                {"label": "Temporary Password", "value": password},
                                {"label": "Company", "value": customer.company_name},
                                {"label": "Phone", "value": customer.phone},
                                {"label": "Portal URL", "value": build_public_url("/customer/login/")},
                            ],
                            note="Please sign in and update your password as soon as possible.",
                            cta_label="Open Customer Login",
                            cta_url=build_public_url("/customer/login/"),
                        )
                        messages.success(
                            self.request,
                            f'Customer "{customer.company_name}" created successfully! '
                            f'Account credentials sent to {email}.'
                        )
                    except Exception as e:
                        messages.warning(
                            self.request,
                            f'Customer "{customer.company_name}" created and account set up, '
                            f'but email could not be sent: {str(e)}. '
                            f'Generated password: {password}'
                        )
                except Exception as e:
                    messages.warning(
                        self.request,
                        f'Customer "{customer.company_name}" created, but account creation failed: {str(e)}'
                    )
        else:
            messages.success(self.request, f'Customer "{customer.company_name}" created successfully (no account created).')

        if self._is_ajax():
            return JsonResponse({
                'success': True,
                'id': customer.pk,
                'detail_url': reverse_lazy('transport:customers:detail', kwargs={'pk': customer.pk}),
                'message': f'Customer "{customer.company_name}" created successfully.',
            })

        return redirect(self.success_url)

    def form_invalid(self, form):
        if self._is_ajax():
            return JsonResponse({
                'success': False,
                'errors': form.errors,
                'non_field_errors': form.non_field_errors(),
            }, status=400)
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['cancel_url'] = reverse_lazy('transport:customers:list')
        context['back_url'] = reverse_lazy('transport:customers:list')
        return context

class CustomerUpdateView(StaffRequiredMixin, UpdateView):
    model = Customer
    form_class = CustomerForm
    template_name = 'transport/customers/edit.html'

    def _is_ajax(self):
        return self.request.headers.get("x-requested-with") == "XMLHttpRequest"

    def get_template_names(self):
        if self.request.GET.get("partial") == "form":
            return ["transport/customers/_modal_form.html"]
        return [self.template_name]
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Remove create_account field on edit
        if 'create_account' in form.fields:
            del form.fields['create_account']
        # Pre-fill email from the model
        if self.object and self.object.email:
            form.fields['email'].initial = self.object.email
        return form
    
    def get_success_url(self):
        return reverse_lazy('transport:customers:detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        if self._is_ajax():
            self.object = form.save()
            return JsonResponse({
                'success': True,
                'id': self.object.pk,
                'detail_url': reverse_lazy('transport:customers:detail', kwargs={'pk': self.object.pk}),
                'message': f'Customer {self.object.company_name} updated successfully!',
            })
        messages.success(self.request, f'Customer {form.instance.company_name} updated successfully!')
        return super().form_valid(form)

    def form_invalid(self, form):
        if self._is_ajax():
            return JsonResponse({
                'success': False,
                'errors': form.errors,
                'non_field_errors': form.non_field_errors(),
            }, status=400)
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['customer'] = self.object
        context['cancel_url'] = reverse_lazy('transport:customers:detail', kwargs={'pk': self.object.pk})
        context['back_url'] = reverse_lazy('transport:customers:detail', kwargs={'pk': self.object.pk})
        return context

@login_required  
def customer_quick_status(request, customer_id):
    """AJAX view to quickly update customer status"""
    if not user_has_role(request.user, SystemGroup.ADMIN, SystemGroup.OPERATIONS_MANAGER, SystemGroup.LOGISTICS_COORDINATOR):
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    customer = get_object_or_404(Customer, id=customer_id)
    new_status = request.POST.get('status')
    
    if new_status in dict(Customer.STATUS_CHOICES):
        customer.status = new_status
        customer.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Customer status updated to {customer.get_status_display()}'
        })
    
    return JsonResponse({'success': False, 'error': 'Invalid status'})
