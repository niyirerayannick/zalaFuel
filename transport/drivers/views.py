# Driver Module Views
from io import BytesIO
from pathlib import Path
import warnings

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib import messages
from django.urls import reverse_lazy
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils.decorators import method_decorator
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings as app_settings
from django.template.loader import render_to_string
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from accounts.rbac import SystemGroup, can_access_fleet, can_manage_fleet, user_has_role
from .models import Driver
from .forms import DriverForm


def _logo_path():
    candidate = Path(__file__).resolve().parents[2] / "static" / "img" / "ZALA Terminal.png"
    return candidate if candidate.exists() else None


def _logo_stream(max_width=900):
    logo = _logo_path()
    if not logo:
        return None

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", PILImage.DecompressionBombWarning)
        img = PILImage.open(logo)
        img.load()
    with img:
        img.thumbnail((max_width, max_width))
        stream = BytesIO()
        img.save(stream, format="PNG", optimize=True)
        stream.seek(0)
        return stream


class StaffRequiredMixin(UserPassesTestMixin):
    """Mixin to require staff access level"""
    def test_func(self):
        return self.request.user.is_authenticated and can_access_fleet(self.request.user)


class FleetWriteRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_authenticated and can_manage_fleet(self.request.user)

class DriverListView(StaffRequiredMixin, ListView):
    model = Driver
    template_name = 'transport/drivers/list.html'
    context_object_name = 'drivers'
    paginate_by = 20

    def get_template_names(self):
        if self.request.GET.get("partial") == "list":
            return ["transport/drivers/_list_content.html"]
        return [self.template_name]

    def get_queryset(self):
        queryset = Driver.objects.all()
        search = self.request.GET.get('search')
        status = self.request.GET.get('status')
        
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(license_number__icontains=search)
            )
        
        if status:
            queryset = queryset.filter(status=status)
        
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Driver statistics
        all_drivers = Driver.objects.all()
        context['total_drivers'] = all_drivers.count()
        context['available_drivers'] = all_drivers.filter(status='AVAILABLE').count()
        context['assigned_drivers'] = all_drivers.filter(status='ASSIGNED').count()
        context['company_drivers'] = all_drivers.filter(work_status=Driver.WorkStatus.COMPANY).count()
        context['external_drivers'] = all_drivers.filter(work_status=Driver.WorkStatus.EXTERNAL).count()
        
        # License alerts (expiring in next 30 days)
        from datetime import timedelta
        next_month = timezone.now().date() + timedelta(days=30)
        context['license_expiring'] = all_drivers.filter(license_expiry__lte=next_month).count()
        
        context['status_choices'] = Driver.STATUS_CHOICES
        context['can_manage_fleet'] = can_manage_fleet(self.request.user)
        return context


class DriverExportMixin(StaffRequiredMixin):
    def get_filtered_queryset(self):
        view = DriverListView()
        view.request = self.request
        return view.get_queryset()


class DriverExcelExportView(DriverExportMixin, ListView):
    def get(self, request, *args, **kwargs):
        queryset = self.get_filtered_queryset()
        filename = "drivers_report.xlsx"
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError:
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="drivers_report.csv"'
            response.write("Name,Phone,Email,License Number,Category,Status,License Expiry\n")
            for driver in queryset:
                response.write(
                    f'{driver.name},{driver.phone},{driver.email or ""},{driver.license_number},{driver.get_license_category_display()},{driver.get_status_display()},{driver.license_expiry}\n'
                )
            return response

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Drivers"

        header_fill = PatternFill(fill_type="solid", fgColor="0F5B2A")
        header_font = Font(color="FFFFFF", bold=True)
        body_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
        alt_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        sheet.merge_cells("A1:G1")
        sheet["A1"] = "ZALA Terminal Driver Report"
        sheet["A1"].font = Font(color="0F5B2A", bold=True, size=16)
        sheet.merge_cells("A2:G2")
        sheet["A2"] = f"Driver register export generated on {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
        sheet["A2"].font = Font(color="475569", italic=True, size=10)

        headers = ["Name", "Phone", "Email", "License Number", "Category", "Status", "License Expiry"]
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=4, column=idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        row_no = 5
        for driver in queryset:
            values = [
                driver.name,
                driver.phone,
                driver.email or "-",
                driver.license_number,
                driver.get_license_category_display(),
                driver.get_status_display(),
                driver.license_expiry,
            ]
            for col_no, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_no, column=col_no, value=value)
                cell.border = thin_border
                cell.fill = body_fill if row_no % 2 == 0 else alt_fill
                cell.alignment = Alignment(vertical="top")
                if col_no == 7 and value:
                    cell.number_format = "yyyy-mm-dd"
            row_no += 1

        sheet.auto_filter.ref = f"A4:G{max(row_no - 1, 4)}"
        sheet.freeze_panes = "A5"
        for column_cells in sheet.columns:
            column_index = column_cells[0].column
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 3, 28)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type=mime_type)
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class DriverPdfExportView(DriverExportMixin, ListView):
    def get(self, request, *args, **kwargs):
        queryset = self.get_filtered_queryset()
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=12 * mm,
            leftMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
        )
        styles = getSampleStyleSheet()

        logo_stream = _logo_stream()
        header_left = []
        if logo_stream:
            header_left.append(Image(logo_stream, width=34 * mm, height=16 * mm))
            header_left.append(Spacer(1, 2 * mm))
        header_left.extend(
            [
                Paragraph("<font color='#0F5B2A'><b>ZALA Terminal Driver Report</b></font>", styles["Title"]),
                Paragraph("Driver register export generated from the ZALA Terminal directory.", styles["Normal"]),
            ]
        )
        header_right = [
            Paragraph("<b>Report</b><br/>Driver Register", styles["Normal"]),
            Spacer(1, 2),
            Paragraph(f"<b>Generated</b><br/>{timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}", styles["Normal"]),
            Spacer(1, 2),
            Paragraph(f"<b>Total Drivers</b><br/>{queryset.count()}", styles["Normal"]),
        ]
        header_table = Table([[header_left, header_right]], colWidths=[170 * mm, 85 * mm])
        header_table.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1E7D7")),
                    ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#EAF7EF")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )

        rows = [["Name", "Phone", "Email", "License Number", "Category", "Status", "Expiry"]]
        for driver in queryset:
            rows.append(
                [
                    driver.name,
                    driver.phone,
                    driver.email or "-",
                    driver.license_number,
                    driver.get_license_category_display(),
                    driver.get_status_display(),
                    driver.license_expiry.strftime("%d/%m/%Y") if driver.license_expiry else "-",
                ]
            )
        if len(rows) == 1:
            rows.append(["-", "-", "-", "-", "-", "-", "-"])

        table = Table(
            rows,
            colWidths=[42 * mm, 28 * mm, 42 * mm, 36 * mm, 34 * mm, 24 * mm, 24 * mm],
            repeatRows=1,
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F5B2A")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1D5DB")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )

        doc.build([header_table, Spacer(1, 5 * mm), table])
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="drivers_report.pdf"'
        return response

class DriverDetailView(StaffRequiredMixin, DetailView):
    model = Driver
    template_name = 'transport/drivers/detail.html'
    context_object_name = 'driver'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        driver = self.get_object()
        
        # Add tabs data
        tab = self.request.GET.get('tab', 'overview')
        context['current_tab'] = tab
        
        # Recent trips ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â real data from Trip model
        from transport.trips.models import Trip
        driver_trips = Trip.objects.filter(driver=driver).select_related(
            'vehicle', 'customer', 'route', 'commodity_type',
        ).order_by('-created_at')
        context['recent_trips'] = driver_trips[:10]
        
        # Performance metrics ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â computed from actual trips
        from django.db.models import Sum, Count, Avg
        stats = driver_trips.aggregate(
            total_trips=Count('id'),
            total_distance=Sum('distance'),
            total_revenue=Sum('revenue'),
            total_cost=Sum('total_cost'),
            total_profit=Sum('profit'),
        )
        completed_trips = driver_trips.filter(
            status__in=[Trip.TripStatus.DELIVERED, Trip.TripStatus.CLOSED],
        ).count()
        context['performance_metrics'] = {
            'total_trips': stats['total_trips'] or 0,
            'total_distance': stats['total_distance'] or 0,
            'total_revenue': stats['total_revenue'] or 0,
            'total_cost': stats['total_cost'] or 0,
            'total_profit': stats['total_profit'] or 0,
            'completed_trips': completed_trips,
            'safety_score': 100,  # TODO: Calculate from incident reports
        }
        
        # Active trip (ASSIGNED or IN_TRANSIT)
        context['active_trip'] = driver_trips.filter(
            status__in=[Trip.TripStatus.ASSIGNED, Trip.TripStatus.IN_TRANSIT],
        ).first()
        
        # License and document alerts
        from datetime import timedelta
        context['alerts'] = []
        if driver.license_expiry:
            days_until_expiry = (driver.license_expiry - timezone.now().date()).days
            if days_until_expiry <= 30:
                context['alerts'].append({
                    'type': 'warning' if days_until_expiry > 7 else 'danger',
                    'message': f'License expires in {days_until_expiry} days'
                })
        
        return context

class DriverCreateView(FleetWriteRequiredMixin, CreateView):
    model = Driver
    form_class = DriverForm
    template_name = 'transport/drivers/create.html'
    success_url = reverse_lazy('transport:drivers:list')

    def _is_ajax(self):
        return self.request.headers.get("x-requested-with") == "XMLHttpRequest"

    def get_template_names(self):
        if self.request.GET.get("partial") == "form":
            return ["transport/drivers/_modal_form.html"]
        return [self.template_name]

    def form_valid(self, form):
        driver = form.save(commit=False)
        email = form.cleaned_data.get('email', '')
        
        driver.email = email
        driver.save()

        messages.success(self.request, f'Driver "{driver.name}" created successfully.')

        if self._is_ajax():
            return JsonResponse({
                'success': True,
                'id': driver.pk,
                'detail_url': reverse_lazy('transport:drivers:detail', kwargs={'pk': driver.pk}),
                'message': f'Driver "{driver.name}" created successfully.',
            })

        return redirect(self.success_url)

    def form_invalid(self, form):
        if self._is_ajax():
            return JsonResponse({
                'success': False,
                'errors': form.errors,
                'non_field_errors': form.non_field_errors(),
            }, status=400)
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['cancel_url'] = reverse_lazy('transport:drivers:list')
        context['back_url'] = reverse_lazy('transport:drivers:list')
        return context

class DriverUpdateView(FleetWriteRequiredMixin, UpdateView):
    model = Driver
    form_class = DriverForm
    template_name = 'transport/drivers/edit.html'

    def _is_ajax(self):
        return self.request.headers.get("x-requested-with") == "XMLHttpRequest"

    def get_template_names(self):
        if self.request.GET.get("partial") == "form":
            return ["transport/drivers/_modal_form.html"]
        return [self.template_name]
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Pre-fill email from the model
        if self.object and self.object.email:
            form.fields['email'].initial = self.object.email
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['driver'] = self.object
        context['cancel_url'] = reverse_lazy('transport:drivers:detail', kwargs={'pk': self.object.pk})
        context['back_url'] = reverse_lazy('transport:drivers:detail', kwargs={'pk': self.object.pk})
        return context

    def get_success_url(self):
        return reverse_lazy('transport:drivers:detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        if self._is_ajax():
            self.object = form.save()
            return JsonResponse({
                'success': True,
                'id': self.object.pk,
                'detail_url': reverse_lazy('transport:drivers:detail', kwargs={'pk': self.object.pk}),
                'message': f'Driver {form.cleaned_data["name"]} updated successfully!',
            })
        messages.success(self.request, f'Driver {form.cleaned_data["name"]} updated successfully!')
        return super().form_valid(form)

    def form_invalid(self, form):
        if self._is_ajax():
            return JsonResponse({
                'success': False,
                'errors': form.errors,
                'non_field_errors': form.non_field_errors(),
            }, status=400)
        return super().form_invalid(form)

@login_required  
def driver_quick_status(request, driver_id):
    """AJAX view to quickly update driver status"""
    if not user_has_role(request.user, SystemGroup.ADMIN, SystemGroup.OPERATIONS_MANAGER, SystemGroup.LOGISTICS_COORDINATOR):
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    driver = get_object_or_404(Driver, id=driver_id)
    new_status = request.POST.get('status')
    
    if new_status in dict(Driver.STATUS_CHOICES):
        driver.status = new_status
        driver.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Driver status updated to {driver.get_status_display()}'
        })
    
    return JsonResponse({'success': False, 'error': 'Invalid status'})
