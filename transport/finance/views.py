import json
import csv
from datetime import timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from collections import OrderedDict
import warnings

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.core.exceptions import PermissionDenied
from django.core import signing
from django.core.paginator import Paginator
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.db.models import Avg, Count, Max, Q, Sum
from django.db.models.functions import TruncMonth
from django.urls import reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import CreateView, DetailView, ListView, TemplateView, UpdateView
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from transport.customers.models import Customer
from transport.drivers.models import Driver
from transport.fuel.models import FuelRequest
from transport.maintenance.models import MaintenanceRecord
from transport.trips.models import Trip
from transport.vehicles.models import Vehicle
from accounts.emailing import approval_recipients, build_public_url, send_atms_email
from accounts.rbac import (
    SystemGroup,
    can_access_finance,
    can_manage_finance,
    user_has_role,
)

from .forms import DriverAllowanceForm, DriverFeeForm, ExpenseForm, PaymentForm
from .models import DriverAllowance, DriverFee, Expense, Payment
from .services import build_invoice_context, render_invoice_pdf


AFRILOTT_GREEN = "0F5B2A"
AFRILOTT_LIGHT_GREEN = "EAF7EF"


def _finance_logo_path():
    candidate = Path(__file__).resolve().parents[2] / "static" / "img" / "ZALA/ECO ENERGY.png"
    return candidate if candidate.exists() else None


def _finance_logo_stream(max_width=900):
    logo = _finance_logo_path()
    if not logo:
        return None

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", PILImage.DecompressionBombWarning)
        img = PILImage.open(logo)
        img.load()
    with img:
        img.thumbnail((max_width, max_width))
        stream = BytesIO()
        img.save(stream, format="PNG", optimize=True)
        stream.seek(0)
        return stream


def canonical_payment_queryset(queryset=None):
    base_queryset = queryset if queryset is not None else Payment.objects.all()
    return base_queryset.exclude(
        trip__isnull=True,
        order__isnull=False,
        notes__icontains="Auto-generated from trip workflow.",
    )


class StaffRequiredMixin(UserPassesTestMixin):
    """Mixin to require staff access level."""

    def test_func(self):
        return self.request.user.is_authenticated and can_access_finance(self.request.user)


class FinanceWriteAccessMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_authenticated and can_manage_finance(self.request.user)


def restrict_payments_for_user(queryset, user):
    if user_has_role(user, SystemGroup.CUSTOMER):
        customer = getattr(user, "customer_profile", None)
        if customer is None:
            return queryset.none()
        return queryset.filter(
            Q(customer=customer)
            | Q(trip__customer=customer)
            | Q(order__customer=customer)
        ).distinct()
    return queryset


class FinanceModalFormMixin:
    """Shared HTMX modal behavior for finance create forms."""

    modal_template_name = "transport/finance/_modal_form.html"
    modal_title = ""
    modal_subtitle = ""
    modal_eyebrow = "Finance"
    form_partial = ""
    submit_label = "Save"
    submit_icon = "save"
    success_message = "Saved successfully."

    def is_htmx(self):
        return self.request.headers.get("HX-Request") == "true"

    def get_template_names(self):
        if self.is_htmx():
            return [self.modal_template_name]
        return [self.template_name]

    def get_htmx_source(self):
        return self.request.GET.get("source") or self.request.POST.get("source")

    def get_htmx_success_url(self):
        return self.get_success_url()

    def get_htmx_refresh_url(self):
        return ""

    def get_htmx_refresh_target(self):
        return ""

    def get_success_url(self):
        raise NotImplementedError

    def get_modal_context(self):
        return {
            "modal_title": self.modal_title,
            "modal_subtitle": self.modal_subtitle,
            "modal_eyebrow": self.modal_eyebrow,
            "modal_icon": getattr(self, "modal_icon", "account_balance"),
            "form_partial": self.form_partial,
            "form_action": self.request.get_full_path(),
            "submit_label": self.submit_label,
            "submit_icon": self.submit_icon,
            "modal_width_class": "max-w-4xl",
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(self.get_modal_context())
        return context

    def form_invalid(self, form):
        response = super().form_invalid(form)
        if self.is_htmx():
            response.status_code = 422
        return response

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, self.success_message)
        if self.is_htmx():
            headers = {
                "HX-Trigger": json.dumps({"showToast": {"message": self.success_message, "level": "success"}}),
            }
            if self.get_htmx_source() == "list":
                headers["HX-Trigger"] = json.dumps({
                    "showToast": {"message": self.success_message, "level": "success"},
                    "financeListRefresh": {
                        "url": str(self.get_htmx_refresh_url()),
                        "target": self.get_htmx_refresh_target(),
                    },
                })
            else:
                headers["HX-Redirect"] = str(self.get_htmx_success_url())
            return HttpResponse(status=204, headers=headers)
        return response


class FinanceOverviewView(StaffRequiredMixin, ListView):
    """Central financial dashboard for transport operations."""

    model = Payment
    template_name = "transport/finance/overview.html"
    context_object_name = "payments"
    paginate_by = 10

    def _apply_date_range(self, queryset, lookup):
        date_from = self.request.GET.get("date_from")
        date_to = self.request.GET.get("date_to")
        if date_from:
            queryset = queryset.filter(**{f"{lookup}__gte": date_from})
        if date_to:
            queryset = queryset.filter(**{f"{lookup}__lte": date_to})
        return queryset

    def _selected_filters(self):
        return {
            "vehicle": self.request.GET.get("vehicle", ""),
            "driver": self.request.GET.get("driver", ""),
            "customer": self.request.GET.get("customer", ""),
            "date_from": self.request.GET.get("date_from", ""),
            "date_to": self.request.GET.get("date_to", ""),
        }

    def _base_trip_queryset(self):
        filters = self._selected_filters()
        queryset = Trip.objects.select_related("customer", "vehicle", "driver", "route")
        queryset = self._apply_date_range(queryset, "created_at__date")
        if filters["vehicle"]:
            queryset = queryset.filter(vehicle_id=filters["vehicle"])
        if filters["driver"]:
            queryset = queryset.filter(driver_id=filters["driver"])
        if filters["customer"]:
            queryset = queryset.filter(customer_id=filters["customer"])
        return queryset

    def _manual_payment_queryset(self):
        filters = self._selected_filters()
        queryset = canonical_payment_queryset(Payment.objects.select_related("trip", "customer", "order"))
        queryset = self._apply_date_range(queryset, "payment_date")
        if filters["customer"]:
            queryset = queryset.filter(Q(customer_id=filters["customer"]) | Q(trip__customer_id=filters["customer"]))
        if filters["vehicle"]:
            queryset = queryset.filter(trip__vehicle_id=filters["vehicle"])
        if filters["driver"]:
            queryset = queryset.filter(trip__driver_id=filters["driver"])
        return queryset

    def _expense_queryset(self):
        filters = self._selected_filters()
        queryset = Expense.objects.select_related("trip", "vehicle")
        queryset = self._apply_date_range(queryset, "expense_date")
        if filters["vehicle"]:
            queryset = queryset.filter(Q(vehicle_id=filters["vehicle"]) | Q(trip__vehicle_id=filters["vehicle"]))
        if filters["driver"]:
            queryset = queryset.filter(trip__driver_id=filters["driver"])
        if filters["customer"]:
            queryset = queryset.filter(trip__customer_id=filters["customer"])
        return queryset

    def _driver_fee_queryset(self):
        filters = self._selected_filters()
        queryset = DriverFee.objects.select_related("trip", "driver")
        queryset = self._apply_date_range(queryset, "fee_date")
        if filters["driver"]:
            queryset = queryset.filter(driver_id=filters["driver"])
        if filters["vehicle"]:
            queryset = queryset.filter(trip__vehicle_id=filters["vehicle"])
        if filters["customer"]:
            queryset = queryset.filter(trip__customer_id=filters["customer"])
        return queryset

    def _fuel_queryset(self):
        filters = self._selected_filters()
        queryset = FuelRequest.objects.filter(is_approved=True).select_related("trip", "station")
        queryset = self._apply_date_range(queryset, "created_at__date")
        if filters["vehicle"]:
            queryset = queryset.filter(trip__vehicle_id=filters["vehicle"])
        if filters["driver"]:
            queryset = queryset.filter(trip__driver_id=filters["driver"])
        if filters["customer"]:
            queryset = queryset.filter(trip__customer_id=filters["customer"])
        return queryset

    def _maintenance_queryset(self):
        filters = self._selected_filters()
        queryset = MaintenanceRecord.objects.select_related("vehicle", "trip").filter(status=MaintenanceRecord.Status.APPROVED)
        queryset = self._apply_date_range(queryset, "service_date")
        if filters["vehicle"]:
            queryset = queryset.filter(vehicle_id=filters["vehicle"])
        if filters["driver"]:
            queryset = queryset.filter(trip__driver_id=filters["driver"])
        if filters["customer"]:
            queryset = queryset.filter(trip__customer_id=filters["customer"])
        return queryset

    def get_queryset(self):
        return self._manual_payment_queryset().order_by("-payment_date", "-created_at")[:10]

    def build_overview_context(self):
        completed_trips = self._base_trip_queryset().filter(status__in=[Trip.TripStatus.DELIVERED, Trip.TripStatus.CLOSED])
        revenue_entries = self._manual_payment_queryset()
        trip_revenue = revenue_entries.filter(trip__isnull=False).aggregate(total=Sum("amount"))["total"] or Decimal("0")
        manual_revenue = revenue_entries.filter(trip__isnull=True).aggregate(total=Sum("amount"))["total"] or Decimal("0")
        total_revenue = trip_revenue + manual_revenue
        collected_payments = sum((payment.collected_amount or Decimal("0")) for payment in revenue_entries)
        pending_revenue = sum((payment.outstanding_amount or Decimal("0")) for payment in revenue_entries)

        expense_entries = self._expense_queryset()
        driver_fee_entries = self._driver_fee_queryset()
        fuel_entries = self._fuel_queryset()
        maintenance_entries = self._maintenance_queryset()

        other_expenses = expense_entries.aggregate(total=Sum("amount"))["total"] or Decimal("0")
        driver_fees = driver_fee_entries.aggregate(total=Sum("amount"))["total"] or Decimal("0")
        fuel_expenses = fuel_entries.aggregate(total=Sum("amount"))["total"] or Decimal("0")
        maintenance_cost = maintenance_entries.aggregate(total=Sum("cost"))["total"] or Decimal("0")
        total_expenses = other_expenses + driver_fees + fuel_expenses + maintenance_cost
        net_profit = total_revenue - total_expenses
        profit_margin_percentage = (net_profit / total_revenue * 100) if total_revenue else Decimal("0")
        expense_coverage_ratio = (collected_payments / total_expenses * 100) if total_expenses else Decimal("0")
        operating_gap = collected_payments - total_expenses
        summary_title = "Stable financial position"
        summary_tone = "emerald"
        summary_note = "Collections are covering operating costs and the current margin remains healthy."
        if operating_gap < 0 or net_profit < 0:
            summary_title = "Attention required"
            summary_tone = "rose"
            summary_note = "Costs are ahead of collections or profitability is negative in the current filter range."
        elif pending_revenue > collected_payments:
            summary_title = "Watch collections"
            summary_tone = "amber"
            summary_note = "A large share of revenue is still outstanding, so follow-up on collections is important."

        summary_points = [
            {
                "label": "Collection Strength",
                "value": f"{expense_coverage_ratio:.1f}%",
                "hint": "Collected payments compared with total expenses.",
            },
            {
                "label": "Outstanding Revenue",
                "value": f"{pending_revenue:,.2f}",
                "hint": "Revenue not yet collected in this view.",
            },
            {
                "label": "Operating Gap",
                "value": f"{operating_gap:,.2f}",
                "hint": "Collected payments minus total expenses.",
            },
        ]
        payment_status_breakdown = [
            ("Paid", revenue_entries.filter(status=Payment.Status.PAID).count()),
            ("Partial", revenue_entries.filter(status=Payment.Status.PARTIAL).count()),
            ("Pending", revenue_entries.filter(status=Payment.Status.PENDING).count()),
            ("Failed", revenue_entries.filter(status=Payment.Status.FAILED).count()),
        ]

        monthly_start = timezone.localdate().replace(day=1)
        period_start = monthly_start - timedelta(days=150)
        monthly_trip_revenue = (
            revenue_entries.filter(trip__isnull=False, payment_date__gte=period_start)
            .annotate(period=TruncMonth("payment_date"))
            .values("period")
            .annotate(total=Sum("amount"))
            .order_by("period")
        )
        monthly_manual_revenue = (
            revenue_entries.filter(trip__isnull=True, payment_date__gte=period_start)
            .annotate(period=TruncMonth("payment_date"))
            .values("period")
            .annotate(total=Sum("amount"))
            .order_by("period")
        )
        monthly_other_expenses = (
            expense_entries.filter(expense_date__gte=period_start)
            .annotate(period=TruncMonth("expense_date"))
            .values("period")
            .annotate(total=Sum("amount"))
            .order_by("period")
        )
        monthly_driver_fees = (
            driver_fee_entries.filter(fee_date__gte=period_start)
            .annotate(period=TruncMonth("fee_date"))
            .values("period")
            .annotate(total=Sum("amount"))
            .order_by("period")
        )
        monthly_maintenance = (
            maintenance_entries.filter(service_date__gte=period_start)
            .annotate(period=TruncMonth("service_date"))
            .values("period")
            .annotate(total=Sum("cost"))
            .order_by("period")
        )
        monthly_fuel = (
            fuel_entries.filter(created_at__date__gte=period_start)
            .annotate(period=TruncMonth("created_at"))
            .values("period")
            .annotate(total=Sum("amount"))
            .order_by("period")
        )

        monthly_map = {}
        for row in monthly_trip_revenue:
            monthly_map.setdefault(row["period"], {"revenue": Decimal("0"), "expenses": Decimal("0")})
            monthly_map[row["period"]]["revenue"] += row["total"] or Decimal("0")
        for queryset in [monthly_manual_revenue]:
            for row in queryset:
                monthly_map.setdefault(row["period"], {"revenue": Decimal("0"), "expenses": Decimal("0")})
                monthly_map[row["period"]]["revenue"] += row["total"] or Decimal("0")
        for queryset in [monthly_other_expenses, monthly_driver_fees, monthly_maintenance, monthly_fuel]:
            for row in queryset:
                monthly_map.setdefault(row["period"], {"revenue": Decimal("0"), "expenses": Decimal("0")})
                monthly_map[row["period"]]["expenses"] += row["total"] or Decimal("0")

        monthly_labels = []
        monthly_revenue_series = []
        monthly_expense_series = []
        monthly_profit_series = []
        for period in sorted(monthly_map.keys(), key=lambda value: value.strftime("%Y-%m-%d")):
            monthly_labels.append(period.strftime("%b %Y"))
            monthly_revenue_series.append(float(monthly_map[period]["revenue"]))
            monthly_expense_series.append(float(monthly_map[period]["expenses"]))
            monthly_profit_series.append(float(monthly_map[period]["revenue"] - monthly_map[period]["expenses"]))

        expense_breakdown = {
            "labels": json.dumps(["Driver Fees", "Fuel", "Maintenance", "Other Expenses"]),
            "values": json.dumps([
                float(driver_fees),
                float(fuel_expenses),
                float(maintenance_cost),
                float(other_expenses),
            ]),
        }

        trip_profitability = []
        for trip in completed_trips[:8]:
            trip_driver_fees = driver_fee_entries.filter(trip=trip).aggregate(total=Sum("amount"))["total"] or Decimal("0")
            trip_expense_total = trip.total_cost + trip_driver_fees
            trip_profitability.append({
                "trip": trip,
                "revenue": trip.revenue or Decimal("0"),
                "fuel_cost": trip.fuel_cost or Decimal("0"),
                "other_expenses": trip.other_expenses or Decimal("0"),
                "driver_fee": trip_driver_fees,
                "profit": (trip.revenue or Decimal("0")) - trip_expense_total,
            })

        vehicles_by_revenue = list(
            completed_trips.values("vehicle__plate_number")
            .annotate(revenue=Sum("revenue"), trips=Count("id"))
            .order_by("-revenue")[:5]
        )
        driver_payment_summary = list(
            driver_fee_entries.values("driver__name", "payment_status")
            .annotate(total=Sum("amount"), entries=Count("id"))
            .order_by("driver__name", "payment_status")[:8]
        )

        return {
            "page_title": "Finance Control Center",
            "total_revenue": total_revenue,
            "trip_revenue": trip_revenue,
            "manual_revenue": manual_revenue,
            "collected_payments": collected_payments,
            "pending_revenue": pending_revenue,
            "total_expenses": total_expenses,
            "driver_fees": driver_fees,
            "fuel_expenses": fuel_expenses,
            "maintenance_cost": maintenance_cost,
            "other_expenses": other_expenses,
            "net_profit": net_profit,
            "profit_margin_percentage": profit_margin_percentage,
            "expense_coverage_ratio": expense_coverage_ratio,
            "operating_gap": operating_gap,
            "financial_summary_title": summary_title,
            "financial_summary_tone": summary_tone,
            "financial_summary_note": summary_note,
            "financial_summary_points": summary_points,
            "payment_count": revenue_entries.count(),
            "expense_count": expense_entries.count(),
            "driver_fee_count": driver_fee_entries.count(),
            "allowance_request_count": DriverAllowance.objects.filter(status=DriverAllowance.Status.PENDING).count(),
            "recent_expenses": expense_entries.order_by("-expense_date", "-created_at")[:8],
            "recent_driver_fees": driver_fee_entries.order_by("-fee_date", "-created_at")[:8],
            "avg_payment": revenue_entries.aggregate(Avg("amount"))["amount__avg"] or 0,
            "avg_expense": expense_entries.aggregate(Avg("amount"))["amount__avg"] or 0,
            "vehicles": Vehicle.objects.order_by("plate_number"),
            "drivers": Driver.objects.order_by("name"),
            "customers": Customer.objects.order_by("company_name"),
            "can_manage_finance": can_manage_finance(self.request.user),
            "selected_vehicle": self.request.GET.get("vehicle", ""),
            "selected_driver": self.request.GET.get("driver", ""),
            "selected_customer": self.request.GET.get("customer", ""),
            "selected_date_from": self.request.GET.get("date_from", ""),
            "selected_date_to": self.request.GET.get("date_to", ""),
            "monthly_labels": json.dumps(monthly_labels),
            "monthly_revenue_series": json.dumps(monthly_revenue_series),
            "monthly_expense_series": json.dumps(monthly_expense_series),
            "monthly_profit_series": json.dumps(monthly_profit_series),
            "payment_expense_comparison_labels": json.dumps(["Collected Payments", "Pending Revenue", "Total Expenses"]),
            "payment_expense_comparison_values": json.dumps([float(collected_payments), float(pending_revenue), float(total_expenses)]),
            "expense_breakdown_labels": expense_breakdown["labels"],
            "expense_breakdown_values": expense_breakdown["values"],
            "payment_status_labels": json.dumps([label for label, _count in payment_status_breakdown]),
            "payment_status_values": json.dumps([count for _label, count in payment_status_breakdown]),
            "trip_profitability": trip_profitability,
            "vehicles_by_revenue": vehicles_by_revenue,
            "driver_payment_summary": driver_payment_summary,
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(self.build_overview_context())
        return context


class FinanceDashboardView(FinanceOverviewView):
    template_name = "transport/finance/dashboard.html"

    def _selected_filters(self):
        return {
            "vehicle": "",
            "driver": "",
            "customer": "",
            "date_from": "",
            "date_to": "",
        }

    def get_queryset(self):
        return canonical_payment_queryset(Payment.objects.none())


class FinanceAnalysisExportMixin(StaffRequiredMixin):
    def _overview_context(self):
        view = FinanceOverviewView()
        view.request = self.request
        view.object_list = view.get_queryset()
        return view.build_overview_context()

    def _monthly_rows(self, context):
        labels = json.loads(context["monthly_labels"])
        revenue = json.loads(context["monthly_revenue_series"])
        expenses = json.loads(context["monthly_expense_series"])
        profit = json.loads(context["monthly_profit_series"])
        rows = []
        for label, revenue_value, expense_value, profit_value in zip(labels, revenue, expenses, profit):
            rows.append(
                {
                    "period": label,
                    "revenue": Decimal(str(revenue_value)),
                    "expenses": Decimal(str(expense_value)),
                    "profit": Decimal(str(profit_value)),
                }
            )
        return rows


class FinanceAnalysisExcelExportView(FinanceAnalysisExportMixin, View):
    def get(self, request, *args, **kwargs):
        context = self._overview_context()
        monthly_rows = self._monthly_rows(context)
        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError:
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="finance_analysis.csv"'
            writer = csv.writer(response)
            writer.writerow(["Metric", "Value"])
            writer.writerow(["Total Revenue", context["total_revenue"]])
            writer.writerow(["Collected Payments", context["collected_payments"]])
            writer.writerow(["Pending Revenue", context["pending_revenue"]])
            writer.writerow(["Total Expenses", context["total_expenses"]])
            writer.writerow(["Net Profit", context["net_profit"]])
            writer.writerow(["Expense Coverage Ratio %", round(context["expense_coverage_ratio"], 2)])
            writer.writerow([])
            writer.writerow(["Month", "Revenue", "Expenses", "Profit"])
            for row in monthly_rows:
                writer.writerow([row["period"], row["revenue"], row["expenses"], row["profit"]])
            return response

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Finance Analysis"
        green_fill = PatternFill(fill_type="solid", fgColor=AFRILOTT_GREEN)
        light_fill = PatternFill(fill_type="solid", fgColor=AFRILOTT_LIGHT_GREEN)
        white_font = Font(color="FFFFFF", bold=True, size=12)
        dark_font = Font(color="0F172A", bold=True)
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        logo_stream = _finance_logo_stream()
        if logo_stream:
            logo_image = XLImage(logo_stream)
            logo_image.width = 170
            logo_image.height = 60
            sheet.add_image(logo_image, "A1")

        sheet.merge_cells("C1:F1")
        sheet["C1"] = "ZALA/ECO ENERGY Financial Analysis"
        sheet["C1"].font = Font(color=AFRILOTT_GREEN, bold=True, size=16)
        sheet["C1"].alignment = Alignment(horizontal="left", vertical="center")
        sheet.merge_cells("C2:F2")
        sheet["C2"] = "Filtered finance overview comparing payments, expenses, and profitability."
        sheet["C2"].font = Font(color="475569", italic=True, size=10)

        metrics = [
            ("Total Revenue", context["total_revenue"]),
            ("Collected Payments", context["collected_payments"]),
            ("Pending Revenue", context["pending_revenue"]),
            ("Total Expenses", context["total_expenses"]),
            ("Net Profit", context["net_profit"]),
            ("Expense Coverage Ratio %", context["expense_coverage_ratio"]),
            ("Operating Gap", context["operating_gap"]),
        ]
        metric_row = 5
        for label, value in metrics:
            sheet.cell(row=metric_row, column=1, value=label).font = dark_font
            sheet.cell(row=metric_row, column=1).fill = light_fill
            sheet.cell(row=metric_row, column=1).border = thin_border
            number_value = float(value) if isinstance(value, Decimal) else value
            sheet.cell(row=metric_row, column=2, value=number_value)
            sheet.cell(row=metric_row, column=2).border = thin_border
            if isinstance(value, Decimal):
                sheet.cell(row=metric_row, column=2).number_format = "#,##0.00"
            metric_row += 1

        header_row = 14
        headers = ["Month", "Revenue", "Expenses", "Profit"]
        for index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=index, value=header)
            cell.fill = green_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        current_row = header_row + 1
        for row in monthly_rows:
            values = [row["period"], float(row["revenue"]), float(row["expenses"]), float(row["profit"])]
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row=current_row, column=column, value=value)
                cell.border = thin_border
                if column > 1:
                    cell.number_format = "#,##0.00"
            current_row += 1

        for column_index, column_cells in enumerate(sheet.columns, start=1):
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 3, 28)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="finance_analysis.xlsx"'
        workbook.save(response)
        return response


class FinanceAnalysisPdfExportView(FinanceAnalysisExportMixin, View):
    def get(self, request, *args, **kwargs):
        context = self._overview_context()
        monthly_rows = self._monthly_rows(context)
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=12 * mm,
            leftMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
        )
        styles = getSampleStyleSheet()
        logo_stream = _finance_logo_stream()

        header_left = []
        if logo_stream:
            header_left.append(Image(logo_stream, width=34 * mm, height=16 * mm))
            header_left.append(Spacer(1, 2 * mm))
        header_left.extend(
            [
                Paragraph("<font color='#0F5B2A'><b>ZALA/ECO ENERGY Financial Analysis</b></font>", styles["Title"]),
                Paragraph("Filtered finance overview comparing payments, expenses, and profitability.", styles["Normal"]),
            ]
        )

        header_right = [
            Paragraph("<b>Generated</b><br/>" + timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M"), styles["Normal"]),
            Spacer(1, 2),
            Paragraph(f"<b>Total Revenue</b><br/><font color='#0F5B2A'>{context['total_revenue']:,.2f}</font>", styles["Normal"]),
            Spacer(1, 2),
            Paragraph(f"<b>Total Expenses</b><br/><font color='#B91C1C'>{context['total_expenses']:,.2f}</font>", styles["Normal"]),
        ]
        header_table = Table([[header_left, header_right]], colWidths=[175 * mm, 80 * mm])
        header_table.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1E7D7")),
                    ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#EAF7EF")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                ]
            )
        )

        summary_rows = [
            ["Metric", "Value"],
            ["Collected Payments", f"{context['collected_payments']:,.2f}"],
            ["Pending Revenue", f"{context['pending_revenue']:,.2f}"],
            ["Net Profit", f"{context['net_profit']:,.2f}"],
            ["Expense Coverage Ratio", f"{context['expense_coverage_ratio']:.1f}%"],
            ["Operating Gap", f"{context['operating_gap']:,.2f}"],
        ]
        summary_table = Table(summary_rows, colWidths=[65 * mm, 45 * mm], repeatRows=1)
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{AFRILOTT_GREEN}")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#cbd5e1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ]
            )
        )

        monthly_table_rows = [["Month", "Revenue", "Expenses", "Profit"]]
        for row in monthly_rows:
            monthly_table_rows.append([row["period"], f"{row['revenue']:,.2f}", f"{row['expenses']:,.2f}", f"{row['profit']:,.2f}"])
        monthly_table = Table(monthly_table_rows, colWidths=[42 * mm, 34 * mm, 34 * mm, 34 * mm], repeatRows=1)
        monthly_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{AFRILOTT_GREEN}")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#cbd5e1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ]
            )
        )

        story = [
            header_table,
            Spacer(1, 10),
            Paragraph("<b>Financial Analysis Summary</b>", styles["Heading2"]),
            Spacer(1, 4),
            summary_table,
            Spacer(1, 10),
            Paragraph("<b>Monthly Revenue vs Expenses</b>", styles["Heading2"]),
            Spacer(1, 4),
            monthly_table,
        ]
        doc.build(story)

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="finance_analysis.pdf"'
        response.write(buffer.getvalue())
        return response


class PaymentListView(StaffRequiredMixin, ListView):
    model = Payment
    template_name = "transport/finance/payments/list.html"
    partial_template_name = "transport/finance/payments/_list_content.html"
    context_object_name = "payments"
    paginate_by = 20

    def get_base_queryset(self):
        queryset = canonical_payment_queryset(
            Payment.objects.select_related("trip", "order", "customer", "trip__customer").all()
        )
        return restrict_payments_for_user(queryset, self.request.user)

    def get_template_names(self):
        if self.request.GET.get("partial") == "list" and (
            self.request.headers.get("HX-Request") == "true"
            or self.request.headers.get("X-Requested-With") == "XMLHttpRequest"
        ):
            return [self.partial_template_name]
        return [self.template_name]

    def get_queryset(self):
        queryset = self.get_base_queryset()
        date_from = self.request.GET.get("date_from")
        date_to = self.request.GET.get("date_to")
        min_amount = self.request.GET.get("min_amount")
        max_amount = self.request.GET.get("max_amount")
        payment_method = self.request.GET.get("payment_method")
        search = self.request.GET.get("search")

        if date_from:
            queryset = queryset.filter(payment_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(payment_date__lte=date_to)
        if min_amount:
            queryset = queryset.filter(amount__gte=min_amount)
        if max_amount:
            queryset = queryset.filter(amount__lte=max_amount)
        if payment_method:
            queryset = queryset.filter(payment_method=payment_method)
        if search:
            queryset = queryset.filter(
                Q(trip__order_number__icontains=search)
                | Q(reference__icontains=search)
                | Q(customer__company_name__icontains=search)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        revenues = self.get_queryset()
        total_amount = sum((payment.amount or Decimal("0")) for payment in revenues)
        paid_revenue_total = sum((payment.collected_amount or Decimal("0")) for payment in revenues)
        pending_revenue_total = sum((payment.outstanding_amount or Decimal("0")) for payment in revenues)
        context.update({
            "total_payments": revenues.count(),
            "total_amount": total_amount,
            "paid_revenue_total": paid_revenue_total,
            "pending_revenue_total": pending_revenue_total,
            "can_manage_finance": can_manage_finance(self.request.user),
            "is_customer": user_has_role(self.request.user, SystemGroup.CUSTOMER),
            "payment_methods": Payment.PaymentMethod.choices,
            "current_date_from": self.request.GET.get("date_from"),
            "current_date_to": self.request.GET.get("date_to"),
            "current_min_amount": self.request.GET.get("min_amount"),
            "current_max_amount": self.request.GET.get("max_amount"),
            "current_payment_method": self.request.GET.get("payment_method"),
            "current_search": self.request.GET.get("search"),
        })
        return context

    def test_func(self):
        return self.request.user.is_authenticated and (
            can_access_finance(self.request.user) or user_has_role(self.request.user, SystemGroup.CUSTOMER)
        )


class PaymentExportMixin(StaffRequiredMixin):
    def get_filtered_queryset(self):
        view = PaymentListView()
        view.request = self.request
        return view.get_queryset().order_by("-payment_date", "-created_at")

    def _payment_customer_name(self, payment):
        if payment.customer:
            return payment.customer.company_name
        if payment.trip and payment.trip.customer:
            return payment.trip.customer.company_name
        return "Unassigned"

    def _payment_trip_or_order(self, payment):
        if payment.trip:
            return payment.trip.order_number
        if payment.order:
            return payment.order.order_number
        return "Manual Revenue"

    def _export_rows(self):
        rows = []
        for payment in self.get_filtered_queryset():
            rows.append(
                {
                    "date": payment.payment_date,
                    "trip_or_order": self._payment_trip_or_order(payment),
                    "customer": self._payment_customer_name(payment),
                    "status": payment.get_status_display(),
                    "amount": payment.amount,
                    "method": payment.get_payment_method_display(),
                    "reference": payment.reference or "-",
                }
            )
        return rows

    def _total_amount(self, rows):
        return sum((row["amount"] or Decimal("0")) for row in rows)


class PaymentExcelExportView(PaymentExportMixin, View):
    def get(self, request, *args, **kwargs):
        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError:
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="revenue_report.csv"'
            writer = csv.writer(response)
            writer.writerow(["Date", "Trip / Order", "Customer", "Status", "Amount", "Method", "Reference"])
            rows = self._export_rows()
            for row in rows:
                writer.writerow(
                    [
                        row["date"].strftime("%Y-%m-%d") if row["date"] else "",
                        row["trip_or_order"],
                        row["customer"],
                        row["status"],
                        row["amount"],
                        row["method"],
                        row["reference"],
                    ]
                )
            writer.writerow([])
            writer.writerow(["", "", "", "Total Amount", self._total_amount(rows), "", ""])
            return response

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Revenue Report"
        rows = self._export_rows()
        total_amount = self._total_amount(rows)
        green_fill = PatternFill(fill_type="solid", fgColor=AFRILOTT_GREEN)
        light_fill = PatternFill(fill_type="solid", fgColor=AFRILOTT_LIGHT_GREEN)
        white_font = Font(color="FFFFFF", bold=True, size=12)
        dark_font = Font(color="0F172A", bold=True)
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        logo_stream = _finance_logo_stream()
        if logo_stream:
            logo_image = XLImage(logo_stream)
            logo_image.width = 170
            logo_image.height = 60
            sheet.add_image(logo_image, "A1")

        sheet.merge_cells("C1:G1")
        sheet["C1"] = "ZALA/ECO ENERGY Revenue Financial Report"
        sheet["C1"].font = Font(color=AFRILOTT_GREEN, bold=True, size=16)
        sheet["C1"].alignment = Alignment(horizontal="left", vertical="center")

        sheet.merge_cells("C2:G2")
        sheet["C2"] = "Generated from the finance payments report with active filters applied."
        sheet["C2"].font = Font(color="475569", italic=True, size=10)
        sheet["C2"].alignment = Alignment(horizontal="left", vertical="center")

        sheet["A5"] = "Total Amount"
        sheet["A5"].font = dark_font
        sheet["A5"].fill = light_fill
        sheet["B5"] = float(total_amount)
        sheet["B5"].font = Font(color=AFRILOTT_GREEN, bold=True)
        sheet["B5"].fill = light_fill
        sheet["B5"].number_format = '#,##0.00'

        headers = ["Date", "Trip / Order", "Customer", "Status", "Amount", "Method", "Reference"]
        header_row = 7
        for index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=index, value=header)
            cell.fill = green_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        current_row = header_row + 1
        for row in rows:
            values = [
                row["date"].strftime("%Y-%m-%d") if row["date"] else "",
                row["trip_or_order"],
                row["customer"],
                row["status"],
                float(row["amount"] or 0),
                row["method"],
                row["reference"],
            ]
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row=current_row, column=column, value=value)
                cell.alignment = Alignment(vertical="top")
                cell.border = thin_border
                if column == 5:
                    cell.number_format = '#,##0.00'
            current_row += 1

        total_row = current_row + 1
        sheet.cell(row=total_row, column=4, value="Total Amount").font = dark_font
        sheet.cell(row=total_row, column=4).fill = light_fill
        sheet.cell(row=total_row, column=4).border = thin_border
        sheet.cell(row=total_row, column=5, value=float(total_amount)).font = Font(color=AFRILOTT_GREEN, bold=True)
        sheet.cell(row=total_row, column=5).fill = light_fill
        sheet.cell(row=total_row, column=5).border = thin_border
        sheet.cell(row=total_row, column=5).number_format = '#,##0.00'

        for column_index, column_cells in enumerate(sheet.columns, start=1):
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 28)

        sheet.freeze_panes = "A8"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="revenue_report.xlsx"'
        workbook.save(response)
        return response


class PaymentPdfExportView(PaymentExportMixin, View):
    def get(self, request, *args, **kwargs):
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=12 * mm,
            leftMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
        )
        styles = getSampleStyleSheet()
        rows_data = self._export_rows()
        total_amount = self._total_amount(rows_data)
        logo_stream = _finance_logo_stream()

        header_left = []
        if logo_stream:
            header_left.append(Image(logo_stream, width=34 * mm, height=16 * mm))
            header_left.append(Spacer(1, 2 * mm))
        header_left.extend(
            [
                Paragraph("<font color='#0F5B2A'><b>ZALA/ECO ENERGY Revenue Financial Report</b></font>", styles["Title"]),
                Paragraph("Generated from the finance payments report with active filters applied.", styles["Normal"]),
            ]
        )

        header_right = [
            Paragraph("<b>Report</b><br/>Revenue Payments", styles["Normal"]),
            Spacer(1, 2),
            Paragraph(f"<b>Generated</b><br/>{timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}", styles["Normal"]),
            Spacer(1, 2),
            Paragraph(f"<b>Total Amount</b><br/><font color='#0F5B2A'>{total_amount:,.2f}</font>", styles["Normal"]),
        ]

        header_table = Table([[header_left, header_right]], colWidths=[170 * mm, 85 * mm])
        header_table.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1E7D7")),
                    ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#EAF7EF")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                ]
            )
        )

        rows = [["Date", "Trip / Order", "Customer", "Status", "Amount", "Method", "Reference"]]
        for row in rows_data:
            rows.append(
                [
                    row["date"].strftime("%d/%m/%Y") if row["date"] else "",
                    row["trip_or_order"],
                    row["customer"],
                    row["status"],
                    f"{row['amount']:,.2f}",
                    row["method"],
                    row["reference"],
                ]
            )
        rows.append(["", "", "", "Total Amount", f"{total_amount:,.2f}", "", ""])

        table = Table(rows, colWidths=[24 * mm, 38 * mm, 42 * mm, 26 * mm, 26 * mm, 32 * mm, 58 * mm], repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{AFRILOTT_GREEN}")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#cbd5e1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("BACKGROUND", (3, -1), (4, -1), colors.HexColor(f"#{AFRILOTT_LIGHT_GREEN}")),
                    ("FONTNAME", (3, -1), (4, -1), "Helvetica-Bold"),
                    ("TEXTCOLOR", (4, -1), (4, -1), colors.HexColor(f"#{AFRILOTT_GREEN}")),
                ]
            )
        )

        story = [
            header_table,
            Spacer(1, 10),
            table,
        ]
        doc.build(story)

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="revenue_report.pdf"'
        response.write(buffer.getvalue())
        return response


class PaymentPanelMixin:
    panel_template_name = "transport/finance/payments/_panel_form.html"
    success_message = "Revenue saved successfully."

    def is_panel_request(self):
        return (
            self.request.headers.get("X-Requested-With") == "XMLHttpRequest"
            and self.request.GET.get("partial") == "form"
        )

    def is_ajax_submit(self):
        return self.request.headers.get("X-Requested-With") == "XMLHttpRequest"

    def get_template_names(self):
        if self.is_panel_request():
            return [self.panel_template_name]
        return super().get_template_names()

    def get_panel_title(self):
        return getattr(self, "panel_title", "Revenue Entry")

    def get_panel_subtitle(self):
        return getattr(self, "panel_subtitle", "")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "panel_title": self.get_panel_title(),
                "panel_subtitle": self.get_panel_subtitle(),
                "submit_label": getattr(self, "submit_label", "Save Revenue"),
            }
        )
        return context

    def _serialize_form_errors(self, form):
        return {field: [str(error) for error in errors] for field, errors in form.errors.items()}

    def form_invalid(self, form):
        if self.is_ajax_submit():
            return JsonResponse(
                {
                    "success": False,
                    "errors": self._serialize_form_errors(form),
                    "non_field_errors": [str(error) for error in form.non_field_errors()],
                },
                status=400,
            )
        return super().form_invalid(form)

    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, self.success_message)
        if self.is_ajax_submit():
            return JsonResponse(
                {
                    "success": True,
                    "payment_id": self.object.pk,
                    "redirect_url": self.get_success_url(),
                    "message": self.success_message,
                }
            )
        return HttpResponseRedirect(self.get_success_url())


class PaymentDetailView(StaffRequiredMixin, DetailView):
    model = Payment
    template_name = "transport/finance/payments/detail.html"
    context_object_name = "payment"

    def get_queryset(self):
        queryset = super().get_queryset().select_related("trip__customer", "trip__vehicle", "trip__driver", "trip__route", "customer")
        return restrict_payments_for_user(queryset, self.request.user)

    def test_func(self):
        return self.request.user.is_authenticated and (
            can_access_finance(self.request.user) or user_has_role(self.request.user, SystemGroup.CUSTOMER)
        )


class PaymentInvoicePdfView(LoginRequiredMixin, View):
    def get(self, request, pk, *args, **kwargs):
        if not (can_access_finance(request.user) or user_has_role(request.user, SystemGroup.CUSTOMER)):
            raise PermissionDenied("You do not have permission to access this invoice.")
        payment = get_object_or_404(
            restrict_payments_for_user(
                Payment.objects.select_related("trip", "order", "customer", "trip__customer", "order__customer", "order__route", "trip__route"),
                request.user,
            ),
            pk=pk,
        )
        pdf_bytes, _context = render_invoice_pdf(payment)
        response = HttpResponse(content_type="application/pdf")
        reference = payment.reference or f"invoice_{payment.pk}"
        response["Content-Disposition"] = f'attachment; filename="{reference}.pdf"'
        response.write(pdf_bytes)
        return response


class PaymentInvoiceVerifyView(View):
    def get(self, request, pk, *args, **kwargs):
        payment = Payment.objects.select_related("trip", "order", "customer", "trip__customer", "order__customer").filter(pk=pk).first()
        token = request.GET.get("token", "")
        is_valid = False
        payload = {}
        error = ""

        try:
            payload = signing.loads(token, salt="atms.invoice", max_age=60 * 60 * 24 * 365)
            is_valid = payment is not None and str(payload.get("p")) == str(pk)
        except signing.BadSignature:
            error = "Invalid verification token."
        except signing.SignatureExpired:
            error = "Verification token has expired."

        customer_name = "-"
        amount_display = "-"
        if payment:
            if payment.customer:
                customer_name = payment.customer.company_name
            elif payment.trip and payment.trip.customer:
                customer_name = payment.trip.customer.company_name
            amount_display = build_invoice_context(payment)["transport_fee_display"]

        status_text = "Verified" if is_valid else "Not Verified"
        status_color = "#0F5B2A" if is_valid else "#B91C1C"
        html = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <title>Invoice Verification</title>
            <style>
                * {{ box-sizing: border-box; }}
                body {{ font-family: Arial, sans-serif; background: #f8fafc; color: #0f172a; margin: 0; padding: 32px 16px; }}
                .card {{ max-width: 720px; margin: 0 auto; background: white; border: 1px solid #dbe5dd; border-radius: 20px; padding: 28px; box-shadow: 0 10px 30px rgba(15, 23, 42, 0.08); }}
                .eyebrow {{ color: #0F5B2A; font-size: 12px; font-weight: 700; text-transform: uppercase; letter-spacing: .12em; }}
                .status {{ color: {status_color}; font-size: 28px; line-height: 1.15; font-weight: 700; margin: 8px 0 18px; }}
                .grid {{ display: grid; grid-template-columns: 180px minmax(0, 1fr); gap: 10px 16px; align-items: start; }}
                .label {{ color: #475569; font-weight: 700; }}
                .value {{ color: #0f172a; overflow-wrap: anywhere; word-break: break-word; }}
                .notice {{ margin-top: 18px; padding: 14px 16px; border-radius: 14px; background: {'#edf8f0' if is_valid else '#fef2f2'}; color: {'#14532d' if is_valid else '#991b1b'}; line-height: 1.5; }}
                @media (max-width: 640px) {{
                    body {{ padding: 16px 12px; }}
                    .card {{ border-radius: 16px; padding: 18px 16px; }}
                    .eyebrow {{ font-size: 11px; letter-spacing: .1em; }}
                    .status {{ font-size: 22px; margin-bottom: 16px; }}
                    .grid {{ grid-template-columns: 1fr; gap: 6px; }}
                    .label {{ font-size: 12px; text-transform: uppercase; letter-spacing: .04em; }}
                    .value {{ margin-bottom: 8px; font-size: 15px; }}
                    .notice {{ margin-top: 14px; padding: 12px 14px; font-size: 14px; }}
                }}
            </style>
        </head>
        <body>
            <div class="card">
                <div class="eyebrow">ZALA/ECO ENERGY Authenticity Check</div>
                <div class="status">{status_text}</div>
                <div class="grid">
                    <div class="label">Invoice Reference</div><div class="value">{getattr(payment, 'reference', '-') if payment else '-'}</div>
                    <div class="label">Amount</div><div class="value">{amount_display}</div>
                    <div class="label">Customer</div><div class="value">{customer_name}</div>
                    <div class="label">Trip</div><div class="value">{getattr(getattr(payment, 'trip', None), 'order_number', '-') if payment else '-'}</div>
                    <div class="label">Order</div><div class="value">{getattr(getattr(payment, 'order', None), 'order_number', '-') if payment else '-'}</div>
                    <div class="label">Payment Status</div><div class="value">{payment.get_status_display() if payment else '-'}</div>
                </div>
                <div class="notice">
                    {"This invoice matches the signed ZALA/ECO ENERGY verification token." if is_valid else (error or "This invoice could not be verified.")}
                </div>
            </div>
        </body>
        </html>
        """
        return HttpResponse(html)


class PaymentCreateView(PaymentPanelMixin, FinanceModalFormMixin, FinanceWriteAccessMixin, CreateView):
    model = Payment
    form_class = PaymentForm
    template_name = "transport/finance/payments/create.html"
    panel_title = "Add Revenue"
    panel_subtitle = "Capture a trip-linked payment or manual revenue entry without leaving the page."
    modal_title = "Add Revenue"
    modal_subtitle = "Capture a completed-trip payment or record manual revenue without leaving the dashboard."
    form_partial = "transport/finance/_payment_form_fields.html"
    submit_label = "Save Revenue"
    submit_icon = "add_card"
    modal_icon = "payments"
    success_message = "Revenue added successfully."

    def get_success_url(self):
        return reverse_lazy("transport:finance:payment-detail", kwargs={"pk": self.object.pk})

    def get_htmx_success_url(self):
        return reverse_lazy("transport:finance:payment-list")

    def get_htmx_refresh_url(self):
        return reverse_lazy("transport:finance:payment-list") + "?partial=list"

    def get_htmx_refresh_target(self):
        return "#payments-list-content"


class PaymentUpdateView(PaymentPanelMixin, FinanceWriteAccessMixin, UpdateView):
    model = Payment
    form_class = PaymentForm
    template_name = "transport/finance/payments/edit.html"
    context_object_name = "payment"
    panel_title = "Edit Revenue"
    panel_subtitle = "Update the revenue entry and keep the finance list in sync."
    submit_label = "Update Revenue"
    success_message = "Revenue updated successfully!"

    def get_queryset(self):
        return super().get_queryset().select_related("trip", "customer")

    def get_success_url(self):
        return reverse_lazy("transport:finance:payment-detail", kwargs={"pk": self.object.pk})


class ExpenseExportMixin(StaffRequiredMixin):
    def get_filtered_queryset(self):
        view = ExpenseListView()
        view.request = self.request
        return view.get_queryset().order_by("-expense_date", "-created_at")

    def _expense_trip_label(self, expense):
        if expense.trip:
            return expense.trip.order_number
        return "Manual Expense"

    def _expense_vehicle_label(self, expense):
        if expense.vehicle:
            return expense.vehicle.plate_number
        if expense.trip and expense.trip.vehicle:
            return expense.trip.vehicle.plate_number
        return "-"

    def _export_rows(self):
        rows = []
        for expense in self.get_filtered_queryset():
            rows.append(
                {
                    "date": expense.expense_date,
                    "trip": self._expense_trip_label(expense),
                    "vehicle": self._expense_vehicle_label(expense),
                    "category": expense.category or (expense.type.name if expense.type else "-"),
                    "status": expense.get_status_display(),
                    "amount": expense.amount or Decimal("0"),
                    "receipt": "Attached" if expense.proof_document else "No receipt",
                    "description": (expense.description or "-").strip() or "-",
                }
            )
        return rows

    def _total_amount(self, rows):
        return sum((row["amount"] or Decimal("0")) for row in rows)


class ExpenseExcelExportView(ExpenseExportMixin, View):
    def get(self, request, *args, **kwargs):
        rows = self._export_rows()
        total_amount = self._total_amount(rows)
        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError:
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="expense_report.csv"'
            writer = csv.writer(response)
            writer.writerow(["Date", "Trip", "Vehicle", "Category", "Status", "Amount", "Receipt", "Description"])
            for row in rows:
                writer.writerow(
                    [
                        row["date"].strftime("%Y-%m-%d") if row["date"] else "",
                        row["trip"],
                        row["vehicle"],
                        row["category"],
                        row["status"],
                        row["amount"],
                        row["receipt"],
                        row["description"],
                    ]
                )
            writer.writerow([])
            writer.writerow(["", "", "", "", "Total Amount", total_amount, "", ""])
            return response

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Expense Report"
        green_fill = PatternFill(fill_type="solid", fgColor=AFRILOTT_GREEN)
        light_fill = PatternFill(fill_type="solid", fgColor=AFRILOTT_LIGHT_GREEN)
        white_font = Font(color="FFFFFF", bold=True, size=12)
        dark_font = Font(color="0F172A", bold=True)
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        logo_stream = _finance_logo_stream()
        if logo_stream:
            logo_image = XLImage(logo_stream)
            logo_image.width = 170
            logo_image.height = 60
            sheet.add_image(logo_image, "A1")

        sheet.merge_cells("C1:H1")
        sheet["C1"] = "ZALA/ECO ENERGY Expense Financial Report"
        sheet["C1"].font = Font(color=AFRILOTT_GREEN, bold=True, size=16)
        sheet["C1"].alignment = Alignment(horizontal="left", vertical="center")

        sheet.merge_cells("C2:H2")
        sheet["C2"] = "Generated from the finance expenses report with active filters applied."
        sheet["C2"].font = Font(color="475569", italic=True, size=10)
        sheet["C2"].alignment = Alignment(horizontal="left", vertical="center")

        sheet["A5"] = "Total Amount"
        sheet["A5"].font = dark_font
        sheet["A5"].fill = light_fill
        sheet["B5"] = float(total_amount)
        sheet["B5"].font = Font(color=AFRILOTT_GREEN, bold=True)
        sheet["B5"].fill = light_fill
        sheet["B5"].number_format = "#,##0.00"

        headers = ["Date", "Trip", "Vehicle", "Category", "Status", "Amount", "Receipt", "Description"]
        header_row = 7
        for index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=index, value=header)
            cell.fill = green_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        current_row = header_row + 1
        for row in rows:
            values = [
                row["date"].strftime("%Y-%m-%d") if row["date"] else "",
                row["trip"],
                row["vehicle"],
                row["category"],
                row["status"],
                float(row["amount"] or 0),
                row["receipt"],
                row["description"],
            ]
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row=current_row, column=column, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border
                if column == 6:
                    cell.number_format = "#,##0.00"
            current_row += 1

        total_row = current_row + 1
        sheet.cell(row=total_row, column=5, value="Total Amount").font = dark_font
        sheet.cell(row=total_row, column=5).fill = light_fill
        sheet.cell(row=total_row, column=5).border = thin_border
        sheet.cell(row=total_row, column=6, value=float(total_amount)).font = Font(color=AFRILOTT_GREEN, bold=True)
        sheet.cell(row=total_row, column=6).fill = light_fill
        sheet.cell(row=total_row, column=6).border = thin_border
        sheet.cell(row=total_row, column=6).number_format = "#,##0.00"

        for column_index, column_cells in enumerate(sheet.columns, start=1):
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 30)

        sheet.freeze_panes = "A8"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="expense_report.xlsx"'
        workbook.save(response)
        return response


class ExpensePdfExportView(ExpenseExportMixin, View):
    def get(self, request, *args, **kwargs):
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=12 * mm,
            leftMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
        )
        styles = getSampleStyleSheet()
        rows_data = self._export_rows()
        total_amount = self._total_amount(rows_data)
        logo_stream = _finance_logo_stream()

        header_left = []
        if logo_stream:
            header_left.append(Image(logo_stream, width=34 * mm, height=16 * mm))
            header_left.append(Spacer(1, 2 * mm))
        header_left.extend(
            [
                Paragraph("<font color='#0F5B2A'><b>ZALA/ECO ENERGY Expense Financial Report</b></font>", styles["Title"]),
                Paragraph("Generated from the finance expenses report with active filters applied.", styles["Normal"]),
            ]
        )

        header_right = [
            Paragraph("<b>Report</b><br/>Expense Entries", styles["Normal"]),
            Spacer(1, 2),
            Paragraph(f"<b>Generated</b><br/>{timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}", styles["Normal"]),
            Spacer(1, 2),
            Paragraph(f"<b>Total Amount</b><br/><font color='#0F5B2A'>{total_amount:,.2f}</font>", styles["Normal"]),
        ]

        header_table = Table([[header_left, header_right]], colWidths=[170 * mm, 85 * mm])
        header_table.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1E7D7")),
                    ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#EAF7EF")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                ]
            )
        )

        rows = [["Date", "Trip", "Vehicle", "Category", "Status", "Amount", "Receipt", "Description"]]
        for row in rows_data:
            rows.append(
                [
                    row["date"].strftime("%d/%m/%Y") if row["date"] else "",
                    row["trip"],
                    row["vehicle"],
                    row["category"],
                    row["status"],
                    f"{row['amount']:,.2f}",
                    row["receipt"],
                    row["description"],
                ]
            )
        rows.append(["", "", "", "", "Total Amount", f"{total_amount:,.2f}", "", ""])

        table = Table(rows, colWidths=[22 * mm, 30 * mm, 28 * mm, 32 * mm, 25 * mm, 24 * mm, 24 * mm, 72 * mm], repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{AFRILOTT_GREEN}")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.2),
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#cbd5e1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("BACKGROUND", (4, -1), (5, -1), colors.HexColor(f"#{AFRILOTT_LIGHT_GREEN}")),
                    ("FONTNAME", (4, -1), (5, -1), "Helvetica-Bold"),
                    ("TEXTCOLOR", (5, -1), (5, -1), colors.HexColor(f"#{AFRILOTT_GREEN}")),
                ]
            )
        )

        story = [header_table, Spacer(1, 10), table]
        doc.build(story)

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="expense_report.pdf"'
        response.write(buffer.getvalue())
        return response


class ExpensePanelMixin:
    panel_template_name = "transport/finance/expenses/_panel_form.html"
    success_message = "Expense saved successfully."

    def is_panel_request(self):
        return (
            self.request.headers.get("X-Requested-With") == "XMLHttpRequest"
            and self.request.GET.get("partial") == "form"
        )

    def is_ajax_submit(self):
        return self.request.headers.get("X-Requested-With") == "XMLHttpRequest"

    def get_template_names(self):
        if self.is_panel_request():
            return [self.panel_template_name]
        return super().get_template_names()

    def get_panel_title(self):
        return getattr(self, "panel_title", "Expense Entry")

    def get_panel_subtitle(self):
        return getattr(self, "panel_subtitle", "")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "panel_title": self.get_panel_title(),
                "panel_subtitle": self.get_panel_subtitle(),
                "submit_label": getattr(self, "submit_label", "Save Expense"),
            }
        )
        return context

    def _serialize_form_errors(self, form):
        return {field: [str(error) for error in errors] for field, errors in form.errors.items()}

    def form_invalid(self, form):
        if self.is_ajax_submit():
            return JsonResponse(
                {
                    "success": False,
                    "errors": self._serialize_form_errors(form),
                    "non_field_errors": [str(error) for error in form.non_field_errors()],
                },
                status=400,
            )
        return super().form_invalid(form)

    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, self.success_message)
        if self.is_ajax_submit():
            return JsonResponse(
                {
                    "success": True,
                    "expense_id": self.object.pk,
                    "redirect_url": self.get_success_url(),
                    "message": self.success_message,
                }
            )
        return HttpResponseRedirect(self.get_success_url())


class ExpenseListView(StaffRequiredMixin, ListView):
    model = Expense
    template_name = "transport/finance/expenses/list.html"
    partial_template_name = "transport/finance/expenses/_list_content.html"
    context_object_name = "expenses"
    paginate_by = 20

    def get_template_names(self):
        if self.request.GET.get("partial") == "list" and (
            self.request.headers.get("HX-Request") == "true"
            or self.request.headers.get("X-Requested-With") == "XMLHttpRequest"
        ):
            return [self.partial_template_name]
        return [self.template_name]

    def get_base_queryset(self):
        return Expense.objects.select_related("trip", "trip__customer", "vehicle", "type").all()

    def get_queryset(self):
        queryset = self.get_base_queryset()
        category = self.request.GET.get("category")
        status = self.request.GET.get("status")
        date_from = self.request.GET.get("date_from")
        date_to = self.request.GET.get("date_to")
        min_amount = self.request.GET.get("min_amount")
        search = self.request.GET.get("search")

        if category:
            queryset = queryset.filter(category__icontains=category)
        if status:
            queryset = queryset.filter(status=status)
        if date_from:
            queryset = queryset.filter(expense_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(expense_date__lte=date_to)
        if min_amount:
            queryset = queryset.filter(amount__gte=min_amount)
        if search:
            queryset = queryset.filter(
                Q(category__icontains=search)
                | Q(description__icontains=search)
                | Q(trip__order_number__icontains=search)
                | Q(vehicle__plate_number__icontains=search)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        expenses = self.get_queryset()
        grouped = OrderedDict()

        for expense in expenses:
            if expense.trip_id:
                key = f"trip-{expense.trip_id}"
                label = expense.trip.order_number
            else:
                key = f"manual-{expense.pk}"
                label = "Manual Expense"

            if key not in grouped:
                grouped[key] = {
                    "key": key,
                    "trip": expense.trip,
                    "label": label,
                    "display_date": expense.expense_date,
                    "latest_amount": expense.amount or Decimal("0"),
                    "total_amount": Decimal("0"),
                    "expense_count": 0,
                    "latest_expense": expense,
                    "expenses": [],
                }

            group = grouped[key]
            group["expenses"].append(expense)
            group["expense_count"] += 1
            group["total_amount"] += expense.amount or Decimal("0")
            if expense.expense_date and expense.expense_date >= group["display_date"]:
                group["display_date"] = expense.expense_date
                group["latest_expense"] = expense
                group["latest_amount"] = expense.amount or Decimal("0")

        grouped_rows = list(grouped.values())
        paginator = Paginator(grouped_rows, self.paginate_by)
        page_number = self.request.GET.get("page") or 1
        page_obj = paginator.get_page(page_number)
        total_amount = expenses.aggregate(total=Sum("amount"))["total"] or Decimal("0")
        approved_amount = expenses.filter(status__in=[Expense.Status.APPROVED, Expense.Status.PAID]).aggregate(total=Sum("amount"))["total"] or Decimal("0")
        pending_amount = expenses.filter(status=Expense.Status.PENDING).aggregate(total=Sum("amount"))["total"] or Decimal("0")
        context.update({
            "expenses": page_obj.object_list,
            "expense_groups": page_obj.object_list,
            "page_obj": page_obj,
            "paginator": paginator,
            "is_paginated": paginator.num_pages > 1,
            "object_list": page_obj.object_list,
            "total_expenses": expenses.count(),
            "total_amount": total_amount,
            "approved_amount": approved_amount,
            "pending_amount": pending_amount,
            "categories": Expense.objects.order_by().values_list("category", flat=True).distinct(),
            "expense_status_choices": Expense.Status.choices,
            "current_category": self.request.GET.get("category"),
            "current_status": self.request.GET.get("status"),
            "current_date_from": self.request.GET.get("date_from"),
            "current_date_to": self.request.GET.get("date_to"),
            "current_min_amount": self.request.GET.get("min_amount"),
            "current_search": self.request.GET.get("search"),
            "can_manage_finance": can_manage_finance(self.request.user),
        })
        return context


class ExpenseDetailView(StaffRequiredMixin, DetailView):
    model = Expense
    template_name = "transport/finance/expenses/detail.html"
    context_object_name = "expense"

    def get_queryset(self):
        return super().get_queryset().select_related("trip__customer", "trip__vehicle", "trip__driver", "trip__route", "vehicle", "type")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["can_manage_finance"] = can_manage_finance(self.request.user)
        return context


class ExpenseCreateView(ExpensePanelMixin, FinanceModalFormMixin, FinanceWriteAccessMixin, CreateView):
    model = Expense
    form_class = ExpenseForm
    template_name = "transport/finance/expenses/create.html"
    panel_title = "Add Expense"
    panel_subtitle = "Capture an operating cost and keep the expense register updated without leaving the page."
    modal_title = "Add Expense"
    modal_subtitle = "Record operating costs and optionally link them to a trip or vehicle."
    form_partial = "transport/finance/_expense_form_fields.html"
    submit_label = "Save Expense"
    submit_icon = "receipt_long"
    modal_icon = "receipt_long"
    success_message = "Expense recorded successfully."

    def get_initial(self):
        initial = super().get_initial()
        trip_id = self.request.GET.get("trip")
        vehicle_id = self.request.GET.get("vehicle")
        amount = self.request.GET.get("amount")
        category = self.request.GET.get("category")
        description = self.request.GET.get("description")

        if trip_id:
            initial["trip"] = trip_id
        if vehicle_id:
            initial["vehicle"] = vehicle_id
        if amount:
            initial["amount"] = amount
        if category:
            initial["category"] = category
        if description:
            initial["description"] = description
        return initial

    def get_success_url(self):
        return reverse_lazy("transport:finance:expense-detail", kwargs={"pk": self.object.pk})

    def get_htmx_success_url(self):
        return reverse_lazy("transport:finance:expense-list")

    def get_htmx_refresh_url(self):
        return reverse_lazy("transport:finance:expense-list") + "?partial=list"

    def get_htmx_refresh_target(self):
        return "#expenses-list-content"

    def form_valid(self, form):
        response = super().form_valid(form)
        manager_emails = approval_recipients(SystemGroup.ADMIN, SystemGroup.OPERATIONS_MANAGER)
        if manager_emails:
            trip_label = self.object.trip.order_number if self.object.trip_id else "Manual Expense"
            vehicle_label = self.object.vehicle.plate_number if self.object.vehicle_id else "Not linked"
            try:
                send_atms_email(
                    subject=f"ZALA/ECO ENERGY approval required for expense #{self.object.pk}",
                    to=manager_emails,
                    greeting="Hello Manager",
                    headline="Expense Approval Required",
                    intro="A new expense was recorded in ZALA/ECO ENERGY and is waiting for approval.",
                    details=[
                        {"label": "Expense ID", "value": str(self.object.pk)},
                        {"label": "Trip", "value": trip_label},
                        {"label": "Vehicle", "value": vehicle_label},
                        {"label": "Category", "value": self.object.category},
                        {"label": "Amount", "value": f"{self.object.amount:,.2f}"},
                    ],
                    note="Review the expense details and approve or reject it in the finance area.",
                    cta_label="Review Expense",
                    cta_url=build_public_url(f"/transport/finance/expenses/{self.object.pk}/"),
                )
            except Exception:
                pass
        return response


class ExpenseUpdateView(ExpensePanelMixin, FinanceWriteAccessMixin, UpdateView):
    model = Expense
    form_class = ExpenseForm
    template_name = "transport/finance/expenses/edit.html"
    context_object_name = "expense"
    panel_title = "Edit Expense"
    panel_subtitle = "Update category, receipt, and status while keeping the register in sync."
    submit_label = "Update Expense"
    success_message = "Expense updated successfully!"

    def get_queryset(self):
        return super().get_queryset().select_related("trip", "vehicle", "type")

    def get_success_url(self):
        return reverse_lazy("transport:finance:expense-detail", kwargs={"pk": self.object.pk})


class DriverFeeListView(StaffRequiredMixin, ListView):
    model = DriverFee
    template_name = "transport/finance/driver_fees/list.html"
    context_object_name = "driver_fees"
    paginate_by = 20

    def get_queryset(self):
        return DriverFee.objects.select_related("trip", "driver").order_by("-fee_date", "-created_at")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["can_manage_finance"] = can_manage_finance(self.request.user)
        return context


class DriverFeeCreateView(FinanceModalFormMixin, FinanceWriteAccessMixin, CreateView):
    model = DriverFee
    form_class = DriverFeeForm
    template_name = "transport/finance/driver_fees/create.html"
    modal_title = "Add Driver Fee"
    modal_subtitle = "Capture trip-based driver payout information with validation before you leave the dashboard."
    form_partial = "transport/finance/_driver_fee_form_fields.html"
    submit_label = "Save Driver Fee"
    submit_icon = "badge"
    modal_icon = "badge"
    success_message = "Driver fee added successfully."

    def get_success_url(self):
        return reverse_lazy("transport:finance:driverfee-list")
