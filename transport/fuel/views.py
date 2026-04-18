from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.db.models import Avg, Count, Q, Sum
from django.db.models.functions import TruncMonth
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.generic import DetailView, ListView, TemplateView, View
from io import BytesIO
from pathlib import Path

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from accounts.decorators import driver_required
from accounts.templatetags.currency_tags import currency_filter
from accounts.rbac import can_access_fuel, can_approve_fuel, can_manage_fuel

from .forms import FuelDocumentForm, FuelRequestForm
from .models import FuelDocument, FuelRequest
from .services import (
    apply_fuel_filters,
    base_fuel_expense_queryset,
    build_chart_data,
    build_fuel_records,
    build_global_stats,
    build_loss_detection,
    build_trip_analysis,
    build_vehicle_summary,
    fuel_filter_options,
    weighted_average_fuel_price,
)
from transport.drivers.models import Driver
from transport.vehicles.models import Vehicle


AFRILOTT_GREEN = "0F5B2A"
AFRILOTT_LIGHT_GREEN = "EAF7EF"


def _sync_trip_fuel_cost_from_requests(trip):
    approved_total = (
        FuelRequest.objects.filter(trip=trip, is_approved=True)
        .aggregate(total=Sum("amount"))
        .get("total")
        or 0
    )
    trip.fuel_cost = approved_total
    # Use full save so Trip model recalculates dependent financial fields.
    trip.save()


def _can_access_fuel_module(user):
    return user.is_authenticated and can_access_fuel(user)


def _logo_path():
    candidate = Path(__file__).resolve().parents[2] / "static" / "img" / "ZALA/ECO ENERGY.png"
    return candidate if candidate.exists() else None


def _fuel_dashboard_filters(request):
    return {
        "vehicle": request.GET.get("vehicle", ""),
        "trip": request.GET.get("trip", ""),
        "date_from": request.GET.get("date_from", ""),
        "date_to": request.GET.get("date_to", ""),
    }


def _filtered_fuel_records(filters):
    queryset = apply_fuel_filters(base_fuel_expense_queryset(), filters)
    average_price = weighted_average_fuel_price(queryset)
    records = build_fuel_records(queryset, average_price=average_price)
    return queryset, records, average_price


@login_required
def fuel_dashboard(request):
    if not _can_access_fuel_module(request.user):
        messages.error(request, "You do not have permission to view fuel management.")
        return redirect("transport:dashboard")

    selected_filters = _fuel_dashboard_filters(request)
    queryset, records, average_price = _filtered_fuel_records(selected_filters)
    vehicle_summary = build_vehicle_summary(records)
    global_stats = build_global_stats(records)

    paginator = Paginator(records, 20)
    page_obj = paginator.get_page(request.GET.get("page") or 1)
    filter_options = fuel_filter_options()

    context = {
        "page_obj": page_obj,
        "fuel_records": page_obj.object_list,
        "vehicle_summary": vehicle_summary,
        "total_records": len(records),
        "average_fuel_price": average_price,
        "selected_vehicle": selected_filters["vehicle"],
        "selected_trip": selected_filters["trip"],
        "selected_date_from": selected_filters["date_from"],
        "selected_date_to": selected_filters["date_to"],
        **filter_options,
        **global_stats,
    }
    return render(request, "transport/fuel/index.html", context)


@login_required
def fuel_analytics(request):
    if not _can_access_fuel_module(request.user):
        messages.error(request, "You do not have permission to view fuel analytics.")
        return redirect("transport:dashboard")

    selected_filters = _fuel_dashboard_filters(request)
    queryset, records, average_price = _filtered_fuel_records(selected_filters)
    vehicle_summary = build_vehicle_summary(records)
    trip_analysis = build_trip_analysis(records)
    loss_detection = build_loss_detection(records)
    global_stats = build_global_stats(records)
    filter_options = fuel_filter_options()
    chart_data = build_chart_data(queryset, records)

    best_vehicle = next((row for row in sorted(vehicle_summary, key=lambda item: item["km_per_liter"], reverse=True) if row_has_efficiency(row)), None)
    worst_vehicle = next((row for row in sorted(vehicle_summary, key=lambda item: item["km_per_liter"]) if row_has_efficiency(row)), None)

    context = {
        "vehicle_summary": vehicle_summary,
        "best_vehicle": best_vehicle,
        "worst_vehicle": worst_vehicle,
        "highest_fuel_usage": trip_analysis["highest_fuel_usage"],
        "highest_cost_per_km": trip_analysis["highest_cost_per_km"],
        "high_consumption_records": loss_detection["flagged_records"],
        "average_fuel_per_km_threshold": loss_detection["threshold"],
        "fleet_average_fuel_per_km": loss_detection["average_fuel_per_km"],
        "average_fuel_price": average_price,
        "selected_vehicle": selected_filters["vehicle"],
        "selected_trip": selected_filters["trip"],
        "selected_date_from": selected_filters["date_from"],
        "selected_date_to": selected_filters["date_to"],
        **filter_options,
        **global_stats,
        **chart_data,
    }
    return render(request, "transport/fuel/analytics.html", context)


class FuelExportMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return _can_access_fuel_module(self.request.user)

    def get_export_data(self):
        filters = _fuel_dashboard_filters(self.request)
        queryset, records, average_price = _filtered_fuel_records(filters)
        global_stats = build_global_stats(records)
        return {
            "filters": filters,
            "queryset": queryset,
            "records": records,
            "average_price": average_price,
            "global_stats": global_stats,
        }


class FuelExcelExportView(FuelExportMixin, View):
    def get(self, request, *args, **kwargs):
        export_data = self.get_export_data()
        records = export_data["records"]
        filename = "fuel_report.xlsx"
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError:
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="fuel_report.csv"'
            response.write("Trip,Vehicle,Driver,Fuel (L),Cost,Distance (km),Fuel/KM,Cost/KM,Route,Date\n")
            for record in records:
                response.write(
                    f'"{getattr(record["trip"], "order_number", "Manual Fuel Expense")}","{getattr(record["vehicle"], "plate_number", "-")}","{getattr(record["driver"], "name", "-")}","{record["liters"]}","{record["cost"]}","{record["distance"]}","{record["fuel_per_km"] or ""}","{record["cost_per_km"] or ""}","{record["route_label"]}","{record["date"]}"\n'
                )
            return response

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Fuel Report"

        header_fill = PatternFill(fill_type="solid", fgColor=AFRILOTT_GREEN)
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        sheet.merge_cells("A1:J1")
        sheet["A1"] = "ZALA/ECO ENERGY Fuel Management Report"
        sheet["A1"].font = Font(color=AFRILOTT_GREEN, bold=True, size=16)
        sheet.merge_cells("A2:J2")
        sheet["A2"] = f"Generated on {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
        sheet["A2"].font = Font(color="475569", italic=True, size=10)

        stats = export_data["global_stats"]
        sheet["A4"] = "Total Fuel Cost"
        sheet["B4"] = currency_filter(stats["total_cost"])
        sheet["D4"] = "Total Fuel Used"
        sheet["E4"] = f'{stats["total_liters"]:.2f} L'
        sheet["G4"] = "Total Distance"
        sheet["H4"] = f'{stats["total_distance"]:.0f} km'

        headers = ["Trip", "Vehicle", "Driver", "Fuel (L)", "Cost", "Distance (km)", "Fuel / KM", "Cost / KM", "Route", "Date"]
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=6, column=idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_idx, record in enumerate(records, start=7):
            values = [
                getattr(record["trip"], "order_number", "Manual Fuel Expense"),
                getattr(record["vehicle"], "plate_number", "-"),
                getattr(record["driver"], "name", "-"),
                float(record["liters"]),
                currency_filter(record["cost"]),
                float(record["distance"]),
                float(record["fuel_per_km"] or 0),
                currency_filter(record["cost_per_km"] or 0),
                record["route_label"],
                record["date"].strftime("%d/%m/%Y") if record["date"] else "",
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

        for col_idx in range(1, 11):
            column_letter = get_column_letter(col_idx)
            max_length = max(len(str(sheet.cell(row=row, column=col_idx).value or "")) for row in range(1, sheet.max_row + 1))
            sheet.column_dimensions[column_letter].width = min(max_length + 3, 28)

        sheet.freeze_panes = "A7"

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type=mime_type)
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class FuelPdfExportView(FuelExportMixin, View):
    def get(self, request, *args, **kwargs):
        export_data = self.get_export_data()
        records = export_data["records"]
        stats = export_data["global_stats"]
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=12 * mm,
            leftMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
        )
        styles = getSampleStyleSheet()
        logo = _logo_path()
        story = []

        header_left = []
        if logo:
            header_left.append(Image(str(logo), width=34 * mm, height=16 * mm))
            header_left.append(Spacer(1, 2 * mm))
        header_left.extend(
            [
                Paragraph("<font color='#0F5B2A'><b>ZALA/ECO ENERGY Fuel Management Report</b></font>", styles["Title"]),
                Paragraph("Fuel expense export generated from the ZALA/ECO ENERGY fuel dashboard.", styles["Normal"]),
            ]
        )
        header_right = [
            Paragraph("<b>Generated</b><br/>%s" % timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M"), styles["Normal"]),
            Spacer(1, 2),
            Paragraph(f"<b>Total Fuel Cost</b><br/>{currency_filter(stats['total_cost'])}", styles["Normal"]),
            Spacer(1, 2),
            Paragraph(f"<b>Total Fuel Used</b><br/>{stats['total_liters']:.2f} L", styles["Normal"]),
        ]
        header_table = Table([[header_left, header_right]], colWidths=[170 * mm, 85 * mm])
        header_table.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1E7D7")),
                    ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#EAF7EF")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                ]
            )
        )
        story.extend([header_table, Spacer(1, 10)])

        rows = [["Trip", "Vehicle", "Driver", "Fuel (L)", "Cost", "Distance", "Fuel / KM", "Cost / KM", "Route", "Date"]]
        for record in records:
            rows.append(
                [
                    getattr(record["trip"], "order_number", "Manual Fuel Expense"),
                    getattr(record["vehicle"], "plate_number", "-"),
                    getattr(record["driver"], "name", "-"),
                    f'{record["liters"]:.2f}',
                    currency_filter(record["cost"]),
                    f'{record["distance"]:.0f} km',
                    f'{record["fuel_per_km"]:.4f}' if record["fuel_per_km"] is not None else "-",
                    currency_filter(record["cost_per_km"]) if record["cost_per_km"] is not None else "-",
                    record["route_label"],
                    record["date"].strftime("%d/%m/%Y") if record["date"] else "",
                ]
            )

        table = Table(
            rows,
            colWidths=[28 * mm, 22 * mm, 28 * mm, 20 * mm, 24 * mm, 22 * mm, 18 * mm, 24 * mm, 42 * mm, 22 * mm],
            repeatRows=1,
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{AFRILOTT_GREEN}")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#cbd5e1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(table)
        doc.build(story)

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="fuel_report.pdf"'
        response.write(buffer.getvalue())
        return response


def row_has_efficiency(row):
    return row["km_per_liter"] is not None and row["km_per_liter"] > 0


class StaffRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return _can_access_fuel_module(self.request.user)


class FuelRequestListView(StaffRequiredMixin, ListView):
    model = FuelRequest
    template_name = "transport/fuel/staff_list.html"
    context_object_name = "fuel_requests"
    paginate_by = 20

    def get_template_names(self):
        if self.request.GET.get("partial") == "list":
            return ["transport/fuel/_staff_list_content.html"]
        return [self.template_name]

    def get_queryset(self):
        qs = FuelRequest.objects.select_related("trip", "station", "driver").order_by("-created_at")
        status = self.request.GET.get("status", "").lower()
        search = self.request.GET.get("search", "").strip()
        vehicle_id = self.request.GET.get("vehicle", "").strip()
        driver_id = self.request.GET.get("driver", "").strip()
        date_from = self.request.GET.get("date_from", "").strip()
        date_to = self.request.GET.get("date_to", "").strip()

        if status == "approved":
            qs = qs.filter(is_approved=True)
        elif status == "pending":
            qs = qs.filter(is_approved=False)

        if vehicle_id:
            qs = qs.filter(trip__vehicle_id=vehicle_id)

        if driver_id:
            qs = qs.filter(driver_id=driver_id)

        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)

        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        if search:
            qs = qs.filter(
                Q(trip__order_number__icontains=search)
                | Q(station__name__icontains=search)
                | Q(driver__full_name__icontains=search)
                | Q(driver__email__icontains=search)
            )

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        all_requests = FuelRequest.objects.select_related("trip", "station", "driver")
        ctx["total_requests"] = all_requests.count()
        ctx["approved_requests"] = all_requests.filter(is_approved=True).count()
        ctx["pending_requests"] = all_requests.filter(is_approved=False).count()
        ctx["total_amount"] = all_requests.aggregate(total=Sum("amount"))["total"] or 0
        ctx["total_liters_used"] = all_requests.aggregate(total=Sum("amount"))["total"] or 0
        ctx["avg_fuel_consumption"] = all_requests.aggregate(avg=Avg("amount"))["avg"] or 0
        ctx["top_vehicle_usage"] = (
            all_requests.values("trip__vehicle__plate_number")
            .annotate(total_amount=Sum("amount"), req_count=Count("id"))
            .order_by("-total_amount")
            .first()
        )
        ctx["selected_status"] = self.request.GET.get("status", "")
        ctx["search_query"] = self.request.GET.get("search", "")
        ctx["selected_vehicle"] = self.request.GET.get("vehicle", "")
        ctx["selected_driver"] = self.request.GET.get("driver", "")
        ctx["selected_date_from"] = self.request.GET.get("date_from", "")
        ctx["selected_date_to"] = self.request.GET.get("date_to", "")
        ctx["vehicles"] = Vehicle.objects.order_by("plate_number")
        ctx["drivers"] = Driver.objects.select_related("user").order_by("name")
        ctx["can_manage_fuel"] = can_manage_fuel(self.request.user)
        ctx["can_approve_fuel"] = can_approve_fuel(self.request.user)
        return ctx


class FuelRequestDetailView(StaffRequiredMixin, DetailView):
    model = FuelRequest
    template_name = "transport/fuel/staff_detail.html"
    context_object_name = "fuel_request"

    def get_queryset(self):
        return FuelRequest.objects.select_related("trip", "station", "driver").prefetch_related("documents")


class FuelRequestAnalyticsView(StaffRequiredMixin, TemplateView):
    template_name = "transport/fuel/staff_analytics.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        requests = FuelRequest.objects.select_related("station", "trip")

        ctx["total_requests"] = requests.count()
        ctx["approved_requests"] = requests.filter(is_approved=True).count()
        ctx["pending_requests"] = requests.filter(is_approved=False).count()
        ctx["total_amount"] = requests.aggregate(total=Sum("amount"))["total"] or 0

        ctx["top_stations"] = (
            requests.values("station__name")
            .annotate(total_amount=Sum("amount"), req_count=Count("id"))
            .order_by("-total_amount")[:5]
        )

        ctx["monthly_requests"] = (
            requests.annotate(month=TruncMonth("created_at"))
            .values("month")
            .annotate(req_count=Count("id"), total_amount=Sum("amount"))
            .order_by("month")
        )
        return ctx


@login_required
@require_POST
def approve_fuel_request(request, pk):
    if not can_approve_fuel(request.user):
        messages.error(request, "You do not have permission to approve fuel requests.")
        return redirect("transport:fuel:detail", pk=pk)

    fuel_request = get_object_or_404(FuelRequest, pk=pk)
    if fuel_request.is_approved:
        messages.info(request, "Fuel request is already approved.")
    else:
        fuel_request.is_approved = True
        fuel_request.approved_by = request.user
        fuel_request.approved_at = timezone.now()
        _sync_trip_fuel_cost_from_requests(fuel_request.trip)
        fuel_request.posted_to_trip = True
        fuel_request.posted_at = timezone.now()
        fuel_request.save(
            update_fields=[
                "is_approved",
                "approved_by",
                "approved_at",
                "posted_to_trip",
                "posted_at",
                "updated_at",
            ]
        )
        messages.success(request, f"Fuel request #{fuel_request.pk} approved.")

    return redirect("transport:fuel:detail", pk=pk)


@driver_required
def request_fuel(request):
    active_trip = (
        FuelRequestForm(driver_user=request.user).fields["trip"].queryset.first()
        if request.method != "POST"
        else None
    )
    if request.method == "POST":
        form = FuelRequestForm(request.POST, request.FILES, driver_user=request.user)
        if form.is_valid():
            fuel_request = form.save(commit=False)
            fuel_request.driver = request.user
            fuel_request.save()
            receipt = form.cleaned_data.get("receipt")
            if receipt:
                FuelDocument.objects.create(fuel_request=fuel_request, document=receipt)
            messages.success(request, "Fuel request submitted successfully.")
            return redirect("transport:driver_fuel")
    else:
        form = FuelRequestForm(driver_user=request.user, trip=active_trip)

    return render(
        request,
        "transport/fuel/request.html",
        {
            "form": form,
            "initial_tab": "fuel",
            "driver_spa": False,
        },
    )


@driver_required
def upload_fuel_document(request, fuel_request_id):
    fuel_request = get_object_or_404(
        FuelRequest,
        pk=fuel_request_id,
        driver=request.user,
    )

    if request.method == "POST":
        form = FuelDocumentForm(request.POST, request.FILES)
        if form.is_valid():
            document = form.save(commit=False)
            document.fuel_request = fuel_request
            document.save()
            messages.success(request, "Fuel document uploaded successfully.")
            return redirect("transport:driver_fuel")
    else:
        form = FuelDocumentForm()

    return render(
        request,
        "transport/fuel/upload_document.html",
        {
            "form": form,
            "fuel_request": fuel_request,
            "initial_tab": "fuel",
            "driver_spa": False,
        },
    )
