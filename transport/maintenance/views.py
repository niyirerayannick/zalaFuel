import warnings
from io import BytesIO
from pathlib import Path

from django.http import HttpResponse, JsonResponse
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView
from django.contrib import messages
from django.urls import reverse_lazy
from django.db.models import Q, Sum, Count, Avg
from django.urls import reverse
from django.conf import settings
from django.utils import timezone
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from django.shortcuts import get_object_or_404, redirect
import json
from urllib import request as urllib_request
from urllib.error import URLError, HTTPError

from .models import MaintenanceRecord, default_service_types
from .forms import MaintenanceRecordForm
from .services import approve_maintenance_record, sync_maintenance_expense
from transport.trips.models import Trip
from transport.vehicles.models import Vehicle
from accounts.rbac import SystemGroup, user_has_role


def _logo_path():
    candidate = Path(__file__).resolve().parents[2] / "static" / "img" / "ZALA Terminal.png"
    return candidate if candidate.exists() else None


def _logo_stream(max_width=900):
    logo = _logo_path()
    if not logo:
        return None

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", PILImage.DecompressionBombWarning)
        img = PILImage.open(logo)
        img.load()
    with img:
        img.thumbnail((max_width, max_width))
        stream = BytesIO()
        img.save(stream, format="PNG", optimize=True)
        stream.seek(0)
        return stream


class StaffRequiredMixin(UserPassesTestMixin):
    """Mixin to require staff access level"""
    def test_func(self):
        return self.request.user.is_authenticated and user_has_role(
            self.request.user,
            SystemGroup.ADMIN,
            SystemGroup.OPERATIONS_MANAGER,
            SystemGroup.LOGISTICS_COORDINATOR,
        )


class MaintenanceListView(StaffRequiredMixin, ListView):
    """List view for maintenance records with filtering and search"""
    model = MaintenanceRecord
    template_name = 'transport/maintenance/list.html'
    context_object_name = 'maintenance_records'
    paginate_by = 20

    def get_template_names(self):
        if self.request.GET.get("partial") == "list":
            return ["transport/maintenance/_list_content.html"]
        return [self.template_name]

    def get_queryset(self):
        queryset = MaintenanceRecord.objects.select_related('vehicle', 'trip', 'expense').all()
        
        # Filtering
        vehicle_id = self.request.GET.get('vehicle')
        trip_id = self.request.GET.get('trip')
        status = self.request.GET.get('status')
        service_type = self.request.GET.get('service_type')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        search = self.request.GET.get('search')
        
        if vehicle_id:
            queryset = queryset.filter(vehicle_id=vehicle_id)
        if trip_id:
            queryset = queryset.filter(trip_id=trip_id)
        if status:
            queryset = queryset.filter(status=status)
        if service_type:
            queryset = queryset.filter(service_type__icontains=service_type)
        if date_from:
            queryset = queryset.filter(service_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(service_date__lte=date_to)
        if search:
            queryset = queryset.filter(
                Q(vehicle__plate_number__icontains=search) |
                Q(vehicle__vehicle_type__icontains=search) |
                Q(service_type__icontains=search) |
                Q(workshop__icontains=search)
            )
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Statistics
        records = MaintenanceRecord.objects.all()
        service_types = list(default_service_types())
        for value in MaintenanceRecord.objects.values_list('service_type', flat=True).distinct():
            value = (value or "").strip()
            if value and value not in service_types:
                service_types.append(value)

        context.update({
            'total_records': records.count(),
            'total_cost': records.filter(status=MaintenanceRecord.Status.APPROVED).aggregate(Sum('cost'))['cost__sum'] or 0,
            'avg_cost': records.aggregate(Avg('cost'))['cost__avg'] or 0,
            'vehicles_in_maintenance': Vehicle.objects.filter(
                status=Vehicle.VehicleStatus.MAINTENANCE
            ).count(),
            'pending_approval_count': records.filter(status=MaintenanceRecord.Status.PENDING).count(),
            'vehicles': Vehicle.objects.all(),
            'trips': Trip.objects.select_related("vehicle").order_by("-created_at")[:100],
            'service_types': service_types,
            'status_choices': MaintenanceRecord.Status.choices,
            
            # Filter values
            'current_vehicle': self.request.GET.get('vehicle'),
            'current_trip': self.request.GET.get('trip'),
            'current_status': self.request.GET.get('status'),
            'current_service_type': self.request.GET.get('service_type'),
            'current_date_from': self.request.GET.get('date_from'),
            'current_date_to': self.request.GET.get('date_to'),
            'current_search': self.request.GET.get('search'),
            'create_form': MaintenanceRecordForm(),
        })
        
        return context


class MaintenanceExportMixin(StaffRequiredMixin):
    def get_filtered_queryset(self):
        view = MaintenanceListView()
        view.request = self.request
        return view.get_queryset()


class MaintenanceExcelExportView(MaintenanceExportMixin, ListView):
    def get(self, request, *args, **kwargs):
        queryset = self.get_filtered_queryset()
        filename = "maintenance_report.xlsx"
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError:
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="maintenance_report.csv"'
            response.write("Vehicle,Trip,Service Type,Date,Status,Odometer,Cost,Workshop,Downtime\n")
            for record in queryset:
                response.write(
                    f'"{record.vehicle.plate_number}","{record.trip.order_number if record.trip else "Standalone"}","{record.service_type}","{record.service_date:%Y-%m-%d}","{record.get_status_display()}","{record.service_km} km","{record.cost}","{record.workshop}","{record.downtime_days} day(s)"\n'
                )
            return response

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Maintenance"
        header_fill = PatternFill(fill_type="solid", fgColor="0F5B2A")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"), top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"))
        sheet.merge_cells("A1:I1")
        sheet["A1"] = "ZALA Terminal Maintenance Report"
        sheet["A1"].font = Font(color="0F5B2A", bold=True, size=16)
        sheet.merge_cells("A2:I2")
        sheet["A2"] = f"Generated on {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
        sheet["A2"].font = Font(color="475569", italic=True, size=10)
        headers = ["Vehicle", "Trip", "Service Type", "Date", "Status", "Odometer", "Cost", "Workshop", "Downtime"]
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=4, column=idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        row_no = 5
        for record in queryset:
            values = [
                record.vehicle.plate_number,
                record.trip.order_number if record.trip else "Standalone",
                record.service_type,
                record.service_date.strftime("%d/%m/%Y"),
                record.get_status_display(),
                f"{record.service_km} km",
                float(record.cost or 0),
                record.workshop,
                f"{record.downtime_days} day(s)",
            ]
            for col_no, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_no, column=col_no, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top")
            row_no += 1
        sheet.auto_filter.ref = f"A4:I{max(row_no - 1, 4)}"
        sheet.freeze_panes = "A5"
        for column_cells in sheet.columns:
            column_index = column_cells[0].column
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 3, 28)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type=mime_type)
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class MaintenancePdfExportView(MaintenanceExportMixin, ListView):
    def get(self, request, *args, **kwargs):
        queryset = self.get_filtered_queryset()
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
        styles = getSampleStyleSheet()
        logo_stream = _logo_stream()
        header_left = []
        if logo_stream:
            header_left.append(Image(logo_stream, width=34 * mm, height=16 * mm))
            header_left.append(Spacer(1, 2 * mm))
        header_left.extend([Paragraph("<font color='#0F5B2A'><b>ZALA Terminal Maintenance Report</b></font>", styles["Title"]), Paragraph("Maintenance register export generated from ZALA Terminal.", styles["Normal"])])
        header_right = [Paragraph("<b>Report</b><br/>Maintenance Register", styles["Normal"]), Spacer(1, 2), Paragraph(f"<b>Generated</b><br/>{timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}", styles["Normal"]), Spacer(1, 2), Paragraph(f"<b>Total Records</b><br/>{queryset.count()}", styles["Normal"])]
        header_table = Table([[header_left, header_right]], colWidths=[170 * mm, 85 * mm])
        header_table.setStyle(TableStyle([("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1E7D7")), ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#EAF7EF")), ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10), ("TOPPADDING", (0, 0), (-1, -1), 10), ("BOTTOMPADDING", (0, 0), (-1, -1), 10), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
        rows = [["Vehicle", "Trip", "Service Type", "Date", "Status", "Odometer", "Cost", "Workshop", "Downtime"]]
        for record in queryset:
            rows.append([record.vehicle.plate_number, record.trip.order_number if record.trip else "Standalone", record.service_type, record.service_date.strftime("%d/%m/%Y"), record.get_status_display(), f"{record.service_km} km", str(record.cost), record.workshop, f"{record.downtime_days} day(s)"])
        if len(rows) == 1:
            rows.append(["-", "-", "-", "-", "-", "-", "-", "-", "-"])
        table = Table(rows, colWidths=[22 * mm, 24 * mm, 33 * mm, 21 * mm, 24 * mm, 24 * mm, 20 * mm, 50 * mm, 20 * mm], repeatRows=1)
        table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F5B2A")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 8), ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1D5DB")), ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]), ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6), ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
        doc.build([header_table, Spacer(1, 5 * mm), table])
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="maintenance_report.pdf"'
        return response


class MaintenanceDetailView(StaffRequiredMixin, DetailView):
    """Detail view for a maintenance record"""
    model = MaintenanceRecord
    template_name = 'transport/maintenance/detail.html'
    context_object_name = 'record'

    def get_queryset(self):
        return MaintenanceRecord.objects.select_related('vehicle', 'trip', 'expense', 'approved_by', 'created_by')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        record = self.get_object()
        
        # Related maintenance records for this vehicle
        context['related_records'] = MaintenanceRecord.objects.filter(
            vehicle=record.vehicle
        ).select_related('vehicle', 'trip').exclude(id=record.id).order_by('-service_date')[:5]
        context["can_approve"] = user_has_role(
            self.request.user,
            SystemGroup.ADMIN,
            SystemGroup.OPERATIONS_MANAGER,
        ) and record.status == MaintenanceRecord.Status.PENDING
        
        return context


class MaintenanceCreateView(StaffRequiredMixin, CreateView):
    """Create view for maintenance records"""
    model = MaintenanceRecord
    form_class = MaintenanceRecordForm
    template_name = 'transport/maintenance/create.html'

    def _is_ajax(self):
        return self.request.headers.get("x-requested-with") == "XMLHttpRequest"

    def get_template_names(self):
        if self.request.GET.get("partial") == "form":
            return ["transport/maintenance/_modal_form.html"]
        return [self.template_name]

    def get_success_url(self):
        messages.success(self.request, 'Maintenance record created successfully!')
        return reverse_lazy('transport:maintenance:detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.status = MaintenanceRecord.Status.PENDING
        if self._is_ajax():
            self.object = form.save()
            return JsonResponse({
                "success": True,
                "id": self.object.pk,
                "detail_url": reverse("transport:maintenance:detail", kwargs={"pk": self.object.pk}),
                "message": "Maintenance record has been created.",
            })
        messages.success(self.request, 'Maintenance record has been created.')
        return super().form_valid(form)

    def form_invalid(self, form):
        if self._is_ajax():
            return JsonResponse({
                "success": False,
                "errors": form.errors,
                "non_field_errors": form.non_field_errors(),
            }, status=400)
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        service_types = list(default_service_types())
        for value in MaintenanceRecord.objects.values_list("service_type", flat=True).distinct():
            value = (value or "").strip()
            if value and value not in service_types:
                service_types.append(value)
        context["service_types"] = service_types
        return context


class MaintenanceUpdateView(StaffRequiredMixin, UpdateView):
    """Update view for maintenance records"""
    model = MaintenanceRecord
    form_class = MaintenanceRecordForm
    template_name = 'transport/maintenance/edit.html'

    def _is_ajax(self):
        return self.request.headers.get("x-requested-with") == "XMLHttpRequest"

    def get_template_names(self):
        if self.request.GET.get("partial") == "form":
            return ["transport/maintenance/_modal_form.html"]
        return [self.template_name]

    def get_success_url(self):
        messages.success(self.request, 'Maintenance record updated successfully!')
        return reverse_lazy('transport:maintenance:detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        if self._is_ajax():
            self.object = form.save()
            sync_maintenance_expense(self.object)
            return JsonResponse({
                "success": True,
                "id": self.object.pk,
                "detail_url": reverse("transport:maintenance:detail", kwargs={"pk": self.object.pk}),
                "message": "Maintenance record has been updated.",
            })
        response = super().form_valid(form)
        sync_maintenance_expense(self.object)
        return response

    def form_invalid(self, form):
        if self._is_ajax():
            return JsonResponse({
                "success": False,
                "errors": form.errors,
                "non_field_errors": form.non_field_errors(),
            }, status=400)
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        service_types = list(default_service_types())
        for value in MaintenanceRecord.objects.values_list("service_type", flat=True).distinct():
            value = (value or "").strip()
            if value and value not in service_types:
                service_types.append(value)
        context["service_types"] = service_types
        return context


def service_types_api(request):
    """
    Returns maintenance service types.
    If MAINTENANCE_SERVICE_TYPES_URL is configured, fetch from that system.
    Falls back to local distinct service types.
    """
    external_url = getattr(settings, "MAINTENANCE_SERVICE_TYPES_URL", "").strip()
    service_types = []
    source = "local"

    if external_url:
        try:
            req = urllib_request.Request(external_url, headers={"Accept": "application/json"})
            with urllib_request.urlopen(req, timeout=5) as response:
                payload = json.loads(response.read().decode("utf-8"))

            # Accept either: ["Oil Change", ...] or {"service_types": ["Oil Change", ...]}
            if isinstance(payload, list):
                service_types = [str(item).strip() for item in payload if str(item).strip()]
            elif isinstance(payload, dict):
                items = payload.get("service_types", [])
                if isinstance(items, list):
                    service_types = [str(item).strip() for item in items if str(item).strip()]

            if service_types:
                source = "external"
        except (URLError, HTTPError, TimeoutError, json.JSONDecodeError, ValueError):
            service_types = []

    if not service_types:
        service_types = list(default_service_types())
        for value in MaintenanceRecord.objects.values_list("service_type", flat=True).distinct():
            value = (value or "").strip()
            if value and value not in service_types:
                service_types.append(value)

    return JsonResponse({"service_types": service_types, "source": source})


def approve_maintenance_request(request, pk):
    if not request.user.is_authenticated or not user_has_role(
        request.user,
        SystemGroup.ADMIN,
        SystemGroup.OPERATIONS_MANAGER,
    ):
        messages.error(request, "You do not have permission to approve maintenance requests.")
        return redirect("transport:maintenance:detail", pk=pk)

    record = get_object_or_404(MaintenanceRecord.objects.select_related("trip", "vehicle", "expense"), pk=pk)
    approve_maintenance_record(record, request.user)
    messages.success(request, "Maintenance request approved and posted to expenses.")
    return redirect("transport:maintenance:detail", pk=pk)
