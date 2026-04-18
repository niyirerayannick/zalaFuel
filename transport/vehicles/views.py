# Vehicle Module Views
from datetime import timedelta
from io import BytesIO
from pathlib import Path
import warnings

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count, Avg
from django.utils.decorators import method_decorator
from django.utils import timezone
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from accounts.rbac import can_access_fleet, can_manage_fleet
from .models import Vehicle, VehicleOwner
from .forms import VehicleForm, VehicleOwnerForm


def _logo_path():
    candidate = Path(__file__).resolve().parents[2] / "static" / "img" / "ZALA Terminal.png"
    return candidate if candidate.exists() else None


def _logo_stream(max_width=900):
    logo = _logo_path()
    if not logo:
        return None

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", PILImage.DecompressionBombWarning)
        img = PILImage.open(logo)
        img.load()
    with img:
        img.thumbnail((max_width, max_width))
        stream = BytesIO()
        img.save(stream, format="PNG", optimize=True)
        stream.seek(0)
        return stream


class StaffRequiredMixin(UserPassesTestMixin):
    """Mixin to require staff access level"""
    def test_func(self):
        return self.request.user.is_authenticated and can_access_fleet(self.request.user)


class FleetWriteRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_authenticated and can_manage_fleet(self.request.user)


class VehicleListView(StaffRequiredMixin, ListView):
    model = Vehicle
    template_name = 'transport/vehicles/list.html'
    context_object_name = 'vehicles'
    paginate_by = 20

    def get_template_names(self):
        if self.request.GET.get("partial") == "list":
            return ["transport/vehicles/_list_content.html"]
        return [self.template_name]

    def get_queryset(self):
        queryset = Vehicle.objects.select_related("owner")
        search = self.request.GET.get('search')
        status = self.request.GET.get('status')
        
        if search:
            queryset = queryset.filter(
                Q(plate_number__icontains=search) | 
                Q(vehicle_model__icontains=search) |
                Q(fuel_type__icontains=search) |
                Q(color__icontains=search)
            )
        
        if status:
            queryset = queryset.filter(status=status)
        
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Fleet Management'
        context['total_vehicles'] = Vehicle.objects.count()
        context['available_vehicles'] = Vehicle.objects.filter(status='AVAILABLE').count()
        context['assigned_vehicles'] = Vehicle.objects.filter(status='ASSIGNED').count()
        context['maintenance_vehicles'] = Vehicle.objects.filter(status='MAINTENANCE').count()
        context['external_vehicles'] = Vehicle.objects.filter(
            ownership_type=Vehicle.OwnershipType.EXTERNAL
        ).count()
        context['company_vehicles'] = Vehicle.objects.filter(
            ownership_type=Vehicle.OwnershipType.COMPANY
        ).count()
        context['status_choices'] = Vehicle.STATUS_CHOICES
        context['search_query'] = self.request.GET.get('search', '')
        context['selected_status'] = self.request.GET.get('status', '')
        context['today'] = timezone.now().date()
        context['warning_date'] = timezone.now().date() + timedelta(days=30)
        context['create_form'] = VehicleForm()
        context['can_manage_fleet'] = can_manage_fleet(self.request.user)
        return context


class VehicleExportMixin(StaffRequiredMixin):
    def get_filtered_queryset(self):
        view = VehicleListView()
        view.request = self.request
        return view.get_queryset()


class VehicleExcelExportView(VehicleExportMixin, ListView):
    def get(self, request, *args, **kwargs):
        queryset = self.get_filtered_queryset()
        filename = "vehicles_report.xlsx"
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError:
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="vehicles_report.csv"'
            response.write("Plate Number,Model,Type,Ownership,Owner,Status,Odometer,Next Service\n")
            for vehicle in queryset:
                response.write(
                    f'{vehicle.plate_number},{vehicle.vehicle_model or ""},{vehicle.get_vehicle_type_display()},{vehicle.get_ownership_type_display()},{vehicle.owner.name if vehicle.owner else ""},{vehicle.get_status_display()},{vehicle.current_odometer},{vehicle.next_service_km}\n'
                )
            return response

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Vehicles"

        header_fill = PatternFill(fill_type="solid", fgColor="0F5B2A")
        header_font = Font(color="FFFFFF", bold=True)
        body_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
        alt_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        sheet.merge_cells("A1:H1")
        sheet["A1"] = "ZALA Terminal Vehicle Report"
        sheet["A1"].font = Font(color="0F5B2A", bold=True, size=16)
        sheet.merge_cells("A2:H2")
        sheet["A2"] = f"Fleet register export generated on {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
        sheet["A2"].font = Font(color="475569", italic=True, size=10)

        headers = ["Plate Number", "Model", "Type", "Ownership", "Owner", "Status", "Odometer (km)", "Next Service (km)"]
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=4, column=idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        row_no = 5
        for vehicle in queryset:
            values = [
                vehicle.plate_number,
                vehicle.vehicle_model or "-",
                vehicle.get_vehicle_type_display(),
                vehicle.get_ownership_type_display(),
                vehicle.owner.name if vehicle.owner else "-",
                vehicle.get_status_display(),
                float(vehicle.current_odometer or 0),
                float(vehicle.next_service_km or 0),
            ]
            for col_no, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_no, column=col_no, value=value)
                cell.border = thin_border
                cell.fill = body_fill if row_no % 2 == 0 else alt_fill
                cell.alignment = Alignment(vertical="top")
                if col_no in {7, 8}:
                    cell.number_format = "#,##0"
            row_no += 1

        sheet.auto_filter.ref = f"A4:H{max(row_no - 1, 4)}"
        sheet.freeze_panes = "A5"
        for column_cells in sheet.columns:
            column_index = column_cells[0].column
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 3, 28)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type=mime_type)
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class VehiclePdfExportView(VehicleExportMixin, ListView):
    def get(self, request, *args, **kwargs):
        queryset = self.get_filtered_queryset()
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=12 * mm,
            leftMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
        )
        styles = getSampleStyleSheet()

        logo_stream = _logo_stream()
        header_left = []
        if logo_stream:
            header_left.append(Image(logo_stream, width=34 * mm, height=16 * mm))
            header_left.append(Spacer(1, 2 * mm))
        header_left.extend(
            [
                Paragraph("<font color='#0F5B2A'><b>ZALA Terminal Vehicle Report</b></font>", styles["Title"]),
                Paragraph("Fleet list export generated from the vehicle register.", styles["Normal"]),
            ]
        )
        header_right = [
            Paragraph("<b>Report</b><br/>Vehicle Register", styles["Normal"]),
            Spacer(1, 2),
            Paragraph(f"<b>Generated</b><br/>{timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}", styles["Normal"]),
            Spacer(1, 2),
            Paragraph(f"<b>Total Vehicles</b><br/>{queryset.count()}", styles["Normal"]),
        ]
        header_table = Table([[header_left, header_right]], colWidths=[170 * mm, 85 * mm])
        header_table.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1E7D7")),
                    ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#EAF7EF")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )

        rows = [["Plate", "Model", "Type", "Ownership", "Owner", "Status", "Odometer", "Next Service"]]
        for vehicle in queryset:
            rows.append(
                [
                    vehicle.plate_number,
                    vehicle.vehicle_model or "-",
                    vehicle.get_vehicle_type_display(),
                    vehicle.get_ownership_type_display(),
                    vehicle.owner.name if vehicle.owner else "-",
                    vehicle.get_status_display(),
                    f"{vehicle.current_odometer:,.0f} km",
                    f"{vehicle.next_service_km:,.0f} km",
                ]
            )
        if len(rows) == 1:
            rows.append(["-", "-", "-", "-", "-", "-", "-", "-"])

        table = Table(
            rows,
            colWidths=[25 * mm, 38 * mm, 28 * mm, 24 * mm, 34 * mm, 24 * mm, 26 * mm, 28 * mm],
            repeatRows=1,
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F5B2A")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1D5DB")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )

        doc.build([header_table, Spacer(1, 5 * mm), table])
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="vehicles_report.pdf"'
        return response


class VehicleDetailView(StaffRequiredMixin, DetailView):
    model = Vehicle
    template_name = 'transport/vehicles/detail.html'
    context_object_name = 'vehicle'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        vehicle = self.get_object()
        today = timezone.now().date()
        
        # Current tab
        tab = self.request.GET.get('tab', 'overview')
        context['current_tab'] = tab
        context['today'] = today
        context['warning_date'] = today + timedelta(days=30)
        
        # Recent trips via reverse relation
        trips = vehicle.trips.all().order_by('-created_at')
        context['recent_trips'] = trips[:10]
        context['total_trips'] = trips.count()
        
        # Trip-based metrics
        trip_stats = trips.aggregate(
            total_revenue=Sum('revenue'),
            total_fuel_cost=Sum('fuel_cost'),
            total_distance=Sum('distance'),
            total_cost=Sum('total_cost'),
            total_profit=Sum('profit'),
            avg_cost_per_km=Avg('cost_per_km'),
        )
        
        context['performance_metrics'] = {
            'total_trips': trips.count(),
            'total_revenue': trip_stats['total_revenue'] or 0,
            'total_fuel_cost': trip_stats['total_fuel_cost'] or 0,
            'total_distance': trip_stats['total_distance'] or 0,
            'total_cost': trip_stats['total_cost'] or 0,
            'total_profit': trip_stats['total_profit'] or 0,
            'avg_cost_per_km': trip_stats['avg_cost_per_km'] or 0,
        }
        
        # Maintenance records
        maintenance_records = vehicle.maintenance_records.all().order_by('-service_date')
        context['maintenance_records'] = maintenance_records[:5]
        maintenance_stats = maintenance_records.aggregate(
            total_maintenance_cost=Sum('cost'),
            total_downtime=Sum('downtime_days'),
        )
        context['total_maintenance_cost'] = maintenance_stats['total_maintenance_cost'] or 0
        context['total_downtime_days'] = maintenance_stats['total_downtime'] or 0
        
        # Fuel entries
        fuel_entries = vehicle.trips.aggregate(
            total_fuel=Sum('fuel_issued'),
            total_fuel_cost=Sum('fuel_cost'),
        )
        context['total_fuel_issued'] = fuel_entries['total_fuel'] or 0
        context['total_fuel_cost'] = fuel_entries['total_fuel_cost'] or 0
        
        # Document status alerts
        alerts = []
        if vehicle.insurance_expiry < today:
            alerts.append({'type': 'danger', 'message': 'Insurance has EXPIRED!'})
        elif vehicle.insurance_expiry < today + timedelta(days=30):
            days_left = (vehicle.insurance_expiry - today).days
            alerts.append({'type': 'warning', 'message': f'Insurance expires in {days_left} days'})
        
        if vehicle.inspection_expiry < today:
            alerts.append({'type': 'danger', 'message': 'Inspection has EXPIRED!'})
        elif vehicle.inspection_expiry < today + timedelta(days=30):
            days_left = (vehicle.inspection_expiry - today).days
            alerts.append({'type': 'warning', 'message': f'Inspection expires in {days_left} days'})
        
        if vehicle.current_odometer >= vehicle.next_service_km:
            km_over = int(vehicle.current_odometer - vehicle.next_service_km)
            alerts.append({'type': 'danger', 'message': f'Service OVERDUE by {km_over:,} km'})
        else:
            km_left = int(vehicle.next_service_km - vehicle.current_odometer)
            if km_left < 1000:
                alerts.append({'type': 'warning', 'message': f'Service due in {km_left:,} km'})
        
        context['alerts'] = alerts
        
        return context


class VehicleCreateView(FleetWriteRequiredMixin, CreateView):
    model = Vehicle
    form_class = VehicleForm
    template_name = 'transport/vehicles/create.html'
    success_url = reverse_lazy('transport:vehicles:list')

    def _is_ajax(self):
        return self.request.headers.get("x-requested-with") == "XMLHttpRequest"

    def get_template_names(self):
        if self.request.GET.get("partial") == "form":
            return ["transport/vehicles/_modal_form.html"]
        return [self.template_name]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form_action"] = self.request.path
        context["not_configured_label"] = "Not configured in current model"
        return context

    def form_valid(self, form):
        if self._is_ajax():
            self.object = form.save()
            return JsonResponse({
                "success": True,
                "id": self.object.pk,
                "detail_url": reverse("transport:vehicles:detail", kwargs={"pk": self.object.pk}),
                "message": f"Vehicle {self.object.plate_number} created successfully!",
            })
        messages.success(self.request, f'Vehicle {form.instance.plate_number} created successfully!')
        return super().form_valid(form)

    def form_invalid(self, form):
        if self._is_ajax():
            return JsonResponse({
                "success": False,
                "errors": form.errors,
                "non_field_errors": form.non_field_errors(),
                "message": "Please correct the highlighted vehicle fields and try again.",
            }, status=400)
        return super().form_invalid(form)


class VehicleOwnerListView(StaffRequiredMixin, ListView):
    model = VehicleOwner
    template_name = "transport/vehicles/owners.html"
    panel_template_name = "transport/vehicles/_owner_modal_form.html"
    context_object_name = "owners"
    paginate_by = 20

    def get_template_names(self):
        if self.request.GET.get("partial") == "form":
            return [self.panel_template_name]
        return [self.template_name]

    def get_queryset(self):
        queryset = VehicleOwner.objects.prefetch_related("vehicles").order_by("name")
        search = self.request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search)
                | Q(phone__icontains=search)
                | Q(bank_name__icontains=search)
                | Q(bank_account__icontains=search)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Car Owners"
        context["search_query"] = self.request.GET.get("search", "")
        context["total_owners"] = VehicleOwner.objects.count()
        context["external_vehicles"] = Vehicle.objects.filter(
            ownership_type=Vehicle.OwnershipType.EXTERNAL
        ).count()
        context["owners_with_vehicles"] = VehicleOwner.objects.filter(vehicles__isnull=False).distinct().count()
        context["create_form"] = kwargs.get("create_form") or VehicleOwnerForm()
        context["form_action"] = self.request.path
        context["can_manage_fleet"] = can_manage_fleet(self.request.user)
        return context

    def post(self, request, *args, **kwargs):
        if not can_manage_fleet(request.user):
            return JsonResponse({"success": False, "message": "You do not have permission to add car owners."}, status=403) if request.headers.get("X-Requested-With") == "XMLHttpRequest" else redirect("transport:vehicles:owner-list")
        form = VehicleOwnerForm(request.POST)
        if form.is_valid():
            owner = form.save()
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse(
                    {
                        "success": True,
                        "id": owner.pk,
                        "detail_url": reverse("transport:vehicles:owner-detail", kwargs={"pk": owner.pk}),
                        "message": f'Car owner "{owner.name}" added successfully.',
                    }
                )
            messages.success(request, f'Car owner "{owner.name}" added successfully.')
            return redirect("transport:vehicles:owner-list")
        self.object_list = self.get_queryset()
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {
                    "success": False,
                    "errors": form.errors,
                    "non_field_errors": form.non_field_errors(),
                    "message": "Please correct the highlighted owner fields and try again.",
                },
                status=400,
            )
        return self.render_to_response(self.get_context_data(create_form=form))


class VehicleOwnerDetailView(StaffRequiredMixin, DetailView):
    model = VehicleOwner
    template_name = "transport/vehicles/owner_detail.html"
    context_object_name = "owner"

    def get_queryset(self):
        return VehicleOwner.objects.prefetch_related("vehicles")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        owner = self.object
        vehicles = owner.vehicles.all().order_by("plate_number")
        context["vehicles"] = vehicles
        context["total_vehicles"] = vehicles.count()
        context["available_vehicles"] = vehicles.filter(status=Vehicle.VehicleStatus.AVAILABLE).count()
        context["assigned_vehicles"] = vehicles.filter(status=Vehicle.VehicleStatus.ASSIGNED).count()
        context["maintenance_vehicles"] = vehicles.filter(status=Vehicle.VehicleStatus.MAINTENANCE).count()
        context["can_manage_fleet"] = can_manage_fleet(self.request.user)
        return context


class VehicleUpdateView(FleetWriteRequiredMixin, UpdateView):
    model = Vehicle
    form_class = VehicleForm
    template_name = 'transport/vehicles/edit.html'

    def _is_ajax(self):
        return self.request.headers.get("x-requested-with") == "XMLHttpRequest"

    def get_template_names(self):
        if self.request.GET.get("partial") == "form":
            return ["transport/vehicles/_modal_form.html"]
        return [self.template_name]
    
    def get_success_url(self):
        return reverse('transport:vehicles:detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['vehicle'] = self.get_object()
        context['today'] = timezone.now().date()
        context["form_action"] = self.request.path
        context["not_configured_label"] = "Not configured in current model"
        return context

    def form_valid(self, form):
        if self._is_ajax():
            self.object = form.save()
            return JsonResponse({
                "success": True,
                "id": self.object.pk,
                "detail_url": reverse("transport:vehicles:detail", kwargs={"pk": self.object.pk}),
                "message": f"Vehicle {self.object.plate_number} updated successfully!",
            })
        messages.success(self.request, f'Vehicle {form.instance.plate_number} updated successfully!')
        return super().form_valid(form)

    def form_invalid(self, form):
        if self._is_ajax():
            return JsonResponse({
                "success": False,
                "errors": form.errors,
                "non_field_errors": form.non_field_errors(),
                "message": "Please correct the highlighted vehicle fields and try again.",
            }, status=400)
        return super().form_invalid(form)


@login_required  
def vehicle_quick_status(request, vehicle_id):
    """AJAX view to quickly update vehicle status"""
    if not can_manage_fleet(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)
    new_status = request.POST.get('status')
    
    if new_status in dict(Vehicle.STATUS_CHOICES):
        vehicle.status = new_status
        vehicle.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Vehicle status updated to {vehicle.get_status_display()}'
        })
    
    return JsonResponse({'success': False, 'error': 'Invalid status'})
