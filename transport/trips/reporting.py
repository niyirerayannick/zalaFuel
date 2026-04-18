from csv import writer as csv_writer
from decimal import Decimal
from io import BytesIO, StringIO
from pathlib import Path
import warnings

from django.conf import settings
from django.utils import timezone
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


BRAND_GREEN = colors.HexColor("#0F5B2A")
BRAND_LIGHT = colors.HexColor("#EAF7EF")
INK = colors.HexColor("#0F172A")
SLATE = colors.HexColor("#475569")
BORDER = colors.HexColor("#D8E3DA")
HEADER_DARK = colors.HexColor("#0B3B1F")
CARD_BG = colors.HexColor("#F8FBF8")
SOFT_GRAY = colors.HexColor("#F8FAFC")


def _logo_path():
    candidate = Path(settings.BASE_DIR) / "static" / "img" / "ZALA/ECO ENERGY.png"
    return candidate if candidate.exists() else None


def _logo_stream(max_width=900):
    logo = _logo_path()
    if not logo:
        return None

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", PILImage.DecompressionBombWarning)
        img = PILImage.open(logo)
        img.load()
    with img:
        img.thumbnail((max_width, max_width))
        stream = BytesIO()
        img.save(stream, format="PNG", optimize=True)
        stream.seek(0)
        return stream


def _display(value, default="-"):
    if value is None:
        return default
    text = str(value).strip()
    return text or default


def _format_decimal(value, places=2):
    amount = value if value is not None else Decimal("0")
    formatted = f"{Decimal(amount):,.{places}f}"
    if "." in formatted:
        formatted = formatted.rstrip("0").rstrip(".")
    return formatted


def _money_display(value, currency_code=None):
    prefix = f"{currency_code} " if currency_code else ""
    return f"{prefix}{_format_decimal(value)}"


def build_trip_report_context(trip):
    shipments = list(trip.shipments.select_related("order", "customer", "order__unit"))
    expenses = list(trip.expenses.select_related("type").order_by("-created_at"))
    invoices = list(trip.payments.select_related("order").order_by("-created_at"))
    related_orders = list(trip.related_orders.select_related("customer", "unit"))

    shipment_rows = []
    for shipment in shipments:
        unit_symbol = getattr(getattr(shipment.order, "unit", None), "symbol", "")
        business_quantity = f"{shipment.quantity} {unit_symbol}".strip() or str(shipment.quantity)
        shipment_rows.append(
            {
                "order_number": shipment.order.order_number,
                "customer": shipment.customer.company_name,
                "business_quantity": business_quantity,
                "weight_kg": shipment.weight_kg,
                "carriage_type": shipment.get_carriage_type_display(),
                "status": shipment.get_status_display(),
            }
        )

    expense_rows = []
    for expense in expenses:
        expense_rows.append(
            {
                "type": expense.type.name if expense.type else expense.category,
                "amount": expense.amount or Decimal("0"),
                "description": expense.description or "-",
            }
        )

    invoice_rows = []
    for invoice in invoices:
        invoice_rows.append(
            {
                "reference": invoice.reference or f"INV-{invoice.pk}",
                "amount": invoice.amount or Decimal("0"),
                "status": invoice.get_status_display(),
            }
        )

    return {
        "trip": trip,
        "generated_at": timezone.now(),
        "shipments": shipment_rows,
        "expenses": expense_rows,
        "invoices": invoice_rows,
        "related_orders": related_orders,
        "summary": {
            "customer": getattr(trip.customer, "company_name", "-"),
            "driver": getattr(trip.driver, "name", "-"),
            "vehicle": getattr(trip.vehicle, "plate_number", "-"),
            "route": f"{_display(getattr(trip.route, 'origin', None))} to {_display(getattr(trip.route, 'destination', None))}",
            "status": trip.get_status_display(),
            "load": f"{_format_decimal(trip.total_load_weight_kg)} kg",
            "shipments_count": len(shipment_rows),
            "orders_count": len(related_orders),
            "expected_revenue": trip.expected_revenue or Decimal("0"),
            "total_revenue": trip.total_revenue or Decimal("0"),
            "total_expenses": trip.total_expenses or Decimal("0"),
            "net_profit": trip.net_profit or Decimal("0"),
        },
    }


def render_trip_report_pdf(trip):
    context = build_trip_report_context(trip)
    currency_code = getattr(settings, "DEFAULT_CURRENCY", "USD")
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=8 * mm,
        leftMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TripReportTitle", parent=styles["Title"], fontSize=16, textColor=BRAND_GREEN, leading=18, spaceAfter=1)
    subtitle_style = ParagraphStyle("TripReportSubtitle", parent=styles["Normal"], fontSize=8, textColor=SLATE, leading=10)
    section_style = ParagraphStyle("TripReportSection", parent=styles["Heading2"], fontSize=9.5, textColor=BRAND_GREEN, spaceAfter=3)
    metric_value_style = ParagraphStyle("TripReportMetricValue", parent=styles["Normal"], fontSize=10.5, textColor=HEADER_DARK, leading=11.5, spaceAfter=0)
    metric_label_style = ParagraphStyle("TripReportMetricLabel", parent=styles["Normal"], fontSize=6.8, textColor=SLATE, leading=8)
    header_meta_label_style = ParagraphStyle("TripReportHeaderMetaLabel", parent=styles["Normal"], fontSize=6.8, textColor=SLATE, leading=7.5)
    header_meta_value_style = ParagraphStyle("TripReportHeaderMetaValue", parent=styles["Normal"], fontSize=8.2, textColor=INK, leading=9.2)

    logo_stream = _logo_stream()
    header_left = []
    if logo_stream:
        header_left.append(Image(logo_stream, width=26 * mm, height=12 * mm))
        header_left.append(Spacer(1, 1 * mm))
    header_left.extend(
        [
            Paragraph("ZALA/ECO ENERGY Trip Report", title_style),
            Paragraph("Operational trip summary generated from ZALA/ECO ENERGY with shipment, financial, and invoice information.", subtitle_style),
        ]
    )
    header_meta_rows = [
        [
            Paragraph("Trip Reference", header_meta_label_style),
            Paragraph(trip.order_number, header_meta_value_style),
        ],
        [
            Paragraph("Generated", header_meta_label_style),
            Paragraph(timezone.localtime(context["generated_at"]).strftime("%d/%m/%Y %H:%M"), header_meta_value_style),
        ],
        [
            Paragraph("Status", header_meta_label_style),
            Paragraph(context["summary"]["status"], header_meta_value_style),
        ],
        [
            Paragraph("Currency", header_meta_label_style),
            Paragraph(currency_code, header_meta_value_style),
        ],
    ]
    header_right = Table(header_meta_rows, colWidths=[18 * mm, 31 * mm])
    header_right.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), BRAND_LIGHT),
                ("BOX", (0, 0), (-1, -1), 0.8, BORDER),
                ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#DCE7DE")),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    header = Table([[header_left, header_right]], colWidths=[118 * mm, 49 * mm])
    header.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.8, BORDER),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    metrics = [
        ("Load", context["summary"]["load"]),
        ("Shipments", str(context["summary"]["shipments_count"])),
        ("Revenue", _money_display(context["summary"]["total_revenue"], currency_code)),
        ("Expenses", _money_display(context["summary"]["total_expenses"], currency_code)),
        ("Net Profit", _money_display(context["summary"]["net_profit"], currency_code)),
    ]
    metric_cells = []
    for label, value in metrics:
        metric_cells.append(
            Table(
                [[Paragraph(value, metric_value_style)], [Paragraph(label, metric_label_style)]],
                colWidths=[33.4 * mm],
            )
        )
    metrics_table = Table([metric_cells], colWidths=[33.4 * mm] * 5)
    metrics_table.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.8, BORDER),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
                ("BACKGROUND", (0, 0), (-1, -1), CARD_BG),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )

    summary_rows = [
        ["Customer", context["summary"]["customer"], "Vehicle", context["summary"]["vehicle"]],
        ["Driver", context["summary"]["driver"], "Route", context["summary"]["route"]],
        ["Load", context["summary"]["load"], "Related Orders", str(context["summary"]["orders_count"])],
        ["Shipments", str(context["summary"]["shipments_count"]), "Net Profit", _money_display(context["summary"]["net_profit"], currency_code)],
    ]
    summary_table = Table(summary_rows, colWidths=[20 * mm, 49 * mm, 20 * mm, 78 * mm])
    summary_table.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.8, BORDER),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E6EEE8")),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F8FBF8")),
                ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#F8FBF8")),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.8),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )

    shipment_data = [["Order", "Customer", "Business Quantity", "Weight (kg)", "Carriage", "Status"]]
    for row in context["shipments"]:
        shipment_data.append(
            [
                row["order_number"],
                row["customer"],
                row["business_quantity"],
                _format_decimal(row["weight_kg"]),
                row["carriage_type"],
                row["status"],
            ]
        )
    if len(shipment_data) == 1:
        shipment_data.append(["-", "-", "-", "-", "-", "-"])
    shipment_table = Table(shipment_data, colWidths=[22 * mm, 31 * mm, 31 * mm, 18 * mm, 20 * mm, 18 * mm], repeatRows=1)
    shipment_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), BRAND_GREEN),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("BOX", (0, 0), (-1, -1), 0.8, BORDER),
                ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BRAND_LIGHT]),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )

    finance_rows = [
        ["Metric", "Value"],
        ["Expected Revenue", _money_display(context["summary"]["expected_revenue"], currency_code)],
        ["Total Revenue", _money_display(context["summary"]["total_revenue"], currency_code)],
        ["Total Expenses", _money_display(context["summary"]["total_expenses"], currency_code)],
        ["Net Profit", _money_display(context["summary"]["net_profit"], currency_code)],
        ["Invoices", str(len(context["invoices"]))],
    ]
    if context["invoices"]:
        for index, row in enumerate(context["invoices"], start=1):
            finance_rows.append(
                [
                    f"Invoice {index}",
                    f"{row['reference']} | {_money_display(row['amount'], currency_code)} | {row['status']}",
                ]
            )
    else:
        finance_rows.append(["Invoice Status", "No invoices"])
    finance_table = Table(finance_rows, colWidths=[37 * mm, 130 * mm], repeatRows=1)
    finance_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), BRAND_GREEN),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.2),
                ("BOX", (0, 0), (-1, -1), 0.8, BORDER),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E6EEE8")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FBF8")]),
                ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )

    expense_rows = [["Type", "Amount", "Description"]]
    for row in context["expenses"]:
        expense_rows.append([row["type"], _format_decimal(row["amount"]), row["description"]])
    if len(expense_rows) == 1:
        expense_rows.append(["No expenses", "-", "-"])
    expense_table = Table(expense_rows, colWidths=[33 * mm, 21 * mm, 113 * mm], repeatRows=1)
    expense_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), BRAND_GREEN),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("BOX", (0, 0), (-1, -1), 0.8, BORDER),
                ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FBF8")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )

    story = [
        header,
        Spacer(1, 3 * mm),
        metrics_table,
        Spacer(1, 3 * mm),
        Paragraph("Trip Summary", section_style),
        summary_table,
        Spacer(1, 3 * mm),
        Paragraph("Shipments", section_style),
        shipment_table,
        Spacer(1, 3 * mm),
        Paragraph("Financial Summary", section_style),
        finance_table,
        Spacer(1, 3 * mm),
        Paragraph("Expenses", section_style),
        expense_table,
        Spacer(1, 2 * mm),
        Paragraph("Generated by ZALA/ECO ENERGY.", subtitle_style),
    ]
    doc.build(story)
    return buffer.getvalue(), context


def render_trip_report_excel(trip):
    context = build_trip_report_context(trip)
    filename = f"trip_report_{trip.order_number}.xlsx"
    mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        csv_buffer = StringIO()
        writer = csv_writer(csv_buffer)
        writer.writerow(["Trip Reference", trip.order_number])
        writer.writerow(["Customer", context["summary"]["customer"]])
        writer.writerow(["Driver", context["summary"]["driver"]])
        writer.writerow(["Vehicle", context["summary"]["vehicle"]])
        writer.writerow(["Route", context["summary"]["route"]])
        writer.writerow(["Status", context["summary"]["status"]])
        writer.writerow([])
        writer.writerow(["Order", "Customer", "Business Quantity", "Weight (kg)", "Carriage", "Status"])
        for row in context["shipments"]:
            writer.writerow(
                [
                    row["order_number"],
                    row["customer"],
                    row["business_quantity"],
                    _format_decimal(row["weight_kg"]),
                    row["carriage_type"],
                    row["status"],
                ]
            )
        writer.writerow([])
        writer.writerow(["Expense Type", "Amount", "Description"])
        for row in context["expenses"]:
            writer.writerow(
                [
                    row["type"],
                    _format_decimal(row["amount"]),
                    row["description"],
                ]
            )
        return csv_buffer.getvalue().encode("utf-8"), context, "text/csv", f"trip_report_{trip.order_number}.csv"

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Trip Summary"
    shipments_sheet = workbook.create_sheet("Shipments")
    expenses_sheet = workbook.create_sheet("Expenses")

    green_fill = PatternFill(fill_type="solid", fgColor="0F5B2A")
    light_fill = PatternFill(fill_type="solid", fgColor="EAF7EF")
    slate_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    dark_fill = PatternFill(fill_type="solid", fgColor="0B3B1F")
    white_font = Font(color="FFFFFF", bold=True, size=11)
    dark_font = Font(color="0F172A", bold=True)
    title_font = Font(color="0F5B2A", bold=True, size=16)
    subtitle_font = Font(color="475569", italic=True, size=10)
    metric_value_font = Font(color="0F172A", bold=True, size=13)
    metric_label_font = Font(color="475569", bold=True, size=9)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    logo_stream = _logo_stream()
    if logo_stream:
        logo = XLImage(logo_stream)
        logo.width = 170
        logo.height = 60
        summary_sheet.add_image(logo, "A1")

    summary_sheet.merge_cells("C1:F1")
    summary_sheet["C1"] = "ZALA/ECO ENERGY Trip Report"
    summary_sheet["C1"].font = title_font
    summary_sheet.merge_cells("C2:F2")
    summary_sheet["C2"] = "Operational summary generated from the selected trip."
    summary_sheet["C2"].font = subtitle_font

    metric_positions = [
        ("A4", "Load", context["summary"]["load"]),
        ("C4", "Shipments", context["summary"]["shipments_count"]),
        ("E4", "Revenue", float(context["summary"]["total_revenue"] or 0)),
        ("A7", "Expenses", float(context["summary"]["total_expenses"] or 0)),
        ("C7", "Net Profit", float(context["summary"]["net_profit"] or 0)),
        ("E7", "Status", context["summary"]["status"]),
    ]
    for start_cell, label, value in metric_positions:
        column = summary_sheet[start_cell].column
        row = summary_sheet[start_cell].row
        summary_sheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + 1)
        summary_sheet.merge_cells(start_row=row + 1, start_column=column, end_row=row + 1, end_column=column + 1)
        label_cell = summary_sheet.cell(row=row, column=column, value=label)
        value_cell = summary_sheet.cell(row=row + 1, column=column, value=value)
        label_cell.font = metric_label_font
        value_cell.font = metric_value_font
        label_cell.fill = light_fill
        value_cell.fill = slate_fill
        label_cell.border = thin_border
        value_cell.border = thin_border
        label_cell.alignment = Alignment(horizontal="left")
        value_cell.alignment = Alignment(horizontal="left")
        if isinstance(value, float):
            value_cell.number_format = "#,##0.00"

    summary_items = [
        ("Trip Reference", trip.order_number),
        ("Customer", context["summary"]["customer"]),
        ("Driver", context["summary"]["driver"]),
        ("Vehicle", context["summary"]["vehicle"]),
        ("Route", context["summary"]["route"]),
        ("Status", context["summary"]["status"]),
        ("Load", context["summary"]["load"]),
        ("Shipments", context["summary"]["shipments_count"]),
        ("Expected Revenue", float(context["summary"]["expected_revenue"])),
        ("Total Revenue", float(context["summary"]["total_revenue"])),
        ("Total Expenses", float(context["summary"]["total_expenses"])),
        ("Net Profit", float(context["summary"]["net_profit"])),
    ]
    row_no = 11
    for label, value in summary_items:
        label_cell = summary_sheet.cell(row=row_no, column=1, value=label)
        value_cell = summary_sheet.cell(row=row_no, column=2, value=value)
        label_cell.font = dark_font
        label_cell.fill = light_fill
        label_cell.border = thin_border
        value_cell.border = thin_border
        value_cell.fill = white_fill
        label_cell.alignment = Alignment(horizontal="left")
        value_cell.alignment = Alignment(horizontal="left")
        if isinstance(value, float):
            value_cell.number_format = "#,##0.00"
        row_no += 1

    shipment_headers = ["Order", "Customer", "Business Quantity", "Weight (kg)", "Carriage", "Status"]
    for idx, header in enumerate(shipment_headers, start=1):
        cell = shipments_sheet.cell(row=1, column=idx, value=header)
        cell.fill = dark_fill
        cell.font = white_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    shipments_sheet.auto_filter.ref = f"A1:F{max(len(context['shipments']) + 1, 2)}"
    current_row = 2
    for row in context["shipments"]:
        values = [
            row["order_number"],
            row["customer"],
            row["business_quantity"],
            float(row["weight_kg"] or 0),
            row["carriage_type"],
            row["status"],
        ]
        for column, value in enumerate(values, start=1):
            cell = shipments_sheet.cell(row=current_row, column=column, value=value)
            cell.border = thin_border
            cell.fill = light_fill if current_row % 2 == 0 else slate_fill
            cell.alignment = Alignment(vertical="top")
            if column == 4:
                cell.number_format = "#,##0.00"
        current_row += 1

    expense_headers = ["Type", "Amount", "Description"]
    for idx, header in enumerate(expense_headers, start=1):
        cell = expenses_sheet.cell(row=1, column=idx, value=header)
        cell.fill = dark_fill
        cell.font = white_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    expenses_sheet.auto_filter.ref = f"A1:C{max(len(context['expenses']) + 1, 2)}"
    current_row = 2
    for row in context["expenses"]:
        values = [row["type"], float(row["amount"] or 0), row["description"]]
        for column, value in enumerate(values, start=1):
            cell = expenses_sheet.cell(row=current_row, column=column, value=value)
            cell.border = thin_border
            cell.fill = light_fill if current_row % 2 == 0 else slate_fill
            cell.alignment = Alignment(vertical="top")
            if column == 2:
                cell.number_format = "#,##0.00"
        current_row += 1
    expenses_sheet.cell(row=current_row + 1, column=1, value="Total Expenses").font = dark_font
    expenses_sheet.cell(row=current_row + 1, column=1).fill = light_fill
    expenses_sheet.cell(row=current_row + 1, column=1).border = thin_border
    expenses_sheet.cell(row=current_row + 1, column=2, value=float(context["summary"]["total_expenses"] or 0)).font = dark_font
    expenses_sheet.cell(row=current_row + 1, column=2).fill = light_fill
    expenses_sheet.cell(row=current_row + 1, column=2).border = thin_border
    expenses_sheet.cell(row=current_row + 1, column=2).number_format = "#,##0.00"

    summary_sheet.sheet_view.showGridLines = False
    shipments_sheet.sheet_view.showGridLines = False
    expenses_sheet.sheet_view.showGridLines = False
    summary_sheet.sheet_properties.tabColor = "0F5B2A"
    shipments_sheet.sheet_properties.tabColor = "0F5B2A"
    expenses_sheet.sheet_properties.tabColor = "475569"

    for sheet in (summary_sheet, shipments_sheet, expenses_sheet):
        for column_cells in sheet.columns:
            column_index = column_cells[0].column
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 3, 32)

    summary_sheet.freeze_panes = "A11"
    shipments_sheet.freeze_panes = "A2"
    expenses_sheet.freeze_panes = "A2"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), context, mime_type, filename


def get_trip_report_export(trip, export_format):
    if export_format == "pdf":
        pdf_bytes, context = render_trip_report_pdf(trip)
        return pdf_bytes, context, "application/pdf", f"trip_report_{trip.order_number}.pdf"
    return render_trip_report_excel(trip)
