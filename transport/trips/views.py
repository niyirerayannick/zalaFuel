import warnings
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.generic import CreateView, DetailView, ListView, TemplateView, UpdateView
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from accounts.templatetags.currency_tags import currency_filter
from accounts.emailing import approval_recipients, build_public_url, send_atms_email
from accounts.rbac import (
    RBACRequiredMixin,
    SystemGroup,
    can_access_finance,
    can_access_operations,
    can_access_reports,
    can_approve_operations,
    can_manage_operations,
    restrict_queryset_for_user,
    user_has_role,
)
from transport.finance.forms import DriverAllowanceForm, ExpenseForm
from transport.finance.models import DriverAllowance, Expense
from transport.finance.services import (
    approve_allowance,
    generate_and_send_invoices_for_trip,
    generate_invoices_for_trip,
    sync_trip_rental_expense,
)
from transport.drivers.models import Driver
from transport.orders.models import Order
from transport.vehicles.models import Vehicle

from .forms import ShipmentForm, TripForm, TripReportEmailForm
from .loading_order_service import send_loading_order_email
from .models import Shipment, Trip
from .reporting import get_trip_report_export
from .services import (
    TripWorkflowError,
    add_shipment,
    approve_trip,
    complete_trip,
    reject_trip,
    sync_trip_shipments,
    start_trip,
    trip_queryset_for_operations,
)


def _logo_path():
    candidate = Path(__file__).resolve().parents[2] / "static" / "img" / "ZALA/ECO ENERGY.png"
    return candidate if candidate.exists() else None


def _logo_stream(max_width=900):
    logo = _logo_path()
    if not logo:
        return None

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", PILImage.DecompressionBombWarning)
        img = PILImage.open(logo)
        img.load()
    with img:
        img.thumbnail((max_width, max_width))
        stream = BytesIO()
        img.save(stream, format="PNG", optimize=True)
        stream.seek(0)
        return stream


class OperationsAccessMixin(RBACRequiredMixin):
    def test_func(self):
        return can_manage_operations(self.request.user)


def customer_trip_price(trip, customer=None):
    if trip is None:
        return Decimal("0")

    if customer is None:
        return trip.expected_revenue or Decimal("0")

    if getattr(trip, "job_id", None) and getattr(trip, "job", None) and trip.job.customer_id == customer.pk:
        return trip.job.quoted_price or Decimal("0")

    total = Decimal("0")
    shipments = trip.shipments.select_related("order").filter(customer=customer)
    for shipment in shipments:
        order = shipment.order
        order_total_quantity = order.total_quantity_value
        if order_total_quantity <= 0 or not order.quoted_price:
            continue
        total += (order.quoted_price * shipment.quantity) / order_total_quantity
    return total


def restrict_trip_queryset_for_user(queryset, user):
    queryset = restrict_queryset_for_user(queryset, user, "customer")
    if user_has_role(user, SystemGroup.CUSTOMER):
        customer = getattr(user, "customer_profile", None)
        if customer is None:
            return queryset.none()
        return queryset.filter(Q(customer=customer) | Q(shipments__customer=customer)).distinct()
    return queryset


class TripFormResponseMixin:
    modal_template_name = "transport/trips/_modal_form.html"

    def is_partial_form_request(self):
        return self.request.GET.get("partial") == "form" or self.request.headers.get("X-Requested-With") == "XMLHttpRequest"

    def get_template_names(self):
        if self.request.method == "GET" and self.is_partial_form_request():
            return [self.modal_template_name]
        return [self.template_name]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.setdefault("cancel_url", reverse_lazy("transport:trips:list"))
        return context

    def form_invalid(self, form):
        if self.request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {
                    "success": False,
                    "errors": {field: [str(error) for error in errors] for field, errors in form.errors.items()},
                    "non_field_errors": [str(error) for error in form.non_field_errors()],
                },
                status=400,
            )
        return super().form_invalid(form)

    def form_valid(self, form):
        response = super().form_valid(form)
        if self.request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {
                    "success": True,
                    "trip_id": self.object.pk,
                    "detail_url": reverse_lazy("transport:trips:detail", kwargs={"pk": self.object.pk}),
                }
            )
        return response


class TripListView(RBACRequiredMixin, ListView):
    model = Trip
    template_name = "transport/trips/list.html"
    context_object_name = "trips"
    paginate_by = 20

    def get_queryset(self):
        queryset = trip_queryset_for_operations()
        queryset = restrict_trip_queryset_for_user(queryset, self.request.user)

        status = self.request.GET.get("status")
        search = self.request.GET.get("search")
        driver = self.request.GET.get("driver")
        if status:
            queryset = queryset.filter(status=status)
        if search:
            queryset = queryset.filter(order_number__icontains=search)
        if driver:
            queryset = queryset.filter(driver_id=driver)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        base_queryset = self.get_queryset()
        is_customer = user_has_role(self.request.user, SystemGroup.CUSTOMER)
        customer = getattr(self.request.user, "customer_profile", None) if is_customer else None

        if is_customer and customer is not None:
            visible_trips = list(context["trips"])
            customer_total_price = Decimal("0")
            for trip in visible_trips:
                trip.customer_price = customer_trip_price(trip, customer)
                customer_total_price += trip.customer_price
        else:
            visible_trips = list(context["trips"])
            for trip in visible_trips:
                trip.customer_price = trip.expected_revenue or Decimal("0")
            customer_total_price = sum((trip.total_revenue for trip in base_queryset), Decimal("0"))

        coordinator_total_cost = sum((trip.total_expenses for trip in base_queryset), Decimal("0"))

        context.update(
            {
                "total_trips": base_queryset.count(),
                "draft_trips": base_queryset.filter(status__in=[Trip.TripStatus.DRAFT, Trip.TripStatus.PENDING_APPROVAL]).count(),
                "active_trips": base_queryset.filter(status=Trip.TripStatus.IN_TRANSIT).count(),
                "pending_approvals": base_queryset.filter(status=Trip.TripStatus.PENDING_APPROVAL).count(),
                "delivered_trips": base_queryset.filter(
                    status__in=[Trip.TripStatus.DELIVERED, Trip.TripStatus.COMPLETED, Trip.TripStatus.CLOSED]
                ).count(),
                "completed_trips": base_queryset.filter(status=Trip.TripStatus.COMPLETED).count(),
                "total_revenue": customer_total_price,
                "total_cost": coordinator_total_cost,
                "is_customer": is_customer,
                "status_choices": Trip.TripStatus.choices,
                "drivers": Driver.objects.order_by("name"),
            }
        )
        return context

    def test_func(self):
        return user_has_role(self.request.user, SystemGroup.CUSTOMER) or can_access_operations(self.request.user)


class TripListExportMixin(RBACRequiredMixin):
    def test_func(self):
        return user_has_role(self.request.user, SystemGroup.CUSTOMER) or can_access_operations(self.request.user)

    def get_filtered_queryset(self):
        view = TripListView()
        view.request = self.request
        return view.get_queryset()


class TripListExcelExportView(TripListExportMixin, ListView):
    def get(self, request, *args, **kwargs):
        queryset = self.get_filtered_queryset()
        filename = "trips_report.xlsx"
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError:
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="trips_report.csv"'
            response.write("Order,Customer,Route,Vehicle,Driver,Status,Value,Created\n")
            for trip in queryset:
                response.write(
                    f'"{trip.order_number}","{trip.customer.company_name}","{trip.route.origin} -> {trip.route.destination}","{trip.vehicle.plate_number}","{trip.driver.name}","{trip.get_status_display()}","{trip.total_revenue}","{trip.created_at:%Y-%m-%d}"\n'
                )
            return response

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Trips"
        header_fill = PatternFill(fill_type="solid", fgColor="0F5B2A")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"), top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"))
        sheet.merge_cells("A1:H1")
        sheet["A1"] = "ZALA/ECO ENERGY Trips Report"
        sheet["A1"].font = Font(color="0F5B2A", bold=True, size=16)
        sheet.merge_cells("A2:H2")
        sheet["A2"] = f"Generated on {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
        sheet["A2"].font = Font(color="475569", italic=True, size=10)
        headers = ["Order", "Customer", "Route", "Vehicle", "Driver", "Status", "Value", "Created"]
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=4, column=idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        row_no = 5
        for trip in queryset:
            values = [
                trip.order_number,
                trip.customer.company_name,
                f"{trip.route.origin} -> {trip.route.destination}",
                trip.vehicle.plate_number,
                trip.driver.name,
                trip.get_status_display(),
                float(trip.total_revenue or Decimal("0")),
                trip.created_at.strftime("%d/%m/%Y"),
            ]
            for col_no, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_no, column=col_no, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top")
            row_no += 1
        sheet.auto_filter.ref = f"A4:H{max(row_no - 1, 4)}"
        sheet.freeze_panes = "A5"
        for column_cells in sheet.columns:
            column_index = column_cells[0].column
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 3, 30)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type=mime_type)
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class TripListPdfExportView(TripListExportMixin, ListView):
    def get(self, request, *args, **kwargs):
        queryset = self.get_filtered_queryset()
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
        styles = getSampleStyleSheet()
        logo_stream = _logo_stream()
        header_left = []
        if logo_stream:
            header_left.append(Image(logo_stream, width=34 * mm, height=16 * mm))
            header_left.append(Spacer(1, 2 * mm))
        header_left.extend([Paragraph("<font color='#0F5B2A'><b>ZALA/ECO ENERGY Trips Report</b></font>", styles["Title"]), Paragraph("Trips overview export generated from ZALA/ECO ENERGY.", styles["Normal"])])
        header_right = [Paragraph("<b>Report</b><br/>Trips Register", styles["Normal"]), Spacer(1, 2), Paragraph(f"<b>Generated</b><br/>{timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}", styles["Normal"]), Spacer(1, 2), Paragraph(f"<b>Total Trips</b><br/>{queryset.count()}", styles["Normal"])]
        header_table = Table([[header_left, header_right]], colWidths=[170 * mm, 85 * mm])
        header_table.setStyle(TableStyle([("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1E7D7")), ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#EAF7EF")), ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10), ("TOPPADDING", (0, 0), (-1, -1), 10), ("BOTTOMPADDING", (0, 0), (-1, -1), 10), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
        rows = [["Order", "Customer", "Route", "Vehicle", "Driver", "Status", "Value", "Created"]]
        for trip in queryset:
            rows.append([trip.order_number, trip.customer.company_name, f"{trip.route.origin} -> {trip.route.destination}", trip.vehicle.plate_number, trip.driver.name, trip.get_status_display(), currency_filter(trip.total_revenue), trip.created_at.strftime("%d/%m/%Y")])
        if len(rows) == 1:
            rows.append(["-", "-", "-", "-", "-", "-", "-", "-"])
        table = Table(rows, colWidths=[32 * mm, 44 * mm, 54 * mm, 25 * mm, 34 * mm, 24 * mm, 24 * mm, 24 * mm], repeatRows=1)
        table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F5B2A")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 8), ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1D5DB")), ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]), ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6), ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
        doc.build([header_table, Spacer(1, 5 * mm), table])
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="trips_report.pdf"'
        return response


class TripDetailView(RBACRequiredMixin, DetailView):
    model = Trip
    template_name = "transport/trips/detail.html"
    context_object_name = "trip"

    def get_queryset(self):
        return restrict_trip_queryset_for_user(trip_queryset_for_operations(), self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        trip = self.object
        from transport.customers.models import Customer
        from transport.orders.models import Order
        customer = getattr(self.request.user, "customer_profile", None)
        is_customer = user_has_role(self.request.user, SystemGroup.CUSTOMER)

        if is_customer and customer is not None:
            shipment_queryset = trip.shipments.select_related("customer", "order").filter(customer=customer)
            available_orders = Order.objects.none()
            available_customers = Customer.objects.none()
            expenses = trip.expenses.none()
            allowances = trip.allowances.none()
            customer_price = customer_trip_price(trip, customer)
        else:
            shipment_queryset = trip.shipments.select_related("customer", "order")
            available_orders = Order.objects.select_related("customer").order_by("-created_at")[:50]
            available_customers = Customer.objects.order_by("company_name")[:50]
            expenses = trip.expenses.select_related("type").order_by("-created_at")
            allowances = trip.allowances.select_related("driver", "approved_by").order_by("-created_at")
            customer_price = trip.expected_revenue or Decimal("0")

        trip_is_locked = trip.status == Trip.TripStatus.REJECTED
        rental_fee_amount = trip.rental_fee or Decimal("0")
        other_expenses_amount = trip.total_expenses - rental_fee_amount
        if other_expenses_amount < 0:
            other_expenses_amount = Decimal("0")

        context.update(
            {
                "expense_form": ExpenseForm(initial={"trip": trip, "vehicle": trip.vehicle}),
                "allowance_form": DriverAllowanceForm(initial={"trip": trip, "driver": trip.driver}),
                "report_email_form": TripReportEmailForm(),
                "available_orders": available_orders,
                "available_customers": available_customers,
                "shipments": shipment_queryset,
                "expenses": expenses,
                "allowances": allowances,
                "can_approve": can_approve_operations(self.request.user)
                and trip.status == Trip.TripStatus.PENDING_APPROVAL,
                "can_reject": can_approve_operations(self.request.user)
                and trip.status == Trip.TripStatus.PENDING_APPROVAL,
                "can_start": can_manage_operations(self.request.user) and trip.status == Trip.TripStatus.APPROVED,
                "can_complete": can_manage_operations(self.request.user) and trip.status == Trip.TripStatus.IN_TRANSIT,
                "can_edit": (not trip_is_locked)
                and can_manage_operations(self.request.user),
                "can_generate_invoice": can_access_finance(self.request.user)
                and (not trip_is_locked) and (trip.job_id is not None or trip.shipments.exists()),
                "can_send_loading_order": user_has_role(
                    self.request.user,
                    SystemGroup.ADMIN,
                    SystemGroup.OPERATIONS_MANAGER,
                    SystemGroup.LOGISTICS_COORDINATOR,
                ) and (not trip_is_locked) and bool(getattr(trip.customer, "email", "")),
                "can_export_reports": can_access_reports(self.request.user),
                "can_modify_trip": (not trip_is_locked) and can_manage_operations(self.request.user),
                "trip_is_locked": trip_is_locked,
                "financial_rental_fee": rental_fee_amount,
                "financial_other_expenses": other_expenses_amount,
                "financial_total_expenses": trip.total_expenses,
                "invoices": trip.payments.select_related("order").order_by("-created_at"),
                "related_orders": trip.related_orders,
                "show_financials": not is_customer,
                "is_customer": is_customer,
                "customer_price": customer_price,
            }
        )
        return context

    def test_func(self):
        return user_has_role(self.request.user, SystemGroup.CUSTOMER) or can_access_operations(self.request.user)


class TripCreateView(TripFormResponseMixin, OperationsAccessMixin, CreateView):
    model = Trip
    form_class = TripForm
    template_name = "transport/trips/create.html"
    success_url = reverse_lazy("transport:trips:list")

    def form_valid(self, form):
        shipments = form.cleaned_data["shipments"]
        try:
            with transaction.atomic():
                response = super().form_valid(form)
                sync_trip_shipments(trip=self.object, shipments=shipments)
                sync_trip_rental_expense(trip=self.object, created_by=self.request.user)
        except (TripWorkflowError, ValidationError) as exc:
            error_map = getattr(exc, "message_dict", None) or {None: [str(exc)]}
            for field, errors in error_map.items():
                for error in errors:
                    form.add_error(field if field in form.fields else None, error)
            return self.form_invalid(form)
        manager_emails = approval_recipients(SystemGroup.ADMIN, SystemGroup.OPERATIONS_MANAGER)
        if manager_emails:
            try:
                send_atms_email(
                    subject=f"ZALA/ECO ENERGY approval required for trip {self.object.order_number}",
                    to=manager_emails,
                    greeting="Hello Manager",
                    headline="Trip Approval Required",
                    intro="A new trip was created and is waiting for approval in ZALA/ECO ENERGY.",
                    details=[
                        {"label": "Trip Reference", "value": self.object.order_number},
                        {"label": "Driver", "value": self.object.driver.name},
                        {"label": "Vehicle", "value": self.object.vehicle.plate_number},
                        {"label": "Route", "value": f"{self.object.route.origin} to {self.object.route.destination}"},
                        {"label": "Created By", "value": self.request.user.full_name},
                    ],
                    note="Review the trip and approve it if all operational details are correct.",
                    cta_label="Review Trip",
                    cta_url=build_public_url(f"/transport/trips/{self.object.pk}/"),
                )
            except Exception:
                pass
        messages.success(self.request, f"Trip {self.object.order_number} created and sent for approval.")
        return response


class TripUpdateView(TripFormResponseMixin, OperationsAccessMixin, UpdateView):
    model = Trip
    form_class = TripForm
    template_name = "transport/trips/edit.html"

    def form_valid(self, form):
        shipments = form.cleaned_data["shipments"]
        try:
            with transaction.atomic():
                response = super().form_valid(form)
                sync_trip_shipments(trip=self.object, shipments=shipments)
                sync_trip_rental_expense(trip=self.object, created_by=self.request.user)
        except (TripWorkflowError, ValidationError) as exc:
            error_map = getattr(exc, "message_dict", None) or {None: [str(exc)]}
            for field, errors in error_map.items():
                for error in errors:
                    form.add_error(field if field in form.fields else None, error)
            return self.form_invalid(form)
        messages.success(self.request, f"Trip {self.object.order_number} updated successfully.")
        return response

    def get_success_url(self):
        return reverse_lazy("transport:trips:detail", kwargs={"pk": self.object.pk})


class ShipmentListView(OperationsAccessMixin, ListView):
    model = Shipment
    template_name = "transport/trips/shipments.html"
    partial_template_name = "transport/trips/_shipments_list_content.html"
    modal_template_name = "transport/trips/_shipment_modal_form.html"
    context_object_name = "shipments"
    paginate_by = 20

    def test_func(self):
        return can_access_operations(self.request.user)

    def get_template_names(self):
        if self.request.headers.get("X-Requested-With") == "XMLHttpRequest":
            partial = self.request.GET.get("partial")
            if partial == "list":
                return [self.partial_template_name]
            if partial == "form":
                return [self.modal_template_name]
        return [self.template_name]

    def get_queryset(self):
        queryset = Shipment.objects.select_related(
            "order",
            "order__route",
            "order__cargo_category",
            "order__unit",
            "customer",
            "trip",
        ).order_by("-created_at")
        status = self.request.GET.get("status")
        search = self.request.GET.get("search")
        if status:
            queryset = queryset.filter(status=status)
        if search:
            queryset = queryset.filter(
                Q(order__order_number__icontains=search)
                | Q(customer__company_name__icontains=search)
                | Q(order__commodity_description__icontains=search)
                | Q(order__route__origin__icontains=search)
                | Q(order__route__destination__icontains=search)
            )
        return queryset

    def get_context_data(self, **kwargs):
        form = kwargs.get("form") or ShipmentForm()
        order_queryset = Shipment.available_orders_queryset()
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "form": form,
                "editing_shipment": kwargs.get("editing_shipment"),
                "status_choices": Shipment.Status.choices,
                "pending_count": Shipment.objects.filter(status=Shipment.Status.PENDING).count(),
                "assigned_count": Shipment.objects.filter(status=Shipment.Status.ASSIGNED).count(),
                "in_transit_count": Shipment.objects.filter(status=Shipment.Status.IN_TRANSIT).count(),
                "delivered_count": Shipment.objects.filter(status=Shipment.Status.DELIVERED).count(),
                "available_orders": order_queryset,
                "can_create_shipments": can_manage_operations(self.request.user),
                "orders_payload": [
                    {
                        "id": str(order.pk),
                        "order_number": order.order_number,
                        "customer": order.customer.company_name,
                        "cargo_category": getattr(order.cargo_category, "name", "Uncategorised"),
                        "commodity_type": order.get_commodity_type_display(),
                        "route": f"{order.origin} -> {order.destination}" if order.origin or order.destination else "Route not set",
                        "total_quantity": str(order.total_quantity or 0),
                        "remaining_quantity": str(order.remaining_quantity or 0),
                        "display_total_quantity": order.display_quantity,
                        "display_remaining_quantity": order.formatted_remaining_weight_kg,
                        "display_business_quantity": f"{order.formatted_total_quantity} {order.quantity_unit_symbol}".strip(),
                        "weight_kg": str(order.weight_kg or 0),
                        "unit_symbol": order.quantity_unit_symbol,
                        "carriage_type_choices": [
                            {"value": value, "label": label} for value, label in Shipment.CarriageType.choices
                        ],
                        "pickup_address": order.pickup_address or "",
                        "delivery_address": order.delivery_address or "",
                    }
                    for order in order_queryset
                ],
            }
        )
        return context

    def post(self, request, *args, **kwargs):
        if not can_manage_operations(request.user):
            messages.error(request, "You do not have permission to create shipments.")
            return redirect("transport:trips:shipment-list")
        form = ShipmentForm(request.POST)
        if form.is_valid():
            shipment = form.save()
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse(
                    {
                        "success": True,
                        "shipment_id": shipment.pk,
                    }
                )
            messages.success(request, f"Shipment prepared for order {shipment.order.order_number}.")
            return redirect("transport:trips:shipment-list")
        self.object_list = self.get_queryset()
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {
                    "success": False,
                    "errors": {field: [str(error) for error in errors] for field, errors in form.errors.items()},
                    "non_field_errors": [str(error) for error in form.non_field_errors()],
                },
                status=400,
            )
        return self.render_to_response(self.get_context_data(form=form))


class ShipmentExportMixin(OperationsAccessMixin):
    def get_filtered_queryset(self):
        view = ShipmentListView()
        view.request = self.request
        return view.get_queryset()


class ShipmentExcelExportView(ShipmentExportMixin, ListView):
    def get(self, request, *args, **kwargs):
        queryset = self.get_filtered_queryset()
        filename = "shipments_report.xlsx"
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError:
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="shipments_report.csv"'
            response.write("Shipment,Order,Cargo,Route,Customer,Shipment Weight,Remaining Weight,Status,Trip\n")
            for shipment in queryset:
                response.write(
                    f'"SHP-{shipment.pk}","{shipment.order.order_number}","{getattr(shipment.order.cargo_category, "name", "Uncategorised")}","{shipment.order.origin} -> {shipment.order.destination}","{shipment.customer.company_name}","{shipment.weight_kg} kg","{shipment.order.formatted_remaining_weight_kg}","{shipment.get_status_display()}","{shipment.trip.order_number if shipment.trip else "Not assigned"}"\n'
                )
            return response

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Shipments"
        header_fill = PatternFill(fill_type="solid", fgColor="0F5B2A")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"), top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"))
        sheet.merge_cells("A1:I1")
        sheet["A1"] = "ZALA/ECO ENERGY Shipments Report"
        sheet["A1"].font = Font(color="0F5B2A", bold=True, size=16)
        sheet.merge_cells("A2:I2")
        sheet["A2"] = f"Generated on {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
        sheet["A2"].font = Font(color="475569", italic=True, size=10)
        headers = ["Shipment", "Order", "Cargo", "Route", "Customer", "Shipment Weight", "Remaining Weight", "Status", "Trip"]
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=4, column=idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        row_no = 5
        for shipment in queryset:
            values = [
                f"SHP-{shipment.pk}",
                shipment.order.order_number,
                getattr(shipment.order.cargo_category, "name", "Uncategorised"),
                f"{shipment.order.origin} -> {shipment.order.destination}",
                shipment.customer.company_name,
                f"{shipment.weight_kg:.2f} kg",
                shipment.order.formatted_remaining_weight_kg,
                shipment.get_status_display(),
                shipment.trip.order_number if shipment.trip else "Not assigned",
            ]
            for col_no, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_no, column=col_no, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top")
            row_no += 1
        sheet.auto_filter.ref = f"A4:I{max(row_no - 1, 4)}"
        sheet.freeze_panes = "A5"
        for column_cells in sheet.columns:
            column_index = column_cells[0].column
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 3, 30)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type=mime_type)
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class ShipmentPdfExportView(ShipmentExportMixin, ListView):
    def get(self, request, *args, **kwargs):
        queryset = self.get_filtered_queryset()
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
        styles = getSampleStyleSheet()
        logo_stream = _logo_stream()
        header_left = []
        if logo_stream:
            header_left.append(Image(logo_stream, width=34 * mm, height=16 * mm))
            header_left.append(Spacer(1, 2 * mm))
        header_left.extend([Paragraph("<font color='#0F5B2A'><b>ZALA/ECO ENERGY Shipments Report</b></font>", styles["Title"]), Paragraph("Shipments overview export generated from ZALA/ECO ENERGY.", styles["Normal"])])
        header_right = [Paragraph("<b>Report</b><br/>Shipments Register", styles["Normal"]), Spacer(1, 2), Paragraph(f"<b>Generated</b><br/>{timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}", styles["Normal"]), Spacer(1, 2), Paragraph(f"<b>Total Shipments</b><br/>{queryset.count()}", styles["Normal"])]
        header_table = Table([[header_left, header_right]], colWidths=[170 * mm, 85 * mm])
        header_table.setStyle(TableStyle([("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1E7D7")), ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#EAF7EF")), ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10), ("TOPPADDING", (0, 0), (-1, -1), 10), ("BOTTOMPADDING", (0, 0), (-1, -1), 10), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
        rows = [["Shipment", "Order", "Cargo", "Route", "Customer", "Shipment Weight", "Remaining Weight", "Status", "Trip"]]
        for shipment in queryset:
            rows.append([f"SHP-{shipment.pk}", shipment.order.order_number, getattr(shipment.order.cargo_category, "name", "Uncategorised"), f"{shipment.order.origin} -> {shipment.order.destination}", shipment.customer.company_name, f"{shipment.weight_kg:.2f} kg", shipment.order.formatted_remaining_weight_kg, shipment.get_status_display(), shipment.trip.order_number if shipment.trip else "Not assigned"])
        if len(rows) == 1:
            rows.append(["-", "-", "-", "-", "-", "-", "-", "-", "-"])
        table = Table(rows, colWidths=[22 * mm, 28 * mm, 28 * mm, 42 * mm, 35 * mm, 26 * mm, 26 * mm, 23 * mm, 25 * mm], repeatRows=1)
        table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F5B2A")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 8), ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1D5DB")), ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]), ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6), ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
        doc.build([header_table, Spacer(1, 5 * mm), table])
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="shipments_report.pdf"'
        return response


class ShipmentUpdateView(OperationsAccessMixin, UpdateView):
    model = Shipment
    form_class = ShipmentForm
    template_name = "transport/trips/shipments.html"
    partial_template_name = "transport/trips/_shipment_modal_form.html"
    allowed_roles = (SystemGroup.ADMIN, SystemGroup.OPERATIONS_MANAGER, SystemGroup.LOGISTICS_COORDINATOR)

    def get_template_names(self):
        if self.request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return [self.partial_template_name]
        return [self.template_name]

    def get_queryset(self):
        return Shipment.objects.select_related(
            "order",
            "order__route",
            "order__cargo_category",
            "order__unit",
            "customer",
            "trip",
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["editing_shipment"] = self.object
        return context

    def form_valid(self, form):
        shipment = form.save()
        if self.request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": True, "shipment_id": shipment.pk})
        messages.success(self.request, f"Shipment {shipment.pk} updated successfully.")
        return redirect("transport:trips:shipment-list")

    def form_invalid(self, form):
        if self.request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {
                    "success": False,
                    "errors": {field: [str(error) for error in errors] for field, errors in form.errors.items()},
                    "non_field_errors": [str(error) for error in form.non_field_errors()],
                },
                status=400,
            )
        return self.render_to_response(self.get_context_data(form=form))


class OperationsDashboardView(OperationsAccessMixin, TemplateView):
    template_name = "transport/trips/operations_dashboard.html"
    allowed_roles = (SystemGroup.ADMIN, SystemGroup.OPERATIONS_MANAGER)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = trip_queryset_for_operations()
        context.update(
            {
                "active_trips": queryset.filter(status=Trip.TripStatus.IN_TRANSIT)[:10],
                "groupage_trips": queryset.filter(shipments__isnull=False).distinct()[:10],
                "pending_approvals": queryset.filter(status=Trip.TripStatus.PENDING_APPROVAL)[:10],
                "allowance_requests": DriverAllowance.objects.filter(status=DriverAllowance.Status.PENDING).select_related("driver", "trip")[:10],
                "available_vehicles": Vehicle.objects.filter(status=Vehicle.VehicleStatus.AVAILABLE)[:10],
                "drivers_in_transit": Driver.objects.filter(availability_status=Driver.AvailabilityStatus.IN_TRANSIT)[:10],
            }
        )
        return context


@login_required
@require_POST
def update_trip_status(request, trip_id):
    trip = get_object_or_404(trip_queryset_for_operations(), pk=trip_id)
    action = request.POST.get("action")
    try:
        if action == "approve" and can_approve_operations(request.user):
            approve_trip(trip, request.user)
            messages.success(request, "Trip approved and invoice prepared.")
        elif action == "reject" and can_approve_operations(request.user):
            reject_trip(trip)
            messages.success(request, "Trip rejected.")
        elif action == "start" and can_manage_operations(request.user):
            start_trip(trip)
            messages.success(request, "Trip is now in transit.")
        elif action == "complete" and can_manage_operations(request.user):
            complete_trip(trip)
            messages.success(request, "Trip marked as completed.")
        else:
            messages.error(request, "You do not have permission for that action.")
    except (TripWorkflowError, ValidationError) as exc:
        error_message = getattr(exc, "message", None)
        if not error_message and hasattr(exc, "message_dict"):
            error_message = "; ".join(
                f"{field}: {' '.join(messages_list)}" for field, messages_list in exc.message_dict.items()
            )
        messages.error(request, error_message or "Invalid trip workflow transition.")
    return redirect("transport:trips:detail", pk=trip.pk)


@login_required
@require_POST
def generate_trip_invoice(request, trip_id):
    trip = get_object_or_404(trip_queryset_for_operations(), pk=trip_id)
    if not can_access_finance(request.user):
        messages.error(request, "You do not have permission to generate invoices.")
        return redirect("transport:trips:detail", pk=trip.pk)

    try:
        invoices = generate_and_send_invoices_for_trip(trip, created_by=request.user)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("transport:trips:detail", pk=trip.pk)
    except Exception as exc:
        messages.error(request, f"Invoice generation completed, but email sending failed: {exc}")
        return redirect("transport:trips:detail", pk=trip.pk)

    if not invoices:
        messages.error(request, "This trip does not have any linked orders to invoice.")
    elif len(invoices) == 1:
        messages.success(
            request,
            f'Invoice {invoices[0].reference or invoices[0].pk} was generated and sent to {trip.customer.email}.',
        )
    else:
        messages.success(
            request,
            f"{len(invoices)} invoices were generated and sent to customers for this trip.",
        )
    return redirect("transport:trips:detail", pk=trip.pk)


@login_required
def export_trip_report_pdf(request, trip_id):
    trip = get_object_or_404(trip_queryset_for_operations(), pk=trip_id)
    trip = get_object_or_404(restrict_trip_queryset_for_user(Trip.objects.filter(pk=trip.pk), request.user), pk=trip.pk)
    if not can_access_reports(request.user):
        messages.error(request, "You do not have permission to export trip reports.")
        return redirect("transport:trips:detail", pk=trip.pk)

    file_bytes, _context, mime_type, filename = get_trip_report_export(trip, "pdf")
    response = HttpResponse(file_bytes, content_type=mime_type)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_trip_report_excel(request, trip_id):
    trip = get_object_or_404(trip_queryset_for_operations(), pk=trip_id)
    trip = get_object_or_404(restrict_trip_queryset_for_user(Trip.objects.filter(pk=trip.pk), request.user), pk=trip.pk)
    if not can_access_reports(request.user):
        messages.error(request, "You do not have permission to export trip reports.")
        return redirect("transport:trips:detail", pk=trip.pk)

    file_bytes, _context, mime_type, filename = get_trip_report_export(trip, "excel")
    response = HttpResponse(file_bytes, content_type=mime_type)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_POST
def email_trip_report(request, trip_id):
    trip = get_object_or_404(trip_queryset_for_operations(), pk=trip_id)
    trip = get_object_or_404(restrict_trip_queryset_for_user(Trip.objects.filter(pk=trip.pk), request.user), pk=trip.pk)
    if not can_access_reports(request.user):
        messages.error(request, "You do not have permission to share trip reports.")
        return redirect("transport:trips:detail", pk=trip.pk)

    form = TripReportEmailForm(request.POST)
    if not form.is_valid():
        error_text = " ".join(str(error) for error_list in form.errors.values() for error in error_list)
        messages.error(request, error_text or "Please select recipients and a valid report format.")
        return redirect("transport:trips:detail", pk=trip.pk)

    report_format = form.cleaned_data["report_format"]
    recipients = form.cleaned_data["recipients"]
    recipient_emails = [user.email for user in recipients if user.email]
    file_bytes, _context, mime_type, filename = get_trip_report_export(trip, report_format)

    try:
        send_atms_email(
            subject=f"Trip Report {trip.order_number}",
            to=recipient_emails,
            greeting="Hello Team",
            headline="Trip Report Shared",
            intro="A trip report has been shared from ZALA/ECO ENERGY.",
            details=[
                {"label": "Trip Reference", "value": trip.order_number},
                {"label": "Customer", "value": trip.customer.company_name},
                {"label": "Route", "value": f"{trip.route.origin} to {trip.route.destination}"},
                {"label": "Report Format", "value": report_format.upper()},
            ],
            note="Please find the attached trip report for operational review.",
            cta_label="Open Trip",
            cta_url=build_public_url(f"/transport/trips/{trip.pk}/"),
            attachments=[(filename, file_bytes, mime_type)],
        )
    except Exception as exc:
        messages.error(request, f"Trip report email could not be sent: {exc}")
        return redirect("transport:trips:detail", pk=trip.pk)

    messages.success(request, f"Trip report sent to {len(recipient_emails)} recipient(s).")
    return redirect("transport:trips:detail", pk=trip.pk)


@login_required
@require_POST
def send_trip_loading_order(request, trip_id):
    trip = get_object_or_404(trip_queryset_for_operations(), pk=trip_id)
    if not user_has_role(
        request.user,
        SystemGroup.ADMIN,
        SystemGroup.OPERATIONS_MANAGER,
        SystemGroup.LOGISTICS_COORDINATOR,
    ):
        messages.error(request, "You do not have permission to send loading orders.")
        return redirect("transport:trips:detail", pk=trip.pk)

    try:
        document, _context = send_loading_order_email(trip, created_by=request.user)
    except ValueError as exc:
        messages.error(request, str(exc))
    except Exception as exc:
        messages.error(request, f"Loading order email could not be sent: {exc}")
    else:
        messages.success(
            request,
            f'Loading order "{document.name}" was generated and sent to {trip.customer.email}.',
        )
    return redirect("transport:trips:detail", pk=trip.pk)


@login_required
@require_POST
def add_trip_expense(request, trip_id):
    trip = get_object_or_404(trip_queryset_for_operations(), pk=trip_id)
    if not can_manage_operations(request.user):
        return JsonResponse({"success": False, "error": "Permission denied."}, status=403)

    form = ExpenseForm(request.POST)
    if form.is_valid():
        expense = form.save(commit=False)
        expense.trip = trip
        expense.vehicle = trip.vehicle
        expense.created_by = request.user
        expense.save()
        rental_fee_amount = trip.rental_fee or Decimal("0")
        other_expenses_amount = trip.total_expenses - rental_fee_amount
        if other_expenses_amount < 0:
            other_expenses_amount = Decimal("0")
        return JsonResponse(
            {
                "success": True,
                "expense": {
                    "id": expense.pk,
                    "type": expense.type.name if expense.type else expense.category,
                    "amount": currency_filter(expense.amount),
                    "liters_raw": str(expense.liters) if expense.liters is not None else "",
                    "fuel_unit_price_raw": str(expense.fuel_unit_price) if expense.fuel_unit_price is not None else "",
                    "description": expense.description,
                },
                "totals": {
                    "other_expenses": currency_filter(other_expenses_amount),
                    "expenses": currency_filter(trip.total_expenses),
                    "net_profit": currency_filter(trip.net_profit),
                    "cost_per_km": str(trip.cost_per_km),
                },
            }
        )
    return JsonResponse({"success": False, "errors": form.errors}, status=400)


@login_required
@require_POST
def edit_trip_expense(request, trip_id, expense_id):
    trip = get_object_or_404(trip_queryset_for_operations(), pk=trip_id)
    expense = get_object_or_404(Expense.objects.select_related("type", "trip"), pk=expense_id, trip=trip)
    if trip.status == Trip.TripStatus.REJECTED:
        return JsonResponse({"success": False, "error": "Rejected trips are read-only."}, status=400)
    if not can_manage_operations(request.user):
        return JsonResponse({"success": False, "error": "Permission denied."}, status=403)

    form = ExpenseForm(request.POST, request.FILES, instance=expense)
    if form.is_valid():
        expense = form.save(commit=False)
        expense.trip = trip
        expense.vehicle = trip.vehicle
        expense.save()
        rental_fee_amount = trip.rental_fee or Decimal("0")
        other_expenses_amount = trip.total_expenses - rental_fee_amount
        if other_expenses_amount < 0:
            other_expenses_amount = Decimal("0")
        return JsonResponse(
            {
                "success": True,
                "expense": {
                    "id": expense.pk,
                    "type": expense.type.name if expense.type else expense.category,
                    "type_id": expense.type_id,
                    "amount": currency_filter(expense.amount),
                    "amount_raw": str(expense.amount),
                    "liters_raw": str(expense.liters) if expense.liters is not None else "",
                    "fuel_unit_price_raw": str(expense.fuel_unit_price) if expense.fuel_unit_price is not None else "",
                    "description": expense.description,
                },
                "totals": {
                    "other_expenses": currency_filter(other_expenses_amount),
                    "expenses": currency_filter(trip.total_expenses),
                    "net_profit": currency_filter(trip.net_profit),
                    "cost_per_km": str(trip.cost_per_km),
                },
            }
        )
    return JsonResponse(
        {
            "success": False,
            "errors": {field: [str(error) for error in errors] for field, errors in form.errors.items()},
        },
        status=400,
    )


@login_required
@require_POST
def delete_trip_expense(request, trip_id, expense_id):
    trip = get_object_or_404(trip_queryset_for_operations(), pk=trip_id)
    expense = get_object_or_404(Expense.objects.select_related("trip"), pk=expense_id, trip=trip)
    if trip.status == Trip.TripStatus.REJECTED:
        return JsonResponse({"success": False, "error": "Rejected trips are read-only."}, status=400)
    if not can_manage_operations(request.user):
        return JsonResponse({"success": False, "error": "Permission denied."}, status=403)

    expense.delete()
    rental_fee_amount = trip.rental_fee or Decimal("0")
    other_expenses_amount = trip.total_expenses - rental_fee_amount
    if other_expenses_amount < 0:
        other_expenses_amount = Decimal("0")
    return JsonResponse(
        {
            "success": True,
            "totals": {
                "other_expenses": currency_filter(other_expenses_amount),
                "expenses": currency_filter(trip.total_expenses),
                "net_profit": currency_filter(trip.net_profit),
                "cost_per_km": str(trip.cost_per_km),
            },
        }
    )


@login_required
@require_POST
def add_trip_shipment(request, trip_id):
    trip = get_object_or_404(trip_queryset_for_operations(), pk=trip_id)
    if not can_manage_operations(request.user):
        return JsonResponse({"success": False, "error": "Permission denied."}, status=403)

    try:
        quantity = Decimal(request.POST.get("weight_kg", "0"))
    except InvalidOperation:
        return JsonResponse({"success": False, "error": "Weight must be a valid number."}, status=400)
    carriage_type = request.POST.get("carriage_type", Shipment.CarriageType.OTHER)
    container_number = (request.POST.get("container_number") or "").strip()

    from transport.orders.models import Order
    from transport.customers.models import Customer

    order = get_object_or_404(Order, pk=request.POST.get("order"))
    customer = get_object_or_404(Customer, pk=request.POST.get("customer"))
    if customer.pk != order.customer_id:
        return JsonResponse({"success": False, "error": "Selected customer must match the order customer."}, status=400)
    try:
        shipment = add_shipment(
            trip=trip,
            order=order,
            customer=customer,
            quantity=quantity,
            carriage_type=carriage_type,
            container_number=container_number,
        )
    except Exception as exc:
        return JsonResponse({"success": False, "error": str(exc)}, status=400)
    return JsonResponse(
        {
            "success": True,
            "shipment": {
                "id": shipment.pk,
                "order": shipment.order.order_number,
                "customer": shipment.customer.company_name,
                "quantity": str(shipment.quantity),
                "weight_kg": str(shipment.weight_kg),
                "carriage_type": shipment.get_carriage_type_display(),
                "container_number": shipment.container_number,
                "sender_name": shipment.sender_name,
            },
            "totals": {"load": str(trip.total_load), "capacity": str(trip.vehicle.load_capacity)},
        }
    )


@login_required
@require_POST
def request_allowance(request, trip_id):
    trip = get_object_or_404(trip_queryset_for_operations(), pk=trip_id)
    if not can_manage_operations(request.user):
        return JsonResponse({"success": False, "error": "Permission denied."}, status=403)
    form = DriverAllowanceForm(request.POST)
    if form.is_valid():
        allowance = form.save(commit=False)
        allowance.trip = trip
        allowance.created_by = request.user
        allowance.save()
        return JsonResponse({"success": True, "allowance_id": allowance.pk, "status": allowance.status})
    return JsonResponse({"success": False, "errors": form.errors}, status=400)


@login_required
@require_POST
def approve_trip_allowance(request, allowance_id):
    allowance = get_object_or_404(DriverAllowance.objects.select_related("trip", "driver"), pk=allowance_id)
    if not user_has_role(request.user, SystemGroup.ADMIN, SystemGroup.OPERATIONS_MANAGER):
        return JsonResponse({"success": False, "error": "Permission denied."}, status=403)
    approve_allowance(allowance, request.user)
    return JsonResponse({"success": True, "status": allowance.status, "trip_expenses": str(allowance.trip.total_expenses)})
