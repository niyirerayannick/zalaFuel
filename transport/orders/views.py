import warnings
from io import BytesIO
from pathlib import Path

from django.contrib import messages
from django.db.models import Prefetch, Q
from django.http import HttpResponse, JsonResponse
from django.urls import reverse_lazy
from django.views.generic import CreateView, DetailView, ListView, UpdateView
from django.utils import timezone
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from accounts.rbac import (
    RBACRequiredMixin,
    SystemGroup,
    can_access_operations,
    can_create_orders,
    can_edit_orders,
    restrict_queryset_for_user,
    user_has_role,
)
from .forms import OrderForm
from .models import Order
from .services import render_order_pdf
from transport.trips.models import Shipment, Trip


def _logo_path():
    candidate = Path(__file__).resolve().parents[2] / "static" / "img" / "ZALA/ECO ENERGY.png"
    return candidate if candidate.exists() else None


def _logo_stream(max_width=900):
    logo = _logo_path()
    if not logo:
        return None

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", PILImage.DecompressionBombWarning)
        img = PILImage.open(logo)
        img.load()
    with img:
        img.thumbnail((max_width, max_width))
        stream = BytesIO()
        img.save(stream, format="PNG", optimize=True)
        stream.seek(0)
        return stream


class OrderCreateAccessMixin(RBACRequiredMixin):
    def test_func(self):
        return can_create_orders(self.request.user)


class OrderEditAccessMixin(RBACRequiredMixin):
    def test_func(self):
        return can_edit_orders(self.request.user)


class OrderFormResponseMixin:
    modal_template_name = "transport/orders/_modal_form.html"

    def is_partial_form_request(self):
        return self.request.GET.get("partial") == "form" or self.request.headers.get("X-Requested-With") == "XMLHttpRequest"

    def get_template_names(self):
        if self.request.method == "GET" and self.is_partial_form_request():
            return [self.modal_template_name]
        return [self.template_name]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        is_update = getattr(self, "object", None) is not None and getattr(self.object, "pk", None) is not None
        context.setdefault("page_title", "Edit Order" if is_update else "Create Order")
        context.setdefault("page_heading", "Edit Order" if is_update else "Create Order")
        context.setdefault("page_intro", "Update the job details while preserving the current workflow." if is_update else "Capture the job details before preparing shipments and assigning trips.")
        context.setdefault("submit_label", "Update Order" if is_update else "Save Order")
        return context

    def form_invalid(self, form):
        if self.request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {
                    "success": False,
                    "errors": {field: [str(error) for error in errors] for field, errors in form.errors.items()},
                    "non_field_errors": [str(error) for error in form.non_field_errors()],
                },
                status=400,
            )
        return super().form_invalid(form)

    def form_valid(self, form):
        response = super().form_valid(form)
        if self.request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {
                    "success": True,
                    "order_id": str(self.object.pk),
                    "detail_url": reverse_lazy("transport:orders:detail", kwargs={"pk": self.object.pk}),
                }
            )
        return response


class OrderListView(RBACRequiredMixin, ListView):
    model = Order
    template_name = "transport/orders/list.html"
    partial_template_name = "transport/orders/_list_content.html"
    context_object_name = "orders"
    paginate_by = 20

    def get_template_names(self):
        if self.request.headers.get("X-Requested-With") == "XMLHttpRequest" and self.request.GET.get("partial") == "list":
            return [self.partial_template_name]
        return [self.template_name]

    def get_queryset(self):
        queryset = Order.objects.select_related("customer", "route", "cargo_category", "unit").order_by("-created_at")
        queryset = restrict_queryset_for_user(queryset, self.request.user, "customer")
        status = self.request.GET.get("status")
        search = self.request.GET.get("search")
        if status:
            queryset = queryset.filter(status=status)
        if search:
            queryset = queryset.filter(
                Q(order_number__icontains=search)
                | Q(customer__company_name__icontains=search)
                | Q(commodity_description__icontains=search)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        base_queryset = self.get_queryset()
        context["can_create_orders"] = can_create_orders(self.request.user)
        context["can_edit_orders"] = can_edit_orders(self.request.user)
        context.update(
            {
                "total_orders": base_queryset.count(),
                "draft_orders": base_queryset.filter(status=Order.Status.DRAFT).count(),
                "pending_orders": base_queryset.filter(status=Order.Status.PENDING_APPROVAL).count(),
                "active_orders": base_queryset.filter(status__in=[Order.Status.ASSIGNED, Order.Status.IN_TRANSIT]).count(),
                "completed_orders": base_queryset.filter(
                    status__in=[Order.Status.DELIVERED, Order.Status.COMPLETED]
                ).count(),
                "status_choices": Order.Status.choices,
            }
        )
        return context

    def test_func(self):
        return user_has_role(self.request.user, SystemGroup.CUSTOMER) or can_access_operations(self.request.user)


class OrderExportMixin(RBACRequiredMixin):
    def test_func(self):
        return user_has_role(self.request.user, SystemGroup.CUSTOMER) or can_access_operations(self.request.user)

    def get_filtered_queryset(self):
        view = OrderListView()
        view.request = self.request
        return view.get_queryset()


class OrderExcelExportView(OrderExportMixin, ListView):
    def get(self, request, *args, **kwargs):
        queryset = self.get_filtered_queryset()
        filename = "orders_report.xlsx"
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError:
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="orders_report.csv"'
            response.write("Order Number,Customer,Route,Cargo,Status,Quantity,Created\n")
            for order in queryset:
                response.write(
                    f'"{order.order_number}","{order.customer.company_name}","{order.origin} -> {order.destination}","{order.get_commodity_type_display()}","{order.get_status_display()}","{order.display_quantity}","{order.created_at:%Y-%m-%d}"\n'
                )
            return response

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Orders"
        header_fill = PatternFill(fill_type="solid", fgColor="0F5B2A")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"), top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"))
        sheet.merge_cells("A1:G1")
        sheet["A1"] = "ZALA/ECO ENERGY Orders Report"
        sheet["A1"].font = Font(color="0F5B2A", bold=True, size=16)
        sheet.merge_cells("A2:G2")
        sheet["A2"] = f"Generated on {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
        sheet["A2"].font = Font(color="475569", italic=True, size=10)
        headers = ["Order Number", "Customer", "Route", "Cargo", "Status", "Quantity", "Created"]
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=4, column=idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        row_no = 5
        for order in queryset:
            values = [order.order_number, order.customer.company_name, f"{order.origin} -> {order.destination}", order.get_commodity_type_display(), order.get_status_display(), order.display_quantity, order.created_at.strftime("%d/%m/%Y")]
            for col_no, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_no, column=col_no, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top")
            row_no += 1
        sheet.auto_filter.ref = f"A4:G{max(row_no - 1, 4)}"
        sheet.freeze_panes = "A5"
        for column_cells in sheet.columns:
            column_index = column_cells[0].column
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 3, 28)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type=mime_type)
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class OrderPdfExportView(OrderExportMixin, ListView):
    def get(self, request, *args, **kwargs):
        queryset = self.get_filtered_queryset()
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
        styles = getSampleStyleSheet()
        logo_stream = _logo_stream()
        header_left = []
        if logo_stream:
            header_left.append(Image(logo_stream, width=34 * mm, height=16 * mm))
            header_left.append(Spacer(1, 2 * mm))
        header_left.extend([Paragraph("<font color='#0F5B2A'><b>ZALA/ECO ENERGY Orders Report</b></font>", styles["Title"]), Paragraph("Order register export generated from ZALA/ECO ENERGY.", styles["Normal"])])
        header_right = [Paragraph("<b>Report</b><br/>Orders Register", styles["Normal"]), Spacer(1, 2), Paragraph(f"<b>Generated</b><br/>{timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}", styles["Normal"]), Spacer(1, 2), Paragraph(f"<b>Total Orders</b><br/>{queryset.count()}", styles["Normal"])]
        header_table = Table([[header_left, header_right]], colWidths=[170 * mm, 85 * mm])
        header_table.setStyle(TableStyle([("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1E7D7")), ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#EAF7EF")), ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10), ("TOPPADDING", (0, 0), (-1, -1), 10), ("BOTTOMPADDING", (0, 0), (-1, -1), 10), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
        rows = [["Order Number", "Customer", "Route", "Cargo", "Status", "Quantity", "Created"]]
        for order in queryset:
            rows.append([order.order_number, order.customer.company_name, f"{order.origin} -> {order.destination}", order.get_commodity_type_display(), order.get_status_display(), order.display_quantity, order.created_at.strftime("%d/%m/%Y")])
        if len(rows) == 1:
            rows.append(["-", "-", "-", "-", "-", "-", "-"])
        table = Table(rows, colWidths=[32 * mm, 48 * mm, 58 * mm, 32 * mm, 24 * mm, 24 * mm, 24 * mm], repeatRows=1)
        table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F5B2A")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 8), ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1D5DB")), ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]), ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6), ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
        doc.build([header_table, Spacer(1, 5 * mm), table])
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="orders_report.pdf"'
        return response


class OrderCreateView(OrderFormResponseMixin, OrderCreateAccessMixin, CreateView):
    model = Order
    form_class = OrderForm
    template_name = "transport/orders/create.html"
    success_url = reverse_lazy("transport:orders:customer-list")

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        if user_has_role(self.request.user, SystemGroup.CUSTOMER):
            customer = getattr(self.request.user, "customer_profile", None)
            if customer is not None:
                form.fields["customer"].queryset = form.fields["customer"].queryset.filter(pk=customer.pk)
                form.initial["customer"] = customer.pk
        return form

    def form_valid(self, form):
        if user_has_role(self.request.user, SystemGroup.CUSTOMER):
            customer = getattr(self.request.user, "customer_profile", None)
            if customer is None:
                form.add_error("customer", "No customer profile is linked to this account.")
                return self.form_invalid(form)
            form.instance.customer = customer
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        if form.instance.status == Order.Status.DRAFT:
            form.instance.status = Order.Status.PENDING_APPROVAL
        response = super().form_valid(form)
        messages.success(self.request, f"Order {self.object.order_number} created successfully.")
        return response


class OrderUpdateView(OrderFormResponseMixin, OrderEditAccessMixin, UpdateView):
    model = Order
    form_class = OrderForm
    template_name = "transport/orders/create.html"

    def get_success_url(self):
        return reverse_lazy("transport:orders:detail", kwargs={"pk": self.object.pk})

    def get_queryset(self):
        queryset = Order.objects.select_related("customer", "route", "cargo_category", "unit")
        return restrict_queryset_for_user(queryset, self.request.user, "customer")

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        messages.success(self.request, f"Order {self.object.order_number} updated successfully.")
        return response


class OrderDetailView(RBACRequiredMixin, DetailView):
    model = Order
    template_name = "transport/orders/detail.html"
    context_object_name = "order"

    def get_queryset(self):
        queryset = Order.objects.select_related("customer", "route", "cargo_category", "unit")
        if user_has_role(self.request.user, SystemGroup.CUSTOMER):
            customer = getattr(self.request.user, "customer_profile", None)
            trip_queryset = Trip.objects.filter(Q(customer=customer) | Q(shipments__customer=customer)).distinct()
            shipment_queryset = Shipment.objects.filter(customer=customer).select_related("trip", "customer", "order")
            queryset = queryset.prefetch_related(
                Prefetch("trip_records", queryset=trip_queryset),
                Prefetch("shipments", queryset=shipment_queryset),
                "payments",
            )
        else:
            queryset = queryset.prefetch_related("trip_records", "shipments", "payments")
        return restrict_queryset_for_user(queryset, self.request.user, "customer")

    def test_func(self):
        return user_has_role(self.request.user, SystemGroup.CUSTOMER) or can_access_operations(self.request.user)


class OrderPdfView(RBACRequiredMixin, DetailView):
    model = Order

    def get_queryset(self):
        queryset = Order.objects.select_related("customer", "route", "cargo_category", "unit")
        return restrict_queryset_for_user(queryset, self.request.user, "customer")

    def render_to_response(self, context, **response_kwargs):
        order = context["object"]
        pdf_bytes, _pdf_context = render_order_pdf(order)
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{order.order_number}.pdf"'
        return response

    def test_func(self):
        return user_has_role(self.request.user, SystemGroup.CUSTOMER) or can_access_operations(self.request.user)
